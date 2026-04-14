import sys
import uuid
import sqlite3
import os
import hashlib
import copy
import json
from dataclasses import dataclass, field
from pypdf import PdfReader, PdfWriter
from pypdf.annotations import FreeText as PdfFreeTextAnnotation
from pypdf.annotations import Text as PdfTextAnnotation
from pypdf.generic import NameObject, DictionaryObject, NumberObject
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QPushButton,
    QLabel, QFileDialog, QVBoxLayout,
    QWidget, QHBoxLayout, QMenu, QDialog,
    QDialogButtonBox, QFormLayout, QComboBox,
    QInputDialog, QLineEdit, QListWidget,
    QListWidgetItem, QAbstractItemView, QMessageBox,
    QTableWidget, QTableWidgetItem, QHeaderView,
    QSplitter, QGroupBox, QTreeWidget, QTreeWidgetItem,
    QPlainTextEdit, QStackedWidget, QTabWidget, QSizePolicy,
    QDockWidget, QTextEdit, QScrollArea, QSpinBox, QStyle
)
from PySide6.QtCore import Qt, Signal, QPointF, QRectF, QTimer
from PySide6.QtGui import (QPainter, QColor, QPen, QPainterPath, QBrush,
                            QAction, QKeySequence, QActionGroup, QFont, QPalette)
from PySide6.QtPdf import QPdfDocument
from PySide6.QtPdfWidgets import QPdfView

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
ZOOM_STEP    = 1.25
ZOOM_MIN     = 0.1
ZOOM_MAX     = 10.0
PAGE_SPACING = 4

MARKER_W  = 22
MARKER_H  = 22
MARKER_TH = 7

# ---------------------------------------------------------------------------
# Theme stylesheets
# ---------------------------------------------------------------------------
_DARK_THEME_SS = """
    QMainWindow, QWidget {
        background: #1E1E1E;
        color: #F0F0F0;
    }
    QDialog { background: #1E1E1E; color: #F0F0F0; }
    QMenuBar { background: #2B2B2B; color: #F0F0F0; }
    QMenuBar::item:selected { background: #3A3A3A; color: #FFFFFF; }
    QMenu { background: #2B2B2B; color: #F0F0F0; border: 1px solid #555555; }
    QMenu::item:selected { background: #3D5A80; color: #FFFFFF; }
    QMenu::item:disabled { color: #777777; }
    QMenu::separator { height: 1px; background: #444444; margin: 3px 6px; }
    QTabWidget::pane { background: #1E1E1E; border: 1px solid #3A3A3A; }
    QTabBar::tab { background: #2B2B2B; color: #CCCCCC; min-width: 140px;
                   padding: 4px 12px; border: 1px solid #3A3A3A; border-bottom: none; }
    QTabBar::tab:selected { background: #1E1E1E; color: #FFFFFF; font-weight: bold; }
    QTabBar::tab:hover:!selected { background: #353535; }
    QStatusBar { background: #2B2B2B; color: #CCCCCC; }
    QStatusBar QLabel { color: #CCCCCC; }
    QDockWidget { color: #F0F0F0; }
    QDockWidget::title { background: #1F4E79; color: #FFFFFF;
                         padding: 4px 8px; font-weight: bold; }
    QScrollBar:vertical { background: #2B2B2B; width: 10px; }
    QScrollBar::handle:vertical { background: #555555; border-radius: 5px; }
    QScrollBar:horizontal { background: #2B2B2B; height: 10px; }
    QScrollBar::handle:horizontal { background: #555555; border-radius: 5px; }
    QPushButton { background: #3A3A3A; color: #F0F0F0;
                  border: 1px solid #555555; border-radius: 4px; padding: 4px 10px; }
    QPushButton:hover   { background: #4A4A4A; }
    QPushButton:pressed { background: #2A2A2A; }
    QPushButton:disabled { color: #666666; background: #2B2B2B; }
    QDialogButtonBox QPushButton { min-width: 72px; }
    QLineEdit { background: #2B2B2B; color: #F0F0F0;
                border: 1px solid #555555; border-radius: 3px; padding: 2px 4px; }
    QTextEdit, QPlainTextEdit { background: #2B2B2B; color: #F0F0F0;
                                 border: 1px solid #555555; }
    QSpinBox { background: #2B2B2B; color: #F0F0F0;
               border: 1px solid #555555; border-radius: 3px; padding: 2px 4px; }
    QComboBox { background: #2B2B2B; color: #F0F0F0;
                border: 1px solid #555555; border-radius: 3px; padding: 2px 4px; }
    QComboBox QAbstractItemView { background: #2B2B2B; color: #F0F0F0;
                                   selection-background-color: #3D5A80;
                                   selection-color: #FFFFFF; }
    QTreeWidget { background: #252525; color: #F0F0F0; border: none;
                  alternate-background-color: #2B2B2B; }
    QTreeWidget::item:selected { background: #3D5A80; color: #FFFFFF; }
    QTreeWidget::item:hover { background: #333333; }
    QTableWidget { background: #252525; color: #F0F0F0;
                   gridline-color: #3A3A3A; border: 1px solid #3A3A3A; }
    QTableWidget::item:selected { background: #3D5A80; color: #FFFFFF; }
    QHeaderView::section { background: #2B2B2B; color: #F0F0F0;
                            border: 1px solid #3A3A3A; padding: 3px 6px;
                            font-weight: bold; }
    QLabel { color: #F0F0F0; background: transparent; }
    QGroupBox { color: #F0F0F0; border: 1px solid #555555;
                border-radius: 4px; margin-top: 8px; padding-top: 8px; }
    QGroupBox::title { color: #F0F0F0; subcontrol-origin: margin; left: 8px; }
    QListWidget { background: #252525; color: #F0F0F0; border: 1px solid #3A3A3A; }
    QListWidget::item:selected { background: #3D5A80; color: #FFFFFF; }
    QCheckBox { color: #F0F0F0; }
    QRadioButton { color: #F0F0F0; }
    QToolTip { background: #2B2B2B; color: #F0F0F0; border: 1px solid #555555; }
"""

_LIGHT_THEME_SS = """
    QMainWindow, QWidget {
        background: #F5F5F5;
        color: #1A1A1A;
    }
    QDialog { background: #F5F5F5; color: #1A1A1A; }
    QMenuBar { background: #E8E8E8; color: #1A1A1A; }
    QMenuBar::item:selected { background: #D0D8E8; color: #000000; }
    QMenu { background: #FFFFFF; color: #1A1A1A; border: 1px solid #BBBBBB; }
    QMenu::item:selected { background: #3D5A80; color: #FFFFFF; }
    QMenu::item:disabled { color: #999999; }
    QMenu::separator { height: 1px; background: #CCCCCC; margin: 3px 6px; }
    QTabWidget::pane { background: #F5F5F5; border: 1px solid #CCCCCC; }
    QTabBar::tab { background: #E0E0E0; color: #333333; min-width: 140px;
                   padding: 4px 12px; border: 1px solid #CCCCCC; border-bottom: none; }
    QTabBar::tab:selected { background: #F5F5F5; color: #000000; font-weight: bold; }
    QTabBar::tab:hover:!selected { background: #EFEFEF; }
    QStatusBar { background: #E8E8E8; color: #333333; }
    QStatusBar QLabel { color: #333333; }
    QDockWidget { color: #1A1A1A; }
    QDockWidget::title { background: #3D5A80; color: #FFFFFF;
                         padding: 4px 8px; font-weight: bold; }
    QScrollBar:vertical { background: #E0E0E0; width: 10px; }
    QScrollBar::handle:vertical { background: #AAAAAA; border-radius: 5px; }
    QScrollBar:horizontal { background: #E0E0E0; height: 10px; }
    QScrollBar::handle:horizontal { background: #AAAAAA; border-radius: 5px; }
    QPushButton { background: #E0E0E0; color: #1A1A1A;
                  border: 1px solid #AAAAAA; border-radius: 4px; padding: 4px 10px; }
    QPushButton:hover   { background: #D0D0D0; }
    QPushButton:pressed { background: #C0C0C0; }
    QPushButton:disabled { color: #999999; background: #EBEBEB; }
    QDialogButtonBox QPushButton { min-width: 72px; }
    QLineEdit { background: #FFFFFF; color: #1A1A1A;
                border: 1px solid #AAAAAA; border-radius: 3px; padding: 2px 4px; }
    QTextEdit, QPlainTextEdit { background: #FFFFFF; color: #1A1A1A;
                                 border: 1px solid #AAAAAA; }
    QSpinBox { background: #FFFFFF; color: #1A1A1A;
               border: 1px solid #AAAAAA; border-radius: 3px; padding: 2px 4px; }
    QComboBox { background: #FFFFFF; color: #1A1A1A;
                border: 1px solid #AAAAAA; border-radius: 3px; padding: 2px 4px; }
    QComboBox QAbstractItemView { background: #FFFFFF; color: #1A1A1A;
                                   selection-background-color: #3D5A80;
                                   selection-color: #FFFFFF; }
    QTreeWidget { background: #FFFFFF; color: #1A1A1A; border: none;
                  alternate-background-color: #F0F4FA; }
    QTreeWidget::item:selected { background: #3D5A80; color: #FFFFFF; }
    QTreeWidget::item:hover { background: #E8EEF8; }
    QTableWidget { background: #FFFFFF; color: #1A1A1A;
                   gridline-color: #DDDDDD; border: 1px solid #CCCCCC; }
    QTableWidget::item:selected { background: #3D5A80; color: #FFFFFF; }
    QHeaderView::section { background: #E8E8E8; color: #1A1A1A;
                            border: 1px solid #CCCCCC; padding: 3px 6px;
                            font-weight: bold; }
    QLabel { color: #1A1A1A; background: transparent; }
    QGroupBox { color: #1A1A1A; border: 1px solid #AAAAAA;
                border-radius: 4px; margin-top: 8px; padding-top: 8px; }
    QGroupBox::title { color: #1A1A1A; subcontrol-origin: margin; left: 8px; }
    QListWidget { background: #FFFFFF; color: #1A1A1A; border: 1px solid #CCCCCC; }
    QListWidget::item:selected { background: #3D5A80; color: #FFFFFF; }
    QCheckBox { color: #1A1A1A; }
    QRadioButton { color: #1A1A1A; }
    QToolTip { background: #FFFDE7; color: #1A1A1A; border: 1px solid #BBBBBB; }
"""

_DARK_TOOLBAR_SS = """
    QToolBar { background: #252525; border-bottom: 1px solid #3A3A3A;
               padding: 2px 4px; spacing: 2px; }
    QToolButton { background: #2E2E2E; color: #CCCCCC;
                  border: 1px solid #484848; border-bottom: 3px solid #181818;
                  border-radius: 5px;
                  padding: 4px 10px; font-size: 9pt; min-width: 44px; }
    QToolButton:hover { background: #3A3A3A; color: #FFFFFF;
                        border: 1px solid #5A8ABB; border-bottom: 3px solid #2D5A80; }
    QToolButton:pressed { background: #1C1C1C; color: #AAAAAA;
                          border: 1px solid #333333; border-top: 3px solid #111111;
                          border-bottom: 1px solid #484848;
                          padding: 5px 9px 3px 11px; }
    QToolButton:checked { background: #1A3A5C; color: #7EC8F0;
                          border: 1px solid #3D5A80; border-bottom: 3px solid #1A3A5C; }
    QToolButton:disabled { background: #252525; color: #484848;
                           border: 1px solid #333333; border-bottom: 3px solid #141414; }
    QToolBar::separator { background: #3A3A3A; width: 1px; margin: 4px 6px; }
"""

_LIGHT_TOOLBAR_SS = """
    QToolBar { background: #E8E8E8; border-bottom: 1px solid #CCCCCC;
               padding: 2px 4px; spacing: 2px; }
    QToolButton { background: #F5F5F5; color: #333333;
                  border: 1px solid #C0C0C0; border-bottom: 3px solid #999999;
                  border-radius: 5px;
                  padding: 4px 10px; font-size: 9pt; min-width: 44px; }
    QToolButton:hover { background: #EBF3FC; color: #000000;
                        border: 1px solid #5A8ABB; border-bottom: 3px solid #3D6FA0; }
    QToolButton:pressed { background: #D5D5D5; color: #111111;
                          border: 1px solid #AAAAAA; border-top: 3px solid #888888;
                          border-bottom: 1px solid #C8C8C8;
                          padding: 5px 9px 3px 11px; }
    QToolButton:checked { background: #C8DCEC; color: #1A3A5C;
                          border: 1px solid #5A8ABB; border-bottom: 3px solid #3D6FA0; }
    QToolButton:disabled { background: #EBEBEB; color: #BBBBBB;
                           border: 1px solid #D8D8D8; border-bottom: 3px solid #C0C0C0; }
    QToolBar::separator { background: #CCCCCC; width: 1px; margin: 4px 6px; }
"""

_DARK_DOCK_SS = """
    QDockWidget::title { background:#1F4E79; color:white;
                         padding:4px 8px; font-weight:bold; }
"""

_LIGHT_DOCK_SS = """
    QDockWidget::title { background:#3D5A80; color:white;
                         padding:4px 8px; font-weight:bold; }
"""

# ─────────────────────────────────────────────────────────────────────
# TextCommentEditDialog — edit or create a text comment marker
# ─────────────────────────────────────────────────────────────────────
class TextCommentEditDialog(QDialog):
    """Dialog for creating or editing a text comment on a PDF."""
    
    def __init__(self, text: str = "", page: int = 0, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Text Comment")
        self.setMinimumWidth(400)
        self.delete_requested = False
        self._page = page
        
        form = QVBoxLayout()
        form.setContentsMargins(14, 14, 14, 10)
        form.setSpacing(10)
        
        form.addWidget(QLabel(f"<b>Text Comment for Page {page + 1}:</b>"))
        
        self.text_edit = QPlainTextEdit(text)
        self.text_edit.setMinimumHeight(120)
        self.text_edit.setPlaceholderText("Enter your comment here…")
        form.addWidget(self.text_edit)
        
        page_info = QLabel(f"Page: {page + 1}")
        page_info.setStyleSheet("color:#AAAAAA; font-size: 8pt;")
        form.addWidget(page_info)
        
        # Buttons
        buttons = QDialogButtonBox()
        ok_btn     = buttons.addButton(QDialogButtonBox.StandardButton.Ok)
        cancel_btn = buttons.addButton(QDialogButtonBox.StandardButton.Cancel)
        del_btn    = buttons.addButton("Delete", QDialogButtonBox.ButtonRole.DestructiveRole)
        
        ok_btn.clicked.connect(self.accept)
        cancel_btn.clicked.connect(self.reject)
        del_btn.clicked.connect(self._on_delete)
        del_btn.setStyleSheet("color:#B71C1C; font-weight:bold;")
        
        form.addWidget(buttons)
        self.setLayout(form)
    
    def _on_delete(self):
        self.delete_requested = True
        self.accept()
    
    @property
    def comment_text(self) -> str:
        return self.text_edit.toPlainText().strip()
    
# ---------------------------------------------------------------------------
# MarkerEditDialog  — edit a signal typical marker
# ---------------------------------------------------------------------------
class MarkerEditDialog(QDialog):
    """
    Opens when a marker is right-click -> Edit.
    Shows the full Signal Typical details; all fields are editable except
    Control Module Name, Field Device Name, Signal Name column, Type column.
    The count (multiplier) spinner is NOT shown.
    Any extra columns defined in the Signal Typical configuration are
    displayed and editable here too.
    """

    @staticmethod
    def _ro_colors() -> tuple:
        """Return (bg QColor, fg QColor) for read-only cells, respecting the current theme."""
        palette = QApplication.instance().palette()
        is_dark = palette.color(QPalette.ColorRole.Window).lightness() < 128
        if is_dark:
            return QColor("#2A2A2A"), QColor("#888888")
        return QColor("#E0E0E0"), QColor("#555555")

    def __init__(self, marker: dict, signal_types: list[dict], parent=None):
        super().__init__(parent)
        self.setWindowTitle("Edit Signal Typical")
        self.setMinimumSize(700, 560)
        self.delete_requested = False
        self._signal_types = signal_types
        self._marker = marker

        lay = QVBoxLayout()
        lay.setContentsMargins(16, 14, 16, 10)
        lay.setSpacing(10)
        self.setLayout(lay)

        composition_id = marker.get("composition_id")
        self._composition = (db_load_signal_composition(composition_id)
                             if composition_id else None)

        if not self._composition:
            lay.addWidget(QLabel("Warning: Signal Typical not found in configuration."))
        else:
            self._build_composition_form(lay)

        # Page info
        page_lbl = QLabel(f"<b>Page:</b> {marker['page'] + 1}")
        page_lbl.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        lay.addWidget(page_lbl)

        # Buttons
        buttons = QDialogButtonBox()
        ok_btn     = buttons.addButton(QDialogButtonBox.StandardButton.Ok)
        cancel_btn = buttons.addButton(QDialogButtonBox.StandardButton.Cancel)
        del_btn    = buttons.addButton("Delete", QDialogButtonBox.ButtonRole.DestructiveRole)
        ok_btn.clicked.connect(self.accept)
        cancel_btn.clicked.connect(self.reject)
        del_btn.clicked.connect(self._on_delete)
        del_btn.setStyleSheet("color: #B71C1C; font-weight: bold;")
        lay.addWidget(buttons)

    def _build_composition_form(self, lay: QVBoxLayout):
        comp = self._composition
        tag_parts = self._marker.get("tag_parts", {})

        # Title
        composition_text = _get_signal_composition(comp)
        title_lbl = QLabel(
            f"<b>{comp['title']}</b>"
            + (f"  <span style='color:#7EC8F0;'>{composition_text}</span>"
               if composition_text else ""))
        title_lbl.setStyleSheet("font-size: 11pt;")
        lay.addWidget(title_lbl)

        # Description
        desc_row = QHBoxLayout()
        desc_row.addWidget(QLabel("<b>Description:</b>"))
        self.desc_edit = QLineEdit(
            tag_parts.get("description", comp.get("description", "")))
        desc_row.addWidget(self.desc_edit)
        lay.addLayout(desc_row)

        # Control Module + Field Device side by side
        cm_fd_lay = QHBoxLayout()

        cm_group = QGroupBox("Control Module")
        cm_group.setStyleSheet(
            "QGroupBox{border:1px solid #555;border-radius:4px;margin-top:6px;"
            "color:#F0F0F0;font-weight:bold;}"
            "QGroupBox::title{subcontrol-origin:margin;left:8px;padding:0 4px;}")
        cm_form = QFormLayout(cm_group)
        cm_form.setContentsMargins(8, 8, 8, 4)
        cm_form.setSpacing(4)
        _cm_name_lbl = QLabel(comp.get("control_module", "NA") or "NA")
        _cm_name_lbl.setStyleSheet("color:#AAAAAA;")
        cm_form.addRow("Name:", _cm_name_lbl)
        self.cm_type_edit = QLineEdit(
            tag_parts.get("cm_type", comp.get("cm_type", "NA") or "NA"))
        self.cm_desc_edit = QLineEdit(
            tag_parts.get("cm_description", comp.get("cm_description", "NA") or "NA"))
        cm_form.addRow("Type:", self.cm_type_edit)
        cm_form.addRow("Description:", self.cm_desc_edit)
        cm_fd_lay.addWidget(cm_group)

        fd_group = QGroupBox("Field Device")
        fd_group.setStyleSheet(
            "QGroupBox{border:1px solid #555;border-radius:4px;margin-top:6px;"
            "color:#F0F0F0;font-weight:bold;}"
            "QGroupBox::title{subcontrol-origin:margin;left:8px;padding:0 4px;}")
        fd_form = QFormLayout(fd_group)
        fd_form.setContentsMargins(8, 8, 8, 4)
        fd_form.setSpacing(4)
        _fd_name_lbl = QLabel(comp.get("field_device", "NA") or "NA")
        _fd_name_lbl.setStyleSheet("color:#AAAAAA;")
        fd_form.addRow("Name:", _fd_name_lbl)
        self.fd_type_edit = QLineEdit(
            tag_parts.get("fd_type", comp.get("fd_type", "NA") or "NA"))
        self.fd_desc_edit = QLineEdit(
            tag_parts.get("fd_description", comp.get("fd_description", "NA") or "NA"))
        fd_form.addRow("Type:", self.fd_type_edit)
        fd_form.addRow("Description:", self.fd_desc_edit)
        cm_fd_lay.addWidget(fd_group)
        lay.addLayout(cm_fd_lay)

        # Signals table — mirrors SignalCompositionConfigDialog layout
        extra_headers = comp.get("extra_column_headers", [])
        fixed_cols = ["Signal Name", "Type", "Description", "Prefix", "Suffix"]
        all_headers = fixed_cols + extra_headers
        lay.addWidget(QLabel("<b>Signals:</b>"))
        self.signals_table = QTableWidget(0, len(all_headers))
        self.signals_table.setHorizontalHeaderLabels(all_headers)
        hdr = self.signals_table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        for ci in range(3, len(all_headers)):
            hdr.setSectionResizeMode(ci, QHeaderView.ResizeMode.Interactive)
        self.signals_table.setColumnWidth(0, 110)
        self.signals_table.setColumnWidth(1, 80)
        self.signals_table.verticalHeader().setVisible(False)
        self.signals_table.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows)
        self.signals_table.setEditTriggers(
            QAbstractItemView.EditTrigger.DoubleClicked |
            QAbstractItemView.EditTrigger.SelectedClicked)

        overrides = tag_parts.get("signal_overrides", [])
        signals = comp.get("signals", [])
        ro_bg, ro_fg = self._ro_colors()
        for i, sig in enumerate(signals):
            ov = overrides[i] if i < len(overrides) else {}
            r = self.signals_table.rowCount()
            self.signals_table.insertRow(r)

            # Read-only: Signal Name (col 0), Type (col 1)
            for ci, val in enumerate([sig.get("signal_name", ""),
                                       sig.get("signal_type", "")]):
                item = QTableWidgetItem(val)
                item.setFlags(Qt.ItemFlag.ItemIsEnabled)
                item.setBackground(QBrush(ro_bg))
                item.setForeground(QBrush(ro_fg))
                self.signals_table.setItem(r, ci, item)

            # Editable: Description (2), Prefix (3), Suffix (4)
            self.signals_table.setItem(r, 2, QTableWidgetItem(
                ov.get("signal_description", sig.get("signal_description", ""))))
            self.signals_table.setItem(r, 3, QTableWidgetItem(
                ov.get("prefix", sig.get("prefix", "NA"))))
            self.signals_table.setItem(r, 4, QTableWidgetItem(
                ov.get("suffix", sig.get("suffix", "NA"))))

            # Extra columns (editable)
            ov_extras = ov.get("extra_column_values",
                               sig.get("extra_column_values", []))
            for ec in range(len(extra_headers)):
                val = ov_extras[ec] if ec < len(ov_extras) else ""
                self.signals_table.setItem(r, 5 + ec, QTableWidgetItem(val))

        lay.addWidget(self.signals_table, stretch=1)

    def _on_delete(self):
        self.delete_requested = True
        self.accept()

    @property
    def tag_parts(self) -> dict:
        """Return all editable overrides (no count)."""
        if not self._composition:
            return {}
        extra_headers = self._composition.get("extra_column_headers", [])
        overrides = []
        for r in range(self.signals_table.rowCount()):
            ev = [
                (self.signals_table.item(r, 5 + ec) or QTableWidgetItem()).text()
                for ec in range(len(extra_headers))
            ]
            overrides.append({
                "signal_description": (self.signals_table.item(r, 2) or QTableWidgetItem()).text(),
                "prefix":             (self.signals_table.item(r, 3) or QTableWidgetItem()).text(),
                "suffix":             (self.signals_table.item(r, 4) or QTableWidgetItem()).text(),
                "extra_column_values": ev,
            })
        return {
            "description":    self.desc_edit.text().strip(),
            "cm_type":        self.cm_type_edit.text().strip(),
            "cm_description": self.cm_desc_edit.text().strip(),
            "fd_type":        self.fd_type_edit.text().strip(),
            "fd_description": self.fd_desc_edit.text().strip(),
            "signal_overrides": overrides,
        }

    @property
    def count(self) -> int:
        """Return existing count (not editable here)."""
        return int(self._marker.get("tag_parts", {}).get(
            "count", self._marker.get("count", 1)))


    # ---------------------------------------------------------------------------
    # MarkerCountDialog  — count + description fields shown after signal selection
    # ---------------------------------------------------------------------------
    _DARK_SS = (
        "* { background-color: #1E1E1E; color: #F0F0F0; }"
        "QDialog { background: #1E1E1E; }"
        "QLabel { color: #F0F0F0; background: transparent; }"
        "QLineEdit { background: #2B2B2B; color: #F0F0F0;"
        " border: 1px solid #555555; border-radius: 3px; padding: 2px 4px; }"
        "QSpinBox { background: #2B2B2B; color: #F0F0F0;"
        " border: 1px solid #555555; border-radius: 3px;"
        " padding: 2px 4px; min-width: 50px; max-width: 60px; }"
        "QSpinBox::up-button   { width: 0px; border: none; }"
        "QSpinBox::down-button { width: 0px; border: none; }"
        "QPushButton { background: #3A3A3A; color: #F0F0F0;"
        " border: 1px solid #555555; border-radius: 4px; padding: 4px 10px; }"
        "QPushButton:hover { background: #4A4A4A; }"
        "QPushButton:pressed { background: #2A2A2A; }"
        "QDialogButtonBox QPushButton { min-width: 72px; }"
    )

class MarkerCountDialog(QDialog):
    """Dialog to set count and descriptions for a marker."""
    
    def __init__(self, label: str, comment: str, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Configure {label}")
        self.setMinimumWidth(450)
        self._descriptions = []
        self._label = label
        self._comment = comment
        self._count = 1
        
        lay = QVBoxLayout()
        lay.setContentsMargins(14, 14, 14, 10)
        lay.setSpacing(10)
        
        # Header
        title = QLabel(f"<b>Configure: {label}</b>")
        if comment:
            title.setText(f"<b>{label}</b>  —  {comment}")
        lay.addWidget(title)
        
        # Count section with manual buttons
        count_lay = QHBoxLayout()
        count_lay.addWidget(QLabel("Count:"))
        
        # Decrease button
        self.dec_btn = QPushButton("−")
        self.dec_btn.setMaximumWidth(40)
        self.dec_btn.clicked.connect(self._decrease_count)
        count_lay.addWidget(self.dec_btn)
        
        # Count display
        self.count_label = QLabel("1")
        self.count_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.count_label.setStyleSheet("font-weight: bold; font-size: 14pt; min-width: 40px;")
        count_lay.addWidget(self.count_label)
        
        # Increase button
        self.inc_btn = QPushButton("+")
        self.inc_btn.setMaximumWidth(40)
        self.inc_btn.clicked.connect(self._increase_count)
        count_lay.addWidget(self.inc_btn)
        
        count_lay.addWidget(QLabel("(number of signals)"))
        count_lay.addStretch()
        lay.addLayout(count_lay)
        
        # Description fields
        desc_lbl = QLabel("<b>Description per signal (optional):</b>")
        lay.addWidget(desc_lbl)
        
        self.desc_layout = QVBoxLayout()
        self.desc_layout.setSpacing(6)
        lay.addLayout(self.desc_layout)
        
        # Buttons
        bb = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok |
                              QDialogButtonBox.StandardButton.Cancel)
        bb.accepted.connect(self.accept)
        bb.rejected.connect(self.reject)
        lay.addWidget(bb)
        
        self.setLayout(lay)
        self._update_description_fields()
    
    def _decrease_count(self):
        """Decrease count by 1 (minimum 1)."""
        if self._count > 1:
            self._count -= 1
            self.count_label.setText(str(self._count))
            self._update_description_fields()
    
    def _increase_count(self):
        """Increase count by 1 (maximum 999)."""
        if self._count < 999:
            self._count += 1
            self.count_label.setText(str(self._count))
            self._update_description_fields()
    
    def _update_description_fields(self):
        """Update description fields based on count."""
        # Clear existing fields
        while self.desc_layout.count():
            item = self.desc_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        
        self._descriptions = []
        
        # Create new fields
        for i in range(self._count):
            hlayout = QHBoxLayout()
            hlayout.setSpacing(8)
            
            label = QLabel(f"Signal {i+1}:")
            label.setMaximumWidth(70)
            hlayout.addWidget(label)
            
            edit = QLineEdit()
            edit.setPlaceholderText("e.g., Valve position")
            hlayout.addWidget(edit)
            
            self.desc_layout.addLayout(hlayout)
            self._descriptions.append(edit)
    
    @property
    def count(self) -> int:
        """Return the count value."""
        return self._count
    
    @property
    def descriptions(self) -> list[str]:
        """Return list of descriptions."""
        return [e.text().strip() for e in self._descriptions]    
                          
# ---------------------------------------------------------------------------
# MarkerOverlay
# ---------------------------------------------------------------------------
class MarkerOverlay(QWidget):
    # Emitted after a marker is edited or deleted so PDFViewer can update its
    # status bar and print log without needing to know overlay internals.
    markers_changed = Signal()

    def __init__(self, pdf_view: "DraggablePdfView", io_list: list):
        super().__init__(pdf_view.viewport())
        self._pdf_view = pdf_view
        self._markers  = io_list

        self.setAttribute(Qt.WidgetAttribute.WA_NoSystemBackground)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setCursor(Qt.CursorShape.ArrowCursor)
        self.raise_()

        # Enable mouse double-click events
        self.setMouseTracking(True)
        self.setAttribute(Qt.WidgetAttribute.WA_Hover)
    
        # Drag-move state
        self._dragging_marker: dict | None = None
        self._drag_offset = QPointF()

        # Selection & clipboard
        self._selected_marker: dict | None = None
        self._clipboard:       dict | None = None

        # Undo callback — set by PDFViewer after construction
        self._push_undo_fn = lambda: None

        self.setFocusPolicy(Qt.FocusPolicy.ClickFocus)

    # ------------------------------------------------------------------
    # Marker management (called from PDFViewer)
    # ------------------------------------------------------------------
    def add_marker(self) -> None:
        # Caller (PDFViewer.register_io) already appended to io_list,
        # so we just need to repaint.
        self.update()

    def clear_markers(self) -> None:
        # Caller clears io_list; we just repaint.
        self.update()

    def sync_size(self) -> None:
        self.setGeometry(self._pdf_view.viewport().rect())
        self.update()

    # ------------------------------------------------------------------
    # Hit testing
    # ------------------------------------------------------------------
    def _marker_at(self, screen_pos: QPointF) -> dict | None:
        from PySide6.QtGui import QFont, QFontMetricsF
        label_fm = QFontMetricsF(QFont("Arial", 8))

        for m in reversed(self._markers):   # top-most first
            if "pdf_x" not in m:
                continue
            vp = self._pdf_view._pdf_to_widget_pos(m["pdf_x"], m["pdf_y"], m["page"])

            if m.get("kind") == "text":
                # Text box is centred on vp; estimate size
                font = QFont("Arial", 8)
                fm   = QFontMetricsF(font)
                text = m.get("type", "")
                max_w = 200.0; pad_x = 8.0; pad_y = 5.0
                words = text.split(); lines, line = [], ""
                for word in words:
                    test = (line + " " + word).strip()
                    if fm.horizontalAdvance(test) <= max_w - 2 * pad_x:
                        line = test
                    else:
                        if line: lines.append(line)
                        line = word
                if line: lines.append(line)
                if not lines: lines = [""]
                inner_w = min(max(fm.horizontalAdvance(l) for l in lines),
                              max_w - 2 * pad_x)
                w = inner_w + 2 * pad_x
                h = fm.height() * len(lines) + 2 * pad_y
                hit_rect = QRectF(vp.x() - w / 2, vp.y() - h / 2, w, h)
            else:
                pad_x, pad_y = 7.0, 5.0
                lw = label_fm.horizontalAdvance(m["type"])
                cmt = m.get("comment", "")
                cw = label_fm.horizontalAdvance(cmt) if cmt else 0
                w  = max(lw, cw, MARKER_W - 2 * pad_x) + 2 * pad_x
                lh = label_fm.height()
                h  = lh + 2 * pad_y
                hit_rect = QRectF(vp.x() - w / 2, vp.y() - h - MARKER_TH, w,
                                  h + MARKER_TH)

            if hit_rect.contains(screen_pos):
                return m
        return None

    # ------------------------------------------------------------------
    # Mouse events
    # ------------------------------------------------------------------
    def mousePressEvent(self, event):
        hit = self._marker_at(event.position())

        if event.button() == Qt.MouseButton.LeftButton:
            if hit:
                # Track this click for double-click detection
                if not hasattr(self, '_click_pos'):
                    self._click_pos = event.position()
                    self._click_marker = hit
                
                # Select and start drag
                self._selected_marker = hit
                tip = self._pdf_view._pdf_to_widget_pos(
                    hit["pdf_x"], hit["pdf_y"], hit["page"])
                self._dragging_marker = hit
                self._drag_offset = event.position() - tip
                self.setCursor(Qt.CursorShape.SizeAllCursor)
                self.setFocus()
                self.update()
            else:
                # Deselect
                self._selected_marker = None
                self.update()
                event.ignore()
            return

        if event.button() == Qt.MouseButton.RightButton:
            if hit:
                self._selected_marker = hit
                self.update()
                self._show_marker_menu(hit, event.globalPosition().toPoint())
            else:
                self._selected_marker = None
                self.update()
                event.ignore()
            return

        event.ignore()
        
    def mouseDoubleClickEvent(self, event):
        """Handle double-click on a marker to show info."""
        hit = self._marker_at(event.position())
        if hit:
            # Find the parent PDFViewer
            parent = self.parent()
            while parent:
                if hasattr(parent, 'show_marker_info'):
                    parent.show_marker_info(hit)
                    break
                parent = parent.parent()
            event.accept()
            return
        event.ignore()

    def mouseMoveEvent(self, event):
        if self._dragging_marker is not None:
            widget_pos = event.position() - self._drag_offset
            pdf_x, pdf_y, page = self._pdf_view._widget_pos_to_pdf(widget_pos)
            self._dragging_marker["pdf_x"] = pdf_x
            self._dragging_marker["pdf_y"] = pdf_y
            self._dragging_marker["page"]  = page
            self.update()
            return

        hit = self._marker_at(event.position())
        if hit:
            self.setCursor(Qt.CursorShape.PointingHandCursor)
            count   = hit.get("count", 1)
            comment = hit.get("comment", "")
            desc    = hit.get("description", "")
            label   = f"{count}{hit['type']}" if count and count > 1 else hit["type"]
            tip     = label
            if comment:
                tip += f"  —  {comment}"
            if desc:
                tip += f"\n{desc}"
            from PySide6.QtWidgets import QToolTip
            from PySide6.QtGui import QCursor
            QToolTip.showText(QCursor.pos(), tip, self)
        else:
            self.setCursor(Qt.CursorShape.ArrowCursor)
            from PySide6.QtWidgets import QToolTip
            QToolTip.hideText()
        event.ignore()

    def mouseReleaseEvent(self, event):
        if self._dragging_marker is not None:
            self._push_undo_fn()
            self._dragging_marker = None
            self.setCursor(Qt.CursorShape.ArrowCursor)
            self.markers_changed.emit()
            self.update()
            return
        event.ignore()

    def mouseDoubleClickEvent(self, event):
        """Handle double-click on a marker to show info."""
        hit = self._marker_at(event.position())
        if hit:
            from PySide6.QtWidgets import QApplication
            # Find the parent PDFViewer to open the dialog
            # Walk up the widget hierarchy
            parent = self.parent()
            while parent:
                if hasattr(parent, 'show_marker_info'):
                    parent.show_marker_info(hit)
                    break
                parent = parent.parent()
            event.accept()
            return
        event.ignore()
    # ------------------------------------------------------------------
    # Keyboard — Ctrl+C copy, Ctrl+V / Ctrl+P paste, Delete, Arrow move
    # ------------------------------------------------------------------
    # Movement step in PDF points; hold Shift for bigger jumps
    _MOVE_STEP       =  2   # pts  (~0.7 mm)
    _MOVE_STEP_LARGE = 10   # pts  (~3.5 mm) with Shift

    def keyPressEvent(self, event):
        key  = event.key()
        mods = event.modifiers()
        ctrl  = bool(mods & Qt.KeyboardModifier.ControlModifier)
        shift = bool(mods & Qt.KeyboardModifier.ShiftModifier)

        # ── Ctrl shortcuts ────────────────────────────────────────────────
        if ctrl:
            if key == Qt.Key.Key_C and self._selected_marker:
                self._copy_marker(self._selected_marker)
                return
            if key in (Qt.Key.Key_V, Qt.Key.Key_P) and self._clipboard:
                self._paste_marker_offset()
                return

        # ── Delete selected marker ────────────────────────────────────────
        if key in (Qt.Key.Key_Delete, Qt.Key.Key_Backspace):
            if self._selected_marker:
                self._push_undo_fn()
                self._markers.remove(self._selected_marker)
                self._selected_marker = None
                self.markers_changed.emit()
                self.update()
                return

        # ── Arrow-key nudge ───────────────────────────────────────────────
        if key in (Qt.Key.Key_Left, Qt.Key.Key_Right,
                   Qt.Key.Key_Up,   Qt.Key.Key_Down):
            if self._selected_marker:
                step = self._MOVE_STEP_LARGE if shift else self._MOVE_STEP
                dx, dy = 0.0, 0.0
                if key == Qt.Key.Key_Left:  dx = -step
                if key == Qt.Key.Key_Right: dx =  step
                if key == Qt.Key.Key_Up:    dy =  step   # PDF Y is bottom-up
                if key == Qt.Key.Key_Down:  dy = -step
                self._push_undo_fn()
                self._selected_marker["pdf_x"] += dx
                self._selected_marker["pdf_y"] += dy
                self.markers_changed.emit()
                self.update()
                return

        event.ignore()

    # ------------------------------------------------------------------
    # Copy / Paste helpers
    # ------------------------------------------------------------------
    def _copy_marker(self, marker: dict) -> None:
        import copy
        self._clipboard = copy.deepcopy(marker)

    def _paste_marker_at(self, pdf_x: float, pdf_y: float, page: int) -> None:
        """Paste clipboard contents at the given PDF coordinates."""
        import copy
        if not self._clipboard:
            return
        self._push_undo_fn()
        new_marker = copy.deepcopy(self._clipboard)
        new_marker["pdf_x"] = pdf_x
        new_marker["pdf_y"] = pdf_y
        new_marker["page"]  = page
        self._markers.append(new_marker)
        self._selected_marker = new_marker
        self.markers_changed.emit()
        self.update()

    def _paste_marker_offset(self) -> None:
        """Paste near the original (keyboard shortcut — no cursor position)."""
        if not self._clipboard:
            return
        OFFSET = 20  # PDF points
        self._paste_marker_at(
            self._clipboard["pdf_x"] + OFFSET,
            self._clipboard["pdf_y"] - OFFSET,
            self._clipboard["page"])

    # ------------------------------------------------------------------
    # Marker right-click menu  (only shown when a marker is hit)
    # ------------------------------------------------------------------
    def _show_marker_menu(self, marker: dict, global_pos) -> None:
        menu = QMenu(self)
        edit_act   = menu.addAction("✏️  Edit…")
        menu.addSeparator()
        copy_act   = menu.addAction("📋  Copy  Ctrl+C")

        paste_act = None
        if self._clipboard:
            paste_act = menu.addAction("📌  Paste here  Ctrl+V")

        # Conversion option — only for IO signal markers, not text comments
        convert_act = None
        if marker.get("kind") != "text":
            menu.addSeparator()
            convert_act = menu.addAction("🔄  Convert to Text Comment…")

        menu.addSeparator()
        del_act = menu.addAction("🗑  Delete")
        del_act.setData("delete")

        chosen = menu.exec(global_pos)
        if not chosen:
            return

        if chosen is edit_act:
            self._open_edit_dialog(marker)
        elif chosen is copy_act:
            self._copy_marker(marker)
        elif paste_act and chosen is paste_act:
            self._paste_marker_offset()
        elif convert_act and chosen is convert_act:
            self._convert_to_text_comment(marker)
        elif chosen is del_act:
            self._push_undo_fn()
            self._markers.remove(marker)
            if self._selected_marker is marker:
                self._selected_marker = None
            self.markers_changed.emit()
            self.update()
    
    def _show_marker_info(self, marker: dict) -> None:
        """Show marker info dialog by finding parent PDFViewer."""
        # Walk up from overlay to find PDFViewer
        widget = self
        while widget:
            if type(widget).__name__ == 'PDFViewer':
                # Found PDFViewer, call show_marker_info
                widget.show_marker_info(marker)
                return
            
            widget = widget.parent() if hasattr(widget, 'parent') else None

    # ------------------------------------------------------------------
    # Convert IO marker → text comment
    # ------------------------------------------------------------------
    def _convert_to_text_comment(self, marker: dict) -> None:
        """
        Convert an IO signal marker into a text comment in-place.

        The marker's data is composed into a readable text string that is
        pre-filled into the TextCommentEditDialog so the user can review
        and edit it before confirming.  On OK the marker kind changes to
        'text' and signal-specific fields are cleared.  The position on
        the page is preserved.  The operation is undo-safe.
        """
        count   = int(marker.get("count", 1))
        label   = f"{count}{marker['type']}" if count > 1 else marker["type"]
        comment = marker.get("comment", "").strip()
        desc    = marker.get("description", "").strip()

        # Build a sensible default text from all available info
        parts = [label]
        if comment:
            parts.append(comment)
        if desc:
            # description may be pipe-separated (multi-field)
            for d in desc.split("|"):
                d = d.strip()
                if d:
                    parts.append(d)
        composed = "\n".join(parts)

        dlg = TextCommentEditDialog(composed, marker["page"], self)
        dlg.setWindowTitle("Convert to Text Comment — Review & Edit")
        if dlg.exec() != QDialog.DialogCode.Accepted or dlg.delete_requested:
            return

        self._push_undo_fn()
        # Mutate the marker in-place so existing references stay valid
        marker["kind"]        = "text"
        marker["type"]        = dlg.comment_text
        marker["comment"]     = ""
        marker["count"]       = 1
        marker["description"] = ""
        self.markers_changed.emit()
        self.update()

    # ------------------------------------------------------------------
    # Edit dialog
    # ------------------------------------------------------------------
    def _open_edit_dialog(self, marker: dict) -> None:
        """Open edit dialog for a marker (composition or custom)."""
        if marker.get("kind") == "text":
            # ── Text comment dialog ────────────────────────────────────
            dlg = TextCommentEditDialog(
                marker["type"], marker["page"], self)
            dlg.exec()
            if dlg.delete_requested:
                self._push_undo_fn()
                self._markers.remove(marker)
                if self._selected_marker is marker:
                    self._selected_marker = None
                self.markers_changed.emit()
                self.update()
            elif dlg.result() == QDialog.DialogCode.Accepted:
                self._push_undo_fn()
                if hasattr(dlg, 'tag_parts'):
                    marker["tag_parts"] = dlg.tag_parts
                    marker["count"]     = dlg.tag_parts.get("count", 1)
                self.markers_changed.emit()
                self.update()
        
        elif marker.get("is_composition"):
            # ── Composition marker edit dialog ─────────────────────────
            dlg = MarkerEditDialog(marker, [], parent=self)
            if dlg.exec() != QDialog.DialogCode.Accepted:
                return
            if dlg.delete_requested:
                self._push_undo_fn()
                self._markers.remove(marker)
                if self._selected_marker is marker:
                    self._selected_marker = None
                self.markers_changed.emit()
                self.update()
                return

            self._push_undo_fn()
            new_tag_parts = dlg.tag_parts
            # Preserve the existing count (not editable in this dialog)
            new_tag_parts["count"] = dlg.count

            composition = dlg._composition
            new_count   = dlg.count
            if composition:
                signal_counts = {}
                for sig in composition.get("signals", []):
                    sig_type = sig.get("signal_type", "")
                    if sig_type:
                        sig_count = int(sig.get("count", 1) or 1)
                        signal_counts[sig_type] = signal_counts.get(sig_type, 0) + sig_count * new_count
                parts = [f"{c}{t}" for t, c in sorted(signal_counts.items())]
                display_text = " ".join(parts) if parts else composition["title"]
                marker["type"] = display_text

            marker["tag_parts"] = new_tag_parts
            marker["count"]     = new_count
            self.markers_changed.emit()
            self.update()
        
        else:
            # ── Unknown marker type (skip) ─────────────────────────────
            pass

    # ------------------------------------------------------------------
    # Painting
    # ------------------------------------------------------------------
    def paintEvent(self, _event):
        if not self._markers:
            return

        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        for m in self._markers:
            if "pdf_x" not in m:
                continue
            vp = self._pdf_view._pdf_to_widget_pos(m["pdf_x"], m["pdf_y"], m["page"])
            is_sel = m is self._selected_marker
            is_drg = m is self._dragging_marker
            
            if m.get("kind") == "text":
                self._draw_text_box(painter, vp.x(), vp.y(),
                                    m["type"], dragging=is_drg, selected=is_sel)
            else:
                # Check if it's a composition marker
                if m.get("is_composition"):
                    # Use the stored type directly (already has count multiplied)
                    display_label = m.get("type", "?")
                else:
                    # Simple marker: show count + type
                    count = m.get("count", 1)
                    display_label = (f"{count}{m['type']}"
                                    if count and count > 1 else m["type"])
                
                self._draw_comment_bubble(painter, vp.x(), vp.y(),
                                        display_label, m.get("comment", ""),
                                        dragging=is_drg, selected=is_sel)

        painter.end()

    def _draw_comment_bubble(self, painter: QPainter,
                              cx: float, cy: float,
                              label: str = "", comment: str = "",
                              dragging: bool = False,
                              selected: bool = False) -> None:
        from PySide6.QtGui import QFont, QFontMetricsF

        tail_h = MARKER_TH
        r      = 5.0
        pad_x  = 7.0
        pad_y  = 5.0

        label_font = QFont("Arial", 8)
        label_font.setBold(True)
        label_fm   = QFontMetricsF(label_font)

       # Handle multi-line labels
        lines = label.split('\n') if label else [""]
        label_widths = [label_fm.horizontalAdvance(line) for line in lines]
        label_w = max(label_widths) if label_widths else 0
        label_h = label_fm.height()

        inner_w = max(label_w, MARKER_W - 2 * pad_x)
        inner_h = label_h * len(lines)  # Multiple lines
        w = inner_w + 2 * pad_x
        h = inner_h + 2 * pad_y

        body_left = cx - w / 2
        body_top  = cy - h - tail_h

        yellow     = QColor("#F9D71C") if not dragging else QColor("#FFE066")
        yellow_drk = QColor("#C9A800") if not dragging else QColor("#F9D71C")

        # Selection ring (drawn first, behind bubble)
        if selected:
            painter.setPen(QPen(QColor("#1565C0"), 2.5, Qt.PenStyle.DashLine))
            painter.setBrush(Qt.BrushStyle.NoBrush)
            painter.drawRoundedRect(
                QRectF(body_left - 3, body_top - 3, w + 6, h + 6), r + 2, r + 2)

        # Drop shadow
        shadow = QPainterPath()
        shadow.addRoundedRect(QRectF(body_left + 1.5, body_top + 1.5, w, h), r, r)
        painter.setBrush(QBrush(QColor(0, 0, 0, 40)))
        painter.setPen(Qt.PenStyle.NoPen)
        painter.drawPath(shadow)

        # Body + tail
        body_path = QPainterPath()
        body_path.addRoundedRect(QRectF(body_left, body_top, w, h), r, r)
        tail_path = QPainterPath()
        tail_path.moveTo(cx - 3,      body_top + h)
        tail_path.lineTo(cx,          cy)
        tail_path.lineTo(cx + w / 4,  body_top + h)
        tail_path.closeSubpath()
        full = body_path.united(tail_path)
        painter.setBrush(QBrush(yellow))
        painter.setPen(QPen(yellow_drk, 1.2))
        painter.drawPath(full)

        # Label text
        # Label text (handle multi-line for complex markers)
        text_x = body_left + pad_x
        text_y = body_top  + pad_y
        painter.setFont(label_font)
        painter.setPen(QPen(QColor("#3A2A00")))

        # Split label by newline for multi-line display
        lines = label.split('\n') if label else [""]
        line_height = label_fm.height()

        for i, line in enumerate(lines):
            y_offset = i * line_height
            painter.drawText(
                QRectF(text_x, text_y + y_offset, inner_w, line_height),
                Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter,
                line)

    def _draw_text_box(self, painter: QPainter,
                       cx: float, cy: float,
                       text: str = "",
                       dragging: bool = False,
                       selected: bool = False) -> None:
        """Draw a plain rounded white text box (no tail) for text comments."""
        from PySide6.QtGui import QFont, QFontMetricsF
        r     = 5.0
        pad_x = 8.0
        pad_y = 5.0
        max_w = 200.0

        font = QFont("Arial", 8)
        fm = QFontMetricsF(font)

        # Word-wrap text into lines that fit max_w
        words = text.split()
        lines, line = [], ""
        for word in words:
            test = (line + " " + word).strip()
            if fm.horizontalAdvance(test) <= max_w - 2 * pad_x:
                line = test
            else:
                if line:
                    lines.append(line)
                line = word
        if line:
            lines.append(line)
        if not lines:
            lines = [""]

        line_h  = fm.height()
        inner_w = min(max(fm.horizontalAdvance(l) for l in lines), max_w - 2 * pad_x)
        inner_h = line_h * len(lines) + (len(lines) - 1) * 2
        w = inner_w + 2 * pad_x
        h = inner_h + 2 * pad_y

        left = cx - w / 2
        top  = cy - h / 2

        # Selection ring
        if selected:
            painter.setPen(QPen(QColor("#1565C0"), 2.5, Qt.PenStyle.DashLine))
            painter.setBrush(Qt.BrushStyle.NoBrush)
            painter.drawRoundedRect(QRectF(left - 3, top - 3, w + 6, h + 6),
                                    r + 2, r + 2)

        # Shadow
        painter.setBrush(QBrush(QColor(0, 0, 0, 30)))
        painter.setPen(Qt.PenStyle.NoPen)
        painter.drawRoundedRect(QRectF(left + 1.5, top + 1.5, w, h), r, r)

        # Box
        bg = QColor("#FFFDE7") if not dragging else QColor("#FFF9C4")
        painter.setBrush(QBrush(bg))
        painter.setPen(QPen(QColor("#90A4AE"), 1.0))
        painter.drawRoundedRect(QRectF(left, top, w, h), r, r)

        # Text
        painter.setFont(font)
        painter.setPen(QPen(QColor("#212121")))  # text inside bubble stays dark (on light bubble bg)
        for i, ln in enumerate(lines):
            y = top + pad_y + i * (line_h + 2)
            painter.drawText(
                QRectF(left + pad_x, y, inner_w, line_h),
                Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter,
                ln)


# ---------------------------------------------------------------------------
# DraggablePdfView
# ---------------------------------------------------------------------------
class DraggablePdfView(QPdfView):
    io_registered   = Signal(str, QPointF, int, float, float, str, int, str, str, str)
    text_registered = Signal(str, int, float, float)   # text, page, pdf_x, pdf_y
    composition_registered = Signal(int, int, float, float, dict, dict)  

    def __init__(self, io_list: list, signal_types: list, parent=None):
        super().__init__(parent)
        self.setMouseTracking(True)
        self._drag_start    = None
        self._zoom          = 1.0
        self._signal_types  = signal_types
        self._input_mode    = "marker"   # "marker" | "text"
        self._pdf_path = ""  # set by PDFViewer after tab creation
        
        self.setPageMode(QPdfView.PageMode.MultiPage)
        self.setZoomMode(QPdfView.ZoomMode.Custom)

        self._overlay = MarkerOverlay(self, io_list)
        self.horizontalScrollBar().valueChanged.connect(self._overlay.update)
        self.verticalScrollBar().valueChanged.connect(self._overlay.update)
        self._overlay.sync_size()

    def _on_composition_registered(self, comp_id, page, pdf_x, pdf_y, composition, tag_parts):
        """Handle composition marker registration."""
        marker = {
            "type": composition["title"],
            "comment": "",
            "page": page,
            "pdf_x": pdf_x,
            "pdf_y": pdf_y,
            "kind": "marker",
            "count": 1,
            "description": "",
            "signal_type": "",
            "signal_type_comment": "",
            "composition_id": comp_id,
            "is_composition": True,
            "tag_parts": tag_parts,  # STORE tag_parts
            "base_tag": "",
            "complete_tag": "",
        }
        
        self._markers.append(marker)
        self._pdf_view.add_marker()
        
    # ------------------------------------------------------------------
    # Zoom
    # ------------------------------------------------------------------
    @property
    def zoom(self) -> float:
        return self._zoom

    def apply_zoom(self, factor: float) -> None:
        self._zoom = max(ZOOM_MIN, min(ZOOM_MAX, self._zoom * factor))
        self.setZoomFactor(self._zoom)
        self._overlay.update()

    # ------------------------------------------------------------------
    # Marker access
    # ------------------------------------------------------------------
    def add_marker(self) -> None:
        self._overlay.add_marker()

    def clear_markers(self) -> None:
        self._overlay.clear_markers()

    @property
    def markers_changed(self):
        return self._overlay.markers_changed

    # ------------------------------------------------------------------
    # Resize / show
    # ------------------------------------------------------------------
    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._overlay.sync_size()
        self._overlay.raise_()

    def showEvent(self, event):
        super().showEvent(event)
        self._overlay.sync_size()
        self._overlay.raise_()

    # ------------------------------------------------------------------
    # Mouse — pan on left drag, IO menu on right click
    # ------------------------------------------------------------------
    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self._drag_start = event.position()
            self.setCursor(Qt.CursorShape.ClosedHandCursor)
        elif event.button() == Qt.MouseButton.RightButton:
            self._show_context_menu(event)
            return
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self._drag_start is not None:
            delta = event.position() - self._drag_start
            self._drag_start = event.position()
            self.horizontalScrollBar().setValue(
                self.horizontalScrollBar().value() - int(delta.x()))
            self.verticalScrollBar().setValue(
                self.verticalScrollBar().value() - int(delta.y()))
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        self._drag_start = None
        self.setCursor(Qt.CursorShape.ArrowCursor)
        super().mouseReleaseEvent(event)

    def wheelEvent(self, event):
        if event.modifiers() & Qt.KeyboardModifier.ControlModifier:
            factor = ZOOM_STEP if event.angleDelta().y() > 0 else 1.0 / ZOOM_STEP
            self.apply_zoom(factor)
            event.accept()
        else:
            super().wheelEvent(event)

    # ------------------------------------------------------------------
    # Context menu — add new marker (right-click on empty space)
    # ------------------------------------------------------------------
    def _show_context_menu(self, event):
        click_pos  = event.position()
        global_pos = event.globalPosition().toPoint()

        # If a marker was hit the overlay already handled it — skip
        if self._overlay._marker_at(click_pos):
            return

        # ── Text mode: open text dialog immediately ───────────────────────
        if self._input_mode == "text":
            pdf_x, pdf_y, page = self._widget_pos_to_pdf(click_pos)
            dlg = TextCommentEditDialog("", page, self)
            if dlg.exec() == QDialog.DialogCode.Accepted and not dlg.delete_requested:
                self.text_registered.emit(dlg.comment_text, page, pdf_x, pdf_y)
            return

        # ── Marker mode: signal composition cascade ─────────────────────────
        menu = QMenu(self)

        # ── Paste option ────────────────────────────────────────────────────
        if self._overlay._clipboard:
            paste_act = menu.addAction("📌  Paste here  Ctrl+V")
            paste_act.setData("__paste__")
            menu.addSeparator()

       # ── Signal Compositions Section ──────────────────────────────────────────

        default_owner_id = db_ensure_default_owner()
        default_compositions = db_load_compositions_by_owner(default_owner_id)

        # Get project compositions for current PDF
        project_groups = _get_projects_for_pdf(self._pdf_path) if self._pdf_path else []

        has_any_composition = bool(default_compositions) or any(
            db_load_compositions_by_owner(pg["owner_id"]) for pg in project_groups
        )

        if not has_any_composition:
            no_comp_act = menu.addAction("(No typicals configured)")
            no_comp_act.setEnabled(False)
            menu.addSeparator()
        else:
            # ── Default submenu ────────────────────────────────────────────────
            if default_compositions:
                default_submenu = QMenu("Default", menu)

                for comp in default_compositions:
                    composition_text = _get_signal_composition(comp)
                    display_text = comp['title']
                    if composition_text:
                        display_text += f"  ({composition_text})"
                    act = default_submenu.addAction(display_text)
                    act.setData({
                        "is_composition": True,
                        "composition": comp
                    })

                menu.addMenu(default_submenu)

            # ── Project submenus ───────────────────────────────────────────────
            for pg in project_groups:
                project_compositions = db_load_compositions_by_owner(pg["owner_id"])
                if not project_compositions:
                    continue

                proj_submenu = QMenu(pg["project_name"], menu)

                for comp in project_compositions:
                    composition_text = _get_signal_composition(comp)
                    display_text = comp['title']
                    if composition_text:
                        display_text += f"  ({composition_text})"
                    act = proj_submenu.addAction(display_text)
                    act.setData({
                        "is_composition": True,
                        "composition": comp
                    })

                menu.addMenu(proj_submenu)

            menu.addSeparator()

        # ── Custom option ───────────────────────────────────────────────────
        custom_action = menu.addAction("Custom...")
        custom_action.setData(None)

        action = menu.exec(global_pos)
        if not action:
            return

        data = action.data()

        # Skip disabled header actions (old flat style remnants)
        if not action.isEnabled():
            return

        # ── Handle paste ─────────────────────────────────────────────────────
        if data == "__paste__":
            pdf_x, pdf_y, page = self._widget_pos_to_pdf(click_pos)
            self._overlay._paste_marker_at(pdf_x, pdf_y, page)
            return

        # ── Handle composition selection ─────────────────────────────────────
        if isinstance(data, dict) and data.get("is_composition"):
            composition = data["composition"]
            dlg = CompositionPlacementDialog(composition, self)
            if dlg.exec() != QDialog.DialogCode.Accepted:
                return
            
            tag_parts = dlg.tag_parts
            pdf_x, pdf_y, page = self._widget_pos_to_pdf(click_pos)
            
            # Emit signal to create composition marker
            self.composition_registered.emit(
                composition["id"], page, pdf_x, pdf_y, composition, tag_parts)
            return

        # ── Handle custom signal type ────────────────────────────────────────
        if data is None:
            text, ok = QInputDialog.getText(
                self, "Custom signal type",
                "Enter label (e.g. DI) or label - comment:",
                QLineEdit.EchoMode.Normal)
            if not ok or not text.strip():
                return
            parts = text.strip().split("-", 1)
            label                = parts[0].strip()
            comment              = parts[1].strip() if len(parts) > 1 else ""
            signal_type          = ""
            signal_type_comment  = ""
        else:
            # Should not reach here
            return

        # ── Count + description popup (for custom signals) ────────────────────
        count_dlg = MarkerCountDialog(label, comment, self)
        if count_dlg.exec() != QDialog.DialogCode.Accepted:
            return
        count       = count_dlg.count
        description = " | ".join(count_dlg.descriptions)

        pdf_x, pdf_y, page = self._widget_pos_to_pdf(click_pos)
        content_pt = QPointF(
            click_pos.x() + self.horizontalScrollBar().value(),
            click_pos.y() + self.verticalScrollBar().value(),
        )

        # For custom signals, build display with count
        display_type = f"{count}{label}" if count > 1 else label

        self.io_registered.emit(
            display_type, content_pt, page, pdf_x, pdf_y, comment, count, description,
            signal_type, signal_type_comment)
    
    @staticmethod
    def _dpi_scale() -> float:
        """
        QPdfView renders pages at the screen's logical DPI, not at 72 DPI.
        PDF points are defined as 1pt = 1/72 inch.
        So 1 PDF point occupies (screen_dpi / 72) pixels at zoom=1.
        On a typical 96 DPI monitor this is 96/72 = 1.333.
        """
        screen = QApplication.primaryScreen()
        if screen is None:
            return 1.0
        return screen.logicalDotsPerInch() / 72.0

    def _effective_scale(self) -> float:
        """Pixels per PDF point at current zoom: dpi_scale * user_zoom."""
        return self._dpi_scale() * self._zoom

    def _page_at_content_y(self, content_y: float) -> int:
        """Return the 0-indexed page that contains the given content-space Y."""
        doc = self.document()
        if doc is None or doc.pageCount() == 0:
            return 0
        eff     = self._effective_scale()
        margins = self.documentMargins()
        acc     = float(margins.top())
        for idx in range(doc.pageCount()):
            h = doc.pagePointSize(idx).height() * eff
            if content_y <= acc + h:
                return idx
            acc += h + margins.bottom() + float(margins.top())
        return doc.pageCount() - 1

    def _widget_pos_to_pdf(self, widget_pos: QPointF) -> tuple[float, float, int]:
        """
        Convert a viewport-pixel click position to PDF user-space points.

        Key insight: QPdfView renders at screen DPI (e.g. 96), not 72 DPI.
        pagePointSize() returns page dimensions in PDF points (1pt = 1/72 in).
        So the rendered pixel size of a page is:
            page_px = page_pts * (screen_dpi / 72) * zoom
                    = page_pts * effective_scale

        Inverting: page_pts = page_px / effective_scale
        """
        doc = self.document()
        if doc is None or doc.pageCount() == 0:
            return 0.0, 0.0, 0

        eff     = self._effective_scale()
        sx      = self.horizontalScrollBar().value()
        sy      = self.verticalScrollBar().value()
        margins = self.documentMargins()
        vp_w    = self.viewport().width()

        abs_x = widget_pos.x() + sx
        abs_y = widget_pos.y() + sy

        acc_y = float(margins.top())
        for page_idx in range(doc.pageCount()):
            page_pts  = doc.pagePointSize(page_idx)
            page_h_px = page_pts.height() * eff
            page_w_px = page_pts.width()  * eff

            if abs_y <= acc_y + page_h_px:
                page_x0    = max(float(margins.left()),
                                 (vp_w - page_w_px) / 2.0)
                local_x_px = abs_x - page_x0
                local_y_px = abs_y - acc_y

                pdf_x = local_x_px / eff
                pdf_y = page_pts.height() - (local_y_px / eff)

                pdf_x = max(0.0, min(float(page_pts.width()),  pdf_x))
                pdf_y = max(0.0, min(float(page_pts.height()), pdf_y))
                return round(pdf_x, 3), round(pdf_y, 3), page_idx

            acc_y += page_h_px + margins.bottom() + float(margins.top())

        last = doc.pageCount() - 1
        return 0.0, 0.0, last

    def _pdf_to_widget_pos(self, pdf_x: float, pdf_y: float,
                            page_idx: int) -> QPointF:
        """Reverse of _widget_pos_to_pdf — PDF points → viewport screen coords."""
        doc = self.document()
        if doc is None:
            return QPointF()

        eff     = self._effective_scale()
        sx      = self.horizontalScrollBar().value()
        sy      = self.verticalScrollBar().value()
        margins = self.documentMargins()
        vp_w    = self.viewport().width()

        acc_y = float(margins.top())
        for idx in range(page_idx):
            h = doc.pagePointSize(idx).height() * eff
            acc_y += h + margins.bottom() + float(margins.top())

        page_pts  = doc.pagePointSize(page_idx)
        page_w_px = page_pts.width() * eff
        page_x0   = max(float(margins.left()), (vp_w - page_w_px) / 2.0)

        screen_x = page_x0 + pdf_x * eff - sx
        screen_y = acc_y + (page_pts.height() - pdf_y) * eff - sy
        return QPointF(screen_x, screen_y)


# ---------------------------------------------------------------------------
# FDF writer
# ---------------------------------------------------------------------------
def _write_fdf(fdf_path: str, pdf_path: str, markers: list[dict]) -> None:
    """
    Write a well-formed FDF 1.2 file.

    Two annotation subtypes are used to match the PDF export:
      - Text comments   → /Subtype /FreeText  with /IT /FreeTextTypewriter
                          (no border, no background — renders as Typewriter text)
      - IO markers      → /Subtype /Text  (sticky-note icon)

    Structure:
      %FDF-1.2 header
      obj 1  — FDF catalog  (references pdf path + annotation list)
      obj 2+ — one annotation per marker
      xref table + trailer  (required by Acrobat for reliable parsing)

    Coordinates are in PDF user-space points, origin bottom-left.
    Each marker must already have pdf_x / pdf_y populated.
    """
    HALF = 12          # half-side of sticky-note Rect in PDF points
    NL   = "\r\n"      # FDF spec recommends CRLF
    FONT_SIZE = 10     # pt — used for FreeText rect estimation

    def _pdf_escape(s: str) -> str:
        """Escape a PDF string literal (parentheses and backslash)."""
        return s.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")

    # ── build annotation objects ─────────────────────────────────────────
    objects: list[bytes] = []
    objects.append(b"")         # placeholder for catalog (obj 1)
    annot_refs: list[str] = []

    for i, m in enumerate(markers, start=2):
        pdf_x = m.get("pdf_x", 0.0)
        pdf_y = m.get("pdf_y", 0.0)
        pg    = int(m["page"])
        nm    = str(uuid.uuid4())
        kind  = m.get("kind", "marker")

        if kind == "text":
            # ── FreeText / Typewriter ────────────────────────────────────
            text    = m.get("type", "")
            lines   = text.split("\n") if text else [""]
            char_w  = FONT_SIZE * 0.55
            line_h  = FONT_SIZE * 1.4
            rect_w  = max(40.0, max(len(l) for l in lines) * char_w)
            rect_h  = max(line_h, len(lines) * line_h)
            x1 = round(pdf_x, 3)
            y1 = round(pdf_y - rect_h, 3)
            x2 = round(pdf_x + rect_w, 3)
            y2 = round(pdf_y, 3)
            # Default Appearance string: Helvetica 10pt black
            da = "/Helvetica 10 Tf 0 0 0 rg"
            obj_str = (
                f"{i} 0 obj{NL}"
                f"<< /Type /Annot{NL}"
                f"   /Subtype /FreeText{NL}"
                f"   /IT /FreeTextTypewriter{NL}"
                f"   /Subj (Typewriter){NL}"
                f"   /Rect [{x1} {y1} {x2} {y2}]{NL}"
                f"   /Contents ({_pdf_escape(text)}){NL}"
                f"   /DA ({_pdf_escape(da)}){NL}"
                f"   /Q 0{NL}"
                f"   /T (Technical Drawing Viewer){NL}"
                f"   /Page {pg}{NL}"
                f"   /NM ({nm}){NL}"
                f"   /F 4{NL}"
                f"   /BS << /W 0 >>{NL}"
                f">>{NL}"
                f"endobj{NL}"
            )
        else:
            # ── Sticky-note Text annotation for IO markers ───────────────
            count   = int(m.get("count", 1))
            # Composition markers store the fully-expanded label in m['type']
            # (e.g. "4HDI 2HDO" when count=2 on a "2HDI 1HDO" composition),
            # so never prepend count again.  Simple markers do need the prefix.
            if m.get("is_composition"):
                label = m["type"]
            else:
                label = f"{count}{m['type']}" if count > 1 else m["type"]
            comment = m.get("comment", "")
            desc    = m.get("description", "")
            parts   = [label]
            if comment:
                parts.append(comment)
            if desc:
                parts.append(desc)
            contents = "\n".join(parts)
            x1 = round(pdf_x - HALF, 3)
            y1 = round(pdf_y - HALF, 3)
            x2 = round(pdf_x + HALF, 3)
            y2 = round(pdf_y + HALF, 3)
            obj_str = (
                f"{i} 0 obj{NL}"
                f"<< /Type /Annot{NL}"
                f"   /Subtype /Text{NL}"
                f"   /Rect [{x1} {y1} {x2} {y2}]{NL}"
                f"   /Contents ({_pdf_escape(contents)}){NL}"
                f"   /T (Technical Drawing Viewer){NL}"
                f"   /Page {pg}{NL}"
                f"   /NM ({nm}){NL}"
                f"   /F 4{NL}"
                f"   /Open false{NL}"
                f">>{NL}"
                f"endobj{NL}"
            )

        objects.append(obj_str.encode("latin-1", errors="replace"))
        annot_refs.append(f"{i} 0 R")

    # ── catalog object (obj 1) ────────────────────────────────────────────
    pdf_fdf_path = pdf_path.replace("\\", "/")
    annots_str   = " ".join(annot_refs)
    catalog_str  = (
        f"1 0 obj{NL}"
        f"<< /FDF{NL}"
        f"   << /Annots [{annots_str}]{NL}"
        f"      /F ({pdf_fdf_path}){NL}"
        f"   >>{NL}"
        f">>{NL}"
        f"endobj{NL}"
    )
    objects[0] = catalog_str.encode("latin-1")

    # ── assemble file, recording byte offsets for xref ───────────────────
    header  = b"%FDF-1.2" + NL.encode() + b"%\xe2\xe3\xcf\xd3" + NL.encode()
    body    = b""
    offsets = []

    for obj_bytes in objects:
        offsets.append(len(header) + len(body))
        body += obj_bytes

    # ── xref table ───────────────────────────────────────────────────────
    xref_offset = len(header) + len(body)
    n_objs      = len(objects) + 1   # include object 0 (free head)

    xref_lines  = [f"xref{NL}0 {n_objs}{NL}"]
    xref_lines.append(f"0000000000 65535 f {NL}")
    for off in offsets:
        xref_lines.append(f"{off:010d} 00000 n {NL}")
    xref_str = "".join(xref_lines)

    trailer_str = (
        f"trailer{NL}"
        f"<< /Size {n_objs}{NL}"
        f"   /Root 1 0 R{NL}"
        f">>{NL}"
        f"startxref{NL}"
        f"{xref_offset}{NL}"
        f"%%EOF{NL}"
    )

    with open(fdf_path, "wb") as f:
        f.write(header)
        f.write(body)
        f.write(xref_str.encode("latin-1"))
        f.write(trailer_str.encode("latin-1"))


# ---------------------------------------------------------------------------
# SQLite persistence
# ---------------------------------------------------------------------------
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "pid_markers.db")

# ---------------------------------------------------------------------------
# PDF fingerprinting
# ---------------------------------------------------------------------------
_PARTIAL_HASH_BYTES = 512 * 1024   # first 512 KB

def _pdf_partial_hash(path: str) -> str:
    """SHA-256 of the first 512 KB of the file."""
    h = hashlib.sha256()
    with open(path, "rb") as f:
        h.update(f.read(_PARTIAL_HASH_BYTES))
    return h.hexdigest()

def _pdf_internal_id(path: str) -> str | None:
    """
    Extract the /ID array from the PDF trailer.
    Returns the first element as a hex string, or None if not present.
    """
    try:
        reader = PdfReader(path)
        trailer = reader.trailer
        if "/ID" in trailer:
            raw = trailer["/ID"][0]
            if isinstance(raw, bytes):
                return raw.hex()
            return str(raw)
    except Exception:
        pass
    return None

def _pdf_fingerprint(path: str) -> tuple[str | None, str]:
    """Return (pdf_id, partial_hash) for a given PDF path."""
    return _pdf_internal_id(path), _pdf_partial_hash(path)


# ---------------------------------------------------------------------------
# Core DB connection
# ---------------------------------------------------------------------------
def _db_connect() -> sqlite3.Connection:
    con = sqlite3.connect(DB_PATH)
    con.execute("PRAGMA foreign_keys = ON")
    
    # ── Clean up any orphaned migration tables ──────────────────────────────
    try:
        # If signal_compositions_old exists but signal_compositions also exists,
        # it means a previous migration was incomplete - clean it up
        old_exists = con.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='signal_compositions_old'"
        ).fetchone()
        
        new_exists = con.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='signal_compositions'"
        ).fetchone()
        
        if old_exists and new_exists:
            print("Cleaning up orphaned migration table...")
            con.execute("DROP TABLE IF EXISTS signal_compositions_old")
            con.commit()
    except Exception as e:
        print(f"Cleanup error (non-fatal): {e}")
    
    # ── Check and migrate signal_compositions table if needed ──────────────
    try:
        table_exists = con.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='signal_compositions'"
        ).fetchone()

        if table_exists:
            schema_query = con.execute(
                "SELECT sql FROM sqlite_master WHERE type='table' AND name='signal_compositions'"
            ).fetchone()

            if schema_query and "UNIQUE" in schema_query[0]:
                print("Migrating signal_compositions table to allow duplicate titles...")
                try:
                    # MUST disable FK enforcement before renaming
                    # (SQLite does not support ALTER TABLE DROP CONSTRAINT)
                    con.execute("PRAGMA foreign_keys = OFF")
                    con.execute("BEGIN TRANSACTION")

                    # Rename old table
                    con.execute("ALTER TABLE signal_compositions RENAME TO signal_compositions_old")

                    # Create new table WITHOUT UNIQUE constraint
                    con.execute("""
                        CREATE TABLE signal_compositions (
                            id              INTEGER PRIMARY KEY AUTOINCREMENT,
                            title           TEXT    NOT NULL,
                            description     TEXT    NOT NULL DEFAULT '',
                            control_module  TEXT    NOT NULL DEFAULT 'NA',
                            field_device    TEXT    NOT NULL DEFAULT 'NA',
                            created         TEXT    NOT NULL,
                            modified        TEXT    NOT NULL
                        )
                    """)

                    # Copy data
                    con.execute("""
                        INSERT INTO signal_compositions (id, title, description, created, modified)
                        SELECT id, title, description, created, modified
                        FROM signal_compositions_old
                    """)

                    # Rebuild signal_composition_signals WITHOUT the broken FK reference
                    # (SQLite can't update FK targets; rebuild the child table too)
                    con.execute("""
                        CREATE TABLE signal_composition_signals_new (
                            id                 INTEGER PRIMARY KEY AUTOINCREMENT,
                            composition_id     INTEGER NOT NULL
                                                REFERENCES signal_compositions(id)
                                                ON DELETE CASCADE,
                            signal_name        TEXT    NOT NULL,
                            signal_type        TEXT    NOT NULL,
                            signal_description TEXT    NOT NULL DEFAULT '',
                            prefix             TEXT    NOT NULL DEFAULT 'NA',
                            suffix             TEXT    NOT NULL DEFAULT 'NA',
                            sort_order         INTEGER NOT NULL DEFAULT 0
                        )
                    """)
                    con.execute("""
                        INSERT INTO signal_composition_signals_new
                            (id, composition_id, signal_name, signal_type, signal_description, sort_order)
                        SELECT id, composition_id, signal_name, signal_type, signal_description, sort_order
                        FROM signal_composition_signals
                    """)
                    con.execute("DROP TABLE signal_composition_signals")
                    con.execute("ALTER TABLE signal_composition_signals_new RENAME TO signal_composition_signals")

                    # Drop the backup
                    con.execute("DROP TABLE signal_compositions_old")

                    con.execute("COMMIT")
                    con.execute("PRAGMA foreign_keys = ON")
                    print("✓ Migration complete - signal_compositions table updated")

                except Exception as e:
                    con.execute("ROLLBACK")
                    con.execute("PRAGMA foreign_keys = ON")
                    print(f"Migration error: {e}")
        else:
            # Table doesn't exist yet — create fresh
            con.execute("""
                CREATE TABLE IF NOT EXISTS signal_compositions (
                    id              INTEGER PRIMARY KEY AUTOINCREMENT,
                    title           TEXT    NOT NULL,
                    description     TEXT    NOT NULL DEFAULT '',
                    control_module  TEXT    NOT NULL DEFAULT 'NA',
                    field_device    TEXT    NOT NULL DEFAULT 'NA',
                    created         TEXT    NOT NULL,
                    modified        TEXT    NOT NULL
                )
            """)
    except Exception as e:
        print(f"Signal compositions table setup error: {e}")
        con.execute("""
            CREATE TABLE IF NOT EXISTS signal_compositions (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                title           TEXT    NOT NULL,
                description     TEXT    NOT NULL DEFAULT '',
                control_module  TEXT    NOT NULL DEFAULT 'NA',
                field     TEXT    NOT NULL DEFAULT 'NA',
                created         TEXT    NOT NULL,
                modified        TEXT    NOT NULL
            )
        """)
    
    # Final safety check - ensure no orphaned tables exist
    try:
        orphan_check = con.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='signal_compositions_old'"
        ).fetchone()
        if orphan_check:
            con.execute("DROP TABLE IF EXISTS signal_compositions_old")
            con.commit()
    except:
        pass

    # ── Migrate signal_compositions: add control_module / field_device / extended columns ──
    try:
        existing_cols = [
            row[1]
            for row in con.execute("PRAGMA table_info(signal_compositions)").fetchall()
        ]
        if "control_module" not in existing_cols:
            con.execute(
                "ALTER TABLE signal_compositions "
                "ADD COLUMN control_module TEXT NOT NULL DEFAULT 'NA'"
            )
            con.commit()
        if "field_device" not in existing_cols:
            con.execute(
                "ALTER TABLE signal_compositions "
                "ADD COLUMN field_device TEXT NOT NULL DEFAULT 'NA'"
            )
            con.commit()
        # Extended fields: control module type/description, field device type/description
        for col_def in (
            ("cm_type",        "TEXT NOT NULL DEFAULT 'NA'"),
            ("cm_description", "TEXT NOT NULL DEFAULT 'NA'"),
            ("fd_type",        "TEXT NOT NULL DEFAULT 'NA'"),
            ("fd_description", "TEXT NOT NULL DEFAULT 'NA'"),
            ("category",       "TEXT NOT NULL DEFAULT ''"),
        ):
            if col_def[0] not in existing_cols:
                con.execute(
                    f"ALTER TABLE signal_compositions ADD COLUMN {col_def[0]} {col_def[1]}"
                )
                con.commit()
    except Exception as e:
        print(f"signal_compositions column migration error (non-fatal): {e}")

    # ── Migrate signal_composition_signals: add prefix / suffix / extra_column_values columns ──
    try:
        existing_sig_cols = [
            row[1]
            for row in con.execute("PRAGMA table_info(signal_composition_signals)").fetchall()
        ]
        if "prefix" not in existing_sig_cols:
            con.execute(
                "ALTER TABLE signal_composition_signals "
                "ADD COLUMN prefix TEXT NOT NULL DEFAULT 'NA'"
            )
            con.commit()
        if "suffix" not in existing_sig_cols:
            con.execute(
                "ALTER TABLE signal_composition_signals "
                "ADD COLUMN suffix TEXT NOT NULL DEFAULT 'NA'"
            )
            con.commit()
        if "extra_column_values" not in existing_sig_cols:
            con.execute(
                "ALTER TABLE signal_composition_signals "
                "ADD COLUMN extra_column_values TEXT NOT NULL DEFAULT '[]'"
            )
            con.commit()
        if "count" not in existing_sig_cols:
            con.execute(
                "ALTER TABLE signal_composition_signals "
                "ADD COLUMN count INTEGER NOT NULL DEFAULT 1"
            )
            con.commit()
    except Exception as e:
        print(f"signal_composition_signals column migration error (non-fatal): {e}")

    # ── Migrate signal_compositions: add extra_column_headers column ──────
    try:
        existing_comp_cols = [
            row[1]
            for row in con.execute("PRAGMA table_info(signal_compositions)").fetchall()
        ]
        if "extra_column_headers" not in existing_comp_cols:
            con.execute(
                "ALTER TABLE signal_compositions "
                "ADD COLUMN extra_column_headers TEXT NOT NULL DEFAULT '[]'"
            )
            con.commit()
    except Exception as e:
        print(f"signal_compositions extra_column_headers migration error (non-fatal): {e}")

    # ── Other tables (these should be fine) ───────────────────────────────
    con.execute("""
        CREATE TABLE IF NOT EXISTS signal_composition_signals (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            composition_id  INTEGER NOT NULL REFERENCES signal_compositions(id) ON DELETE CASCADE,
            signal_name     TEXT    NOT NULL,
            signal_type     TEXT    NOT NULL,
            signal_description TEXT NOT NULL DEFAULT '',
            prefix          TEXT    NOT NULL DEFAULT 'NA',
            suffix          TEXT    NOT NULL DEFAULT 'NA',
            extra_column_values TEXT NOT NULL DEFAULT '[]',
            count           INTEGER NOT NULL DEFAULT 1,
            sort_order      INTEGER NOT NULL DEFAULT 0
        )
    """)
    
    con.execute("""
        CREATE TABLE IF NOT EXISTS composition_owners (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            owner_name  TEXT    NOT NULL UNIQUE,
            owner_type  TEXT    NOT NULL,
            project_id  INTEGER DEFAULT NULL
        )
    """)
    
    con.execute("""
        CREATE TABLE IF NOT EXISTS composition_ownership (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            composition_id  INTEGER NOT NULL REFERENCES signal_compositions(id) ON DELETE CASCADE,
            owner_id        INTEGER NOT NULL REFERENCES composition_owners(id) ON DELETE CASCADE,
            sort_order      INTEGER NOT NULL DEFAULT 0
        )
    """)
    
    con.commit()
    
    # ── Rest of the original function (markers table, etc.) ────────────────
    con.execute("""
        CREATE TABLE IF NOT EXISTS markers (
            id                   INTEGER PRIMARY KEY AUTOINCREMENT,
            pdf                  TEXT    NOT NULL,
            type                 TEXT    NOT NULL,
            comment              TEXT    NOT NULL DEFAULT '',
            page                 INTEGER NOT NULL,
            pdf_x                REAL    NOT NULL,
            pdf_y                REAL    NOT NULL,
            kind                 TEXT    NOT NULL DEFAULT 'marker',
            count                INTEGER NOT NULL DEFAULT 1,
            description          TEXT    NOT NULL DEFAULT '',
            project_file_id      INTEGER REFERENCES project_files(id) ON DELETE SET NULL,
            signal_type          TEXT    NOT NULL DEFAULT '',
            signal_type_comment  TEXT    NOT NULL DEFAULT ''
        )
    """)
        
    # Migrate existing tables that may be missing newer columns
    for col, dflt in (
        ("comment",              "''"),
        ("kind",                 "'marker'"),
        ("count",                "1"),
        ("description",          "''"),
        ("project_file_id",      "NULL"),
        ("signal_type",          "''"),
        ("signal_type_comment",  "''"),
    ):
        try:
            col_type = "INTEGER" if col in ("count", "project_file_id") else "TEXT"
            null_clause = "" if col == "project_file_id" else f"NOT NULL DEFAULT {dflt}"
            con.execute(f"ALTER TABLE markers ADD COLUMN {col} {col_type} {null_clause}")
        except Exception:
            pass
    
    # Migrate complex object columns
    for col, dflt in (
        ("base_tag",            "''"),
        ("complex_object_id",   "NULL"),
        ("is_complex",          "0"),
        ("complete_tag",        "''"),  # NEW: Complete tag with all parts
        ("tag_parts",           "'{}'"),  # NEW: JSON object for tag parts
    ):
        try:
            col_type = "INTEGER" if col in ("complex_object_id", "is_complex") else "TEXT"
            null_clause = "" if col in ("complex_object_id",) else f"NOT NULL DEFAULT {dflt}"
            con.execute(f"ALTER TABLE markers ADD COLUMN {col} {col_type} {null_clause}")
        except Exception:
            pass
    
    # Migrate composition columns
    for col, dflt in (
        ("composition_id", "NULL"),
        ("is_composition", "0"),
    ):
        try:
            col_type = "INTEGER" if col == "composition_id" else "INTEGER"
            con.execute(f"ALTER TABLE markers ADD COLUMN {col} {col_type} DEFAULT {dflt}")
        except Exception:
            pass
     
    # Extend projects table with drawing metadata columns
    for col in ("identifier", "drw_name", "area"):
        try:
            con.execute(f"ALTER TABLE projects ADD COLUMN {col} TEXT NOT NULL DEFAULT ''")
        except Exception:
            pass

    # View: project → file → IO markers (excludes text comments)
    con.execute("""
        CREATE VIEW IF NOT EXISTS project_marker_view AS
        SELECT
            p.id            AS project_id,
            p.name          AS project_name,
            p.number        AS project_number,
            p.identifier    AS identifier,
            p.drw_name      AS drw_name,
            p.area          AS area,
            pf.id           AS project_file_id,
            pf.file_path    AS file_path,
            pf.file_name    AS file_name,
            m.id            AS marker_id,
            m.type          AS signal_type,
            m.comment       AS signal_comment,
            m.count         AS count,
            m.description   AS description,
            m.page          AS page,
            m.pdf_x         AS pdf_x,
            m.pdf_y         AS pdf_y
        FROM projects p
        JOIN project_files pf ON pf.project_id = p.id
        JOIN markers m        ON m.project_file_id = pf.id
        WHERE m.kind != 'text'
        ORDER BY p.name, pf.sort_order, m.page, m.type
    """)

    con.commit()
    return con


# ---------------------------------------------------------------------------
# Settings DB helpers  (theme persistence)
# ---------------------------------------------------------------------------
def _ensure_settings_table(con: sqlite3.Connection) -> None:
    con.execute("""
        CREATE TABLE IF NOT EXISTS app_settings (
            key   TEXT PRIMARY KEY,
            value TEXT NOT NULL
        )
    """)
    con.commit()


def db_load_theme() -> str:
    """Return the saved theme name ('dark', 'light', or 'system'). Defaults to 'dark'."""
    with _db_connect() as con:
        _ensure_settings_table(con)
        row = con.execute(
            "SELECT value FROM app_settings WHERE key = 'theme'"
        ).fetchone()
    return row[0] if row else "dark"


def db_save_theme(theme: str) -> None:
    """Persist the active theme name."""
    with _db_connect() as con:
        _ensure_settings_table(con)
        con.execute(
            "INSERT INTO app_settings (key, value) VALUES ('theme', ?) "
            "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
            (theme,))
        con.commit()


# ---------------------------------------------------------------------------
# Signal Typical Templates DB helpers
# ---------------------------------------------------------------------------
def _ensure_templates_table(con: sqlite3.Connection) -> None:
    con.execute("""
        CREATE TABLE IF NOT EXISTS signal_typical_templates (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            title       TEXT    NOT NULL,
            description TEXT    NOT NULL DEFAULT '',
            created     TEXT    NOT NULL,
            modified    TEXT    NOT NULL
        )
    """)
    con.execute("""
        CREATE TABLE IF NOT EXISTS signal_typical_template_signals (
            id                 INTEGER PRIMARY KEY AUTOINCREMENT,
            template_id        INTEGER NOT NULL
                               REFERENCES signal_typical_templates(id) ON DELETE CASCADE,
            signal_name        TEXT    NOT NULL,
            signal_type        TEXT    NOT NULL,
            signal_description TEXT    NOT NULL DEFAULT '',
            sort_order         INTEGER NOT NULL DEFAULT 0
        )
    """)
    con.commit()


def db_load_all_templates() -> dict:
    """Return all signal typical templates keyed by id."""
    with _db_connect() as con:
        _ensure_templates_table(con)
        rows = con.execute(
            "SELECT id, title, description FROM signal_typical_templates ORDER BY id"
        ).fetchall()
        result = {}
        for (tid, title, desc) in rows:
            sigs = con.execute(
                "SELECT signal_name, signal_type, signal_description "
                "FROM signal_typical_template_signals "
                "WHERE template_id=? ORDER BY sort_order",
                (tid,)).fetchall()
            result[tid] = {
                "id":          tid,
                "title":       title,
                "description": desc,
                "signals": [
                    {"signal_name": s[0], "signal_type": s[1],
                     "signal_description": s[2]}
                    for s in sigs
                ],
            }
    return result


def db_save_new_template(title: str, description: str,
                         signals: list[dict]) -> int:
    """Insert a new template and return its id."""
    from datetime import datetime
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    with _db_connect() as con:
        _ensure_templates_table(con)
        cur = con.execute(
            "INSERT INTO signal_typical_templates (title, description, created, modified) "
            "VALUES (?, ?, ?, ?)",
            (title, description or "", now, now))
        tid = cur.lastrowid
        for order, sig in enumerate(signals):
            con.execute(
                "INSERT INTO signal_typical_template_signals "
                "(template_id, signal_name, signal_type, signal_description, sort_order) "
                "VALUES (?, ?, ?, ?, ?)",
                (tid, sig["signal_name"], sig["signal_type"],
                 sig.get("signal_description", ""), order))
        con.commit()
    return tid


def db_update_template(template_id: int, title: str, description: str,
                       signals: list[dict]) -> None:
    """Replace an existing template's data."""
    from datetime import datetime
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    with _db_connect() as con:
        _ensure_templates_table(con)
        con.execute(
            "UPDATE signal_typical_templates "
            "SET title=?, description=?, modified=? WHERE id=?",
            (title, description or "", now, template_id))
        con.execute(
            "DELETE FROM signal_typical_template_signals WHERE template_id=?",
            (template_id,))
        for order, sig in enumerate(signals):
            con.execute(
                "INSERT INTO signal_typical_template_signals "
                "(template_id, signal_name, signal_type, signal_description, sort_order) "
                "VALUES (?, ?, ?, ?, ?)",
                (template_id, sig["signal_name"], sig["signal_type"],
                 sig.get("signal_description", ""), order))
        con.commit()


def db_delete_template(template_id: int) -> None:
    """Delete a template and all its signals."""
    with _db_connect() as con:
        _ensure_templates_table(con)
        con.execute(
            "DELETE FROM signal_typical_templates WHERE id=?", (template_id,))
        con.commit()


# ---------------------------------------------------------------------------
# Session DB helpers
# ---------------------------------------------------------------------------
def _ensure_sessions_table(con: sqlite3.Connection) -> None:
    con.execute("""
        CREATE TABLE IF NOT EXISTS pdf_sessions (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            pdf_path      TEXT    NOT NULL UNIQUE,
            pdf_filename  TEXT    NOT NULL,
            pdf_id        TEXT,
            partial_hash  TEXT    NOT NULL,
            last_opened   TEXT    NOT NULL,
            marker_count  INTEGER NOT NULL DEFAULT 0
        )
    """)
    con.commit()

def db_upsert_session(pdf_path: str, pdf_id: str | None,
                      partial_hash: str, marker_count: int) -> None:
    """Insert or update the session record for the given path."""
    from datetime import datetime
    filename = os.path.basename(pdf_path)
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    with _db_connect() as con:
        _ensure_sessions_table(con)
        con.execute("""
            INSERT INTO pdf_sessions
                (pdf_path, pdf_filename, pdf_id, partial_hash, last_opened, marker_count)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(pdf_path) DO UPDATE SET
                pdf_filename = excluded.pdf_filename,
                pdf_id       = excluded.pdf_id,
                partial_hash = excluded.partial_hash,
                last_opened  = excluded.last_opened,
                marker_count = excluded.marker_count
        """, (pdf_path, filename, pdf_id, partial_hash, now, marker_count))
        con.commit()

def db_find_matching_sessions(pdf_path: str,
                               pdf_id: str | None,
                               partial_hash: str) -> list[dict]:
    """
    Find sessions that share /ID or partial_hash with the given fingerprint,
    but have a DIFFERENT path (i.e. the file was renamed/moved).
    Returns a list of match dicts, each with a 'match_reason' field.
    """
    with _db_connect() as con:
        _ensure_sessions_table(con)
        rows = con.execute("""
            SELECT id, pdf_path, pdf_filename, pdf_id, partial_hash,
                   last_opened, marker_count
            FROM pdf_sessions
            WHERE pdf_path != ?
              AND partial_hash != ''
        """, (pdf_path,)).fetchall()

    matches = []
    seen_ids = set()
    for row in rows:
        sid, old_path, old_name, old_pid, old_hash, last_opened, mcount = row
        reasons = []
        if pdf_id and old_pid and pdf_id == old_pid:
            reasons.append("PDF internal ID matched")
        if partial_hash == old_hash:
            reasons.append("file content matched")
        if reasons and sid not in seen_ids:
            seen_ids.add(sid)
            matches.append({
                "session_id":   sid,
                "old_path":     old_path,
                "old_filename": old_name,
                "pdf_id":       old_pid,
                "partial_hash": old_hash,
                "last_opened":  last_opened,
                "marker_count": mcount,
                "match_reason": " + ".join(reasons),
            })
    return matches

def db_migrate_session(old_path: str, new_path: str,
                       pdf_id: str | None, partial_hash: str) -> int:
    """
    Move all markers from old_path to new_path.
    Update the session record in-place.
    Returns number of markers migrated.
    """
    from datetime import datetime
    new_name = os.path.basename(new_path)
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    with _db_connect() as con:
        _ensure_sessions_table(con)
        # Remove any stub session row that was auto-created for new_path
        # (by _on_markers_changed fallback upsert) — this prevents the
        # UNIQUE constraint from firing when we rename the old row.
        con.execute("DELETE FROM pdf_sessions WHERE pdf_path = ?", (new_path,))
        # Also remove any orphaned markers already keyed to new_path
        con.execute("DELETE FROM markers WHERE pdf = ?", (new_path,))
        # Reassign markers from old path to new path
        con.execute("UPDATE markers SET pdf = ? WHERE pdf = ?",
                    (new_path, old_path))
        count = con.execute(
            "SELECT COUNT(*) FROM markers WHERE pdf = ?",
            (new_path,)).fetchone()[0]
        # Update the old session record in-place to the new path
        con.execute("""
            UPDATE pdf_sessions
            SET pdf_path     = ?,
                pdf_filename = ?,
                pdf_id       = ?,
                partial_hash = ?,
                last_opened  = ?,
                marker_count = ?
            WHERE pdf_path = ?
        """, (new_path, new_name, pdf_id, partial_hash, now, count, old_path))
        con.commit()
    return count

def db_all_sessions() -> list[dict]:
    """Return all sessions for the manual link dialog."""
    with _db_connect() as con:
        _ensure_sessions_table(con)
        rows = con.execute("""
            SELECT id, pdf_path, pdf_filename, last_opened, marker_count
            FROM pdf_sessions
            ORDER BY last_opened DESC
        """).fetchall()
    return [{"session_id": r[0], "pdf_path": r[1], "pdf_filename": r[2],
             "last_opened": r[3], "marker_count": r[4]} for r in rows]

def _build_complete_tag(tag_parts: dict, signal_name: str = "") -> str:
    """Build a complete tag from parts."""
    if not tag_parts:
        return signal_name
    
    parts = []
    prefix = tag_parts.get("prefix", "")
    middle_fields = tag_parts.get("middle_fields", [])
    suffix = tag_parts.get("suffix", "")
    
    if prefix:
        parts.append(prefix)
    if signal_name:
        parts.append(signal_name)
    if middle_fields:
        parts.extend(middle_fields)
    if suffix:
        parts.append(suffix)
    
    return "-".join(parts)

def db_save_markers(pdf_path: str, markers: list[dict]) -> None:
    """Replace all markers for this PDF with the current list."""
    
    with _db_connect() as con:
        row = con.execute(
            "SELECT id FROM project_files WHERE file_path = ? LIMIT 1",
            (pdf_path,)).fetchone()
        project_file_id = row[0] if row else None

        con.execute("DELETE FROM markers WHERE pdf = ?", (pdf_path,))
        con.executemany(
            "INSERT INTO markers "
            "(pdf, type, comment, page, pdf_x, pdf_y, kind, count, description, "
            " project_file_id, signal_type, signal_type_comment, base_tag, "
            " composition_id, is_composition, complete_tag, tag_parts) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            [(pdf_path,
            m["type"],
            m.get("comment", ""),
            m["page"],
            m["pdf_x"],
            m["pdf_y"],
            m.get("kind", "marker"),
            int(m.get("count", 1)),
            m.get("description", ""),
            project_file_id,
            m.get("signal_type", ""),
            m.get("signal_type_comment", ""),
            "",  # base_tag
            m.get("composition_id", None),
            1 if m.get("is_composition") else 0,
            _build_complete_tag(m.get("tag_parts", {}), ""),
            json.dumps(m.get("tag_parts", {})))  # ADD THIS LINE
            for m in markers]
        )
        con.commit()
        
def db_load_markers(pdf_path: str) -> list[dict]:
    """Return saved markers for this PDF, or empty list if none."""
    
    with _db_connect() as con:
        rows = con.execute(
            "SELECT type, comment, page, pdf_x, pdf_y, kind, count, description, "
            "signal_type, signal_type_comment, base_tag, composition_id, is_composition, complete_tag, tag_parts "
            "FROM markers WHERE pdf = ? ORDER BY id",
            (pdf_path,)
        ).fetchall()
    
    result = []
    for r in rows:
        # Parse tag_parts JSON
        tag_parts = {}
        if r[14]:  # tag_parts column
            try:
                tag_parts = json.loads(r[14])
            except (json.JSONDecodeError, TypeError):
                tag_parts = {}
        
        # Default structure if empty
        if not tag_parts:
            tag_parts = {
                "prefix": "",
                "middle_fields": [],
                "suffix": "",
            }
        
        marker = {
            "type": r[0],
            "comment": r[1],
            "page": r[2],
            "pdf_x": r[3],
            "pdf_y": r[4],
            "kind": r[5] or "marker",
            "count": int(r[6]) if r[6] else 1,
            "description": r[7] or "",
            "signal_type": r[8] or "",
            "signal_type_comment": r[9] or "",
            "base_tag": r[10] or "",
            "composition_id": r[11],
            "is_composition": bool(r[12]) if r[12] else False,
            "complete_tag": r[13] or "",
            "tag_parts": tag_parts,
        }
        result.append(marker)
    
    return result

def db_load_project_markers(project_id: int,
                             file_path: str | None = None) -> list[dict]:
    """
    Return all IO signal markers (not text comments) for every PDF
    registered under `project_id`, or just for `file_path` if given.

    Joins via pdf path directly — does NOT rely on project_file_id FK
    so it works for all markers regardless of when they were saved.
    """
    with _db_connect() as con:
        if file_path:
            # Normalise to forward slashes for consistent comparison
            fp = file_path.replace("\\", "/")
            rows = con.execute("""
                SELECT
                    pf.file_name,
                    pf.file_path,
                    m.page,
                    m.type,
                    m.comment,
                    m.count,
                    m.description,
                    m.pdf_x,
                    m.pdf_y,
                    m.signal_type,
                    m.signal_type_comment,
                    m.is_composition,
                    m.composition_id,
                    m.tag_parts,
                    COALESCE(pf.drw_name,''),
                    COALESCE(pf.drw_number,'')
                FROM project_files pf
                JOIN markers m
                  ON REPLACE(m.pdf, '\\', '/') = REPLACE(pf.file_path, '\\', '/')
                WHERE pf.project_id = ?
                  AND REPLACE(pf.file_path, '\\', '/') = ?
                  AND m.kind != 'text'
                ORDER BY m.page, m.type
            """, (project_id, fp)).fetchall()
        else:
            rows = con.execute("""
                SELECT
                    pf.file_name,
                    pf.file_path,
                    m.page,
                    m.type,
                    m.comment,
                    m.count,
                    m.description,
                    m.pdf_x,
                    m.pdf_y,
                    m.signal_type,
                    m.signal_type_comment,
                    m.is_composition,
                    m.composition_id,
                    m.tag_parts,
                    COALESCE(pf.drw_name,''),
                    COALESCE(pf.drw_number,'')
                FROM project_files pf
                JOIN markers m
                  ON REPLACE(m.pdf, '\\', '/') = REPLACE(pf.file_path, '\\', '/')
                WHERE pf.project_id = ?
                  AND m.kind != 'text'
                ORDER BY pf.sort_order, pf.id, m.page, m.type
            """, (project_id,)).fetchall()

    result = []
    for r in rows:
        tag_parts = {}
        if r[13]:
            try:
                tag_parts = json.loads(r[13])
            except (json.JSONDecodeError, TypeError):
                tag_parts = {}
        result.append({
            "file_name":            r[0],
            "file_path":            r[1],
            "page":                 r[2],
            "signal_type":          r[3],
            "signal_comment":       r[4],
            "count":                int(r[5]) if r[5] else 1,
            "description":          r[6] or "",
            "pdf_x":                r[7],
            "pdf_y":                r[8],
            "parent_signal_type":   r[9]  or "",
            "parent_signal_comment":r[10] or "",
            "is_composition":       bool(r[11]) if r[11] else False,
            "composition_id":       r[12],
            "tag_parts":            tag_parts,
            "drw_name":             r[14] if len(r) > 14 else "",
            "drw_number":           r[15] if len(r) > 15 else "",
        })
    return result


def db_get_project_drawing_meta(project_id: int) -> dict:
    """Return the drawing metadata stored on the project record."""
    with _db_connect() as con:
        row = con.execute(
            "SELECT name, number, description, identifier, drw_name, area "
            "FROM projects WHERE id = ?", (project_id,)).fetchone()
    if not row:
        return {"name": "", "number": "", "description": "",
                "identifier": "", "drw_name": "", "area": ""}
    return {"name": row[0], "number": row[1], "description": row[2],
            "identifier": row[3] or "", "drw_name": row[4] or "",
            "area": row[5] or ""}
    
# ─────────────────────────────────────────────────────────────────────
# Signal Composition Database Functions
# ─────────────────────────────────────────────────────────────────────

def db_ensure_default_owner() -> int:
    """
    Ensure a 'Default' owner exists in the database.
    Returns the owner_id of the Default owner.
    """
    with _db_connect() as con:
        # Check if Default owner exists
        row = con.execute(
            "SELECT id FROM composition_owners WHERE owner_name = 'Default'").fetchone()
        
        if row:
            return row[0]
        
        # Create Default owner if it doesn't exist
        cur = con.execute(
            "INSERT INTO composition_owners (owner_name, owner_type, project_id) "
            "VALUES ('Default', 'Default', NULL)")
        con.commit()
        return cur.lastrowid


def db_get_or_create_project_owner(project_id: int) -> int:
    """
    Get the owner_id for a project, creating it if necessary.
    The owner_name is the project's name from the projects table.
    
    Args:
        project_id: ID from projects table
        
    Returns:
        owner_id from composition_owners table
    """
    with _db_connect() as con:
        # Get project name
        project = con.execute(
            "SELECT name FROM projects WHERE id = ?", (project_id,)).fetchone()
        
        if not project:
            return None
        
        project_name = project[0]
        
        # Check if owner already exists
        row = con.execute(
            "SELECT id FROM composition_owners WHERE project_id = ?",
            (project_id,)).fetchone()
        
        if row:
            return row[0]
        
        # Create owner for this project
        cur = con.execute(
            "INSERT INTO composition_owners (owner_name, owner_type, project_id) "
            "VALUES (?, 'Project', ?)",
            (project_name, project_id))
        con.commit()
        return cur.lastrowid
 
def db_save_signal_composition(title: str, description: str,
                              signals: list[dict],
                              control_module: str = "NA",
                              field_device: str = "NA",
                              extra_column_headers: list = None,
                              cm_type: str = "NA",
                              cm_description: str = "NA",
                              fd_type: str = "NA",
                              fd_description: str = "NA",
                              category: str = "") -> int:
    """
    Save a new signal composition.
    """
    from datetime import datetime
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    extra_headers_json = json.dumps(extra_column_headers or [])
    
    with _db_connect() as con:
        # Insert composition
        cur = con.execute(
            "INSERT INTO signal_compositions "
            "(title, description, control_module, field_device, extra_column_headers, "
            "cm_type, cm_description, fd_type, fd_description, category, created, modified) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (title, description or "", control_module or "NA",
             field_device or "NA", extra_headers_json,
             cm_type or "NA", cm_description or "NA",
             fd_type or "NA", fd_description or "NA",
             category or "",
             now, now))
        composition_id = cur.lastrowid
        
        # Insert signals
        for order, sig in enumerate(signals):
            extra_values_json = json.dumps(sig.get("extra_column_values", []))
            con.execute(
                "INSERT INTO signal_composition_signals "
                "(composition_id, signal_name, signal_type, signal_description, "
                "prefix, suffix, extra_column_values, count, sort_order) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (composition_id, sig["signal_name"], sig["signal_type"],
                 sig.get("signal_description", ""),
                 sig.get("prefix") or "NA",
                 sig.get("suffix") or "NA",
                 extra_values_json,
                 int(sig.get("count", 1) or 1),
                 order))
        
        con.commit()
    
    return composition_id


def db_load_signal_composition(composition_id: int) -> dict:
    """
    Load a signal composition with all its signals.
    
    Returns:
        {
            "id": int,
            "title": str,
            "description": str,
            "control_module": str,   # cm_name
            "cm_type": str,
            "cm_description": str,
            "field_device": str,      # fd_name
            "fd_type": str,
            "fd_description": str,
            "extra_column_headers": [str, ...],
            "signals": [{"signal_name", "signal_type", "signal_description",
                         "count", "prefix", "suffix", "extra_column_values": [str, ...]}, ...]
        }
    """
    with _db_connect() as con:
        # Load composition
        comp = con.execute(
            "SELECT id, title, description, control_module, field_device, extra_column_headers, "
            "cm_type, cm_description, fd_type, fd_description, category "
            "FROM signal_compositions WHERE id = ?",
            (composition_id,)).fetchone()
        
        if not comp:
            return None
        
        # Load signals
        signals = con.execute(
            "SELECT signal_name, signal_type, signal_description, prefix, suffix, extra_column_values, count "
            "FROM signal_composition_signals WHERE composition_id = ? "
            "ORDER BY sort_order",
            (composition_id,)).fetchall()
    
    return {
        "id": comp[0],
        "title": comp[1],
        "description": comp[2] or "",
        "control_module": comp[3] or "NA",
        "field_device": comp[4] or "NA",
        "extra_column_headers": json.loads(comp[5] or "[]"),
        "cm_type": comp[6] or "NA",
        "cm_description": comp[7] or "NA",
        "fd_type": comp[8] or "NA",
        "fd_description": comp[9] or "NA",
        "category": comp[10] or "",
        "signals": [
            {
                "signal_name": s[0],
                "signal_type": s[1],
                "signal_description": s[2] or "",
                "prefix": s[3] or "NA",
                "suffix": s[4] or "NA",
                "extra_column_values": json.loads(s[5] or "[]"),
                "count": int(s[6] or 1),
            }
            for s in signals
        ]
    }


def db_load_compositions_by_owner(owner_id: int) -> list[dict]:
    """
    Load all signal compositions for a specific owner.

    Uses a single database connection and two bulk queries (compositions + signals)
    instead of opening a new connection per composition, which avoids the N+1
    performance problem when a project has many typicals.

    Args:
        owner_id: From composition_owners table

    Returns:
        List of composition dicts with signals
    """
    with _db_connect() as con:
        # Load all compositions for this owner in one query
        rows = con.execute(
            "SELECT sc.id, sc.title, sc.description, sc.control_module, sc.field_device, "
            "sc.extra_column_headers, sc.cm_type, sc.cm_description, sc.fd_type, "
            "sc.fd_description, sc.category "
            "FROM signal_compositions sc "
            "JOIN composition_ownership co ON co.composition_id = sc.id "
            "WHERE co.owner_id = ? "
            "ORDER BY co.sort_order, sc.id",
            (owner_id,)).fetchall()

        if not rows:
            return []

        # Load all signals for these compositions in one bulk query
        comp_ids = [r[0] for r in rows]
        placeholders = ",".join("?" for _ in comp_ids)
        sig_rows = con.execute(
            f"SELECT composition_id, signal_name, signal_type, signal_description, "
            f"prefix, suffix, extra_column_values, count "
            f"FROM signal_composition_signals "
            f"WHERE composition_id IN ({placeholders}) "
            f"ORDER BY composition_id, sort_order",
            comp_ids).fetchall()

    # Group signals by composition id
    signals_by_comp: dict = {}
    for s in sig_rows:
        signals_by_comp.setdefault(s[0], []).append({
            "signal_name": s[1],
            "signal_type": s[2],
            "signal_description": s[3] or "",
            "prefix": s[4] or "NA",
            "suffix": s[5] or "NA",
            "extra_column_values": json.loads(s[6] or "[]"),
            "count": int(s[7] or 1),
        })

    return [
        {
            "id": r[0],
            "title": r[1],
            "description": r[2] or "",
            "control_module": r[3] or "NA",
            "field_device": r[4] or "NA",
            "extra_column_headers": json.loads(r[5] or "[]"),
            "cm_type": r[6] or "NA",
            "cm_description": r[7] or "NA",
            "fd_type": r[8] or "NA",
            "fd_description": r[9] or "NA",
            "category": r[10] or "",
            "signals": signals_by_comp.get(r[0], []),
        }
        for r in rows
    ]


def db_load_all_compositions_for_project(project_id: int) -> dict:
    """
    Load all signal compositions visible to a project.
    Returns compositions grouped by owner.
    
    Returns:
        {
            "Default": [composition, composition, ...],
            "Project Name": [composition, ...],
            ...
        }
    """
    # Get Default owner compositions
    result = {}
    
    default_owner_id = db_ensure_default_owner()
    default_comps = db_load_compositions_by_owner(default_owner_id)
    if default_comps:
        result["Default"] = default_comps
    
    # Get project-specific compositions
    project_owner_id = db_get_or_create_project_owner(project_id)
    if project_owner_id:
        project_comps = db_load_compositions_by_owner(project_owner_id)
        if project_comps:
            # Get project name for grouping
            with _db_connect() as con:
                project = con.execute(
                    "SELECT name FROM projects WHERE id = ?",
                    (project_id,)).fetchone()
            
            if project:
                result[project[0]] = project_comps
    
    return result


def db_assign_composition_to_owner(composition_id: int, owner_id: int,
                                  sort_order: int = 999) -> None:
    """
    Assign a composition to an owner (create ownership link).
    
    Args:
        composition_id: From signal_compositions
        owner_id: From composition_owners
        sort_order: Display order (default 999 = append at end)
    """
    with _db_connect() as con:
        # Check if link already exists
        existing = con.execute(
            "SELECT id FROM composition_ownership "
            "WHERE composition_id = ? AND owner_id = ?",
            (composition_id, owner_id)).fetchone()
        
        if not existing:
            con.execute(
                "INSERT INTO composition_ownership "
                "(composition_id, owner_id, sort_order) "
                "VALUES (?, ?, ?)",
                (composition_id, owner_id, sort_order))
            con.commit()


def db_delete_signal_composition(composition_id: int) -> None:
    """
    Delete a signal composition (cascades to signals and ownership).
    """
    with _db_connect() as con:
        con.execute("DELETE FROM signal_compositions WHERE id = ?",
                   (composition_id,))
        con.commit()


def db_update_signal_composition(composition_id: int, title: str,
                                description: str, signals: list[dict],
                                control_module: str = "NA",
                                field_device: str = "NA",
                                extra_column_headers: list = None,
                                cm_type: str = "NA",
                                cm_description: str = "NA",
                                fd_type: str = "NA",
                                fd_description: str = "NA",
                                category: str = "") -> None:
    """
    Update an existing signal composition.
    """
    from datetime import datetime
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    extra_headers_json = json.dumps(extra_column_headers or [])
    
    with _db_connect() as con:
        # Update composition header
        con.execute(
            "UPDATE signal_compositions "
            "SET title = ?, description = ?, control_module = ?, field_device = ?, "
            "extra_column_headers = ?, cm_type = ?, cm_description = ?, "
            "fd_type = ?, fd_description = ?, category = ?, modified = ? "
            "WHERE id = ?",
            (title, description or "", control_module or "NA",
             field_device or "NA", extra_headers_json,
             cm_type or "NA", cm_description or "NA",
             fd_type or "NA", fd_description or "NA",
             category or "",
             now, composition_id))
        
        # Delete old signals
        con.execute(
            "DELETE FROM signal_composition_signals WHERE composition_id = ?",
            (composition_id,))
        
        # Insert new signals
        for order, sig in enumerate(signals):
            extra_values_json = json.dumps(sig.get("extra_column_values", []))
            con.execute(
                "INSERT INTO signal_composition_signals "
                "(composition_id, signal_name, signal_type, signal_description, "
                "prefix, suffix, extra_column_values, count, sort_order) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (composition_id, sig["signal_name"], sig["signal_type"],
                 sig.get("signal_description", ""),
                 sig.get("prefix") or "NA",
                 sig.get("suffix") or "NA",
                 extra_values_json,
                 int(sig.get("count", 1) or 1),
                 order))
        
        con.commit()
                
def db_update_project_drawing_meta(project_id: int,
                                   identifier: str,
                                   drw_name: str,
                                   area: str) -> None:
    """Save drawing metadata columns back to the projects table."""
    from datetime import datetime
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    with _db_connect() as con:
        con.execute(
            "UPDATE projects SET identifier=?, drw_name=?, area=?, modified=? "
            "WHERE id=?",
            (identifier, drw_name, area, now, project_id))
        con.commit()



# ---------------------------------------------------------------------------
# Signal-tree DB — single self-referential table, unlimited nesting
# ---------------------------------------------------------------------------
def _ensure_signal_tree_table(con: sqlite3.Connection) -> None:
    con.execute("""
        CREATE TABLE IF NOT EXISTS signal_tree (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            parent_id  INTEGER REFERENCES signal_tree(id) ON DELETE CASCADE,
            label      TEXT    NOT NULL,
            comment    TEXT    NOT NULL DEFAULT \'\',
            sort_order INTEGER NOT NULL DEFAULT 0
        )
    """)
    con.commit()

def _insert_tree_nodes(con, nodes: list[dict], parent_id) -> None:
    for order, node in enumerate(nodes):
        cur = con.execute(
            "INSERT INTO signal_tree (parent_id, label, comment, sort_order) "
            "VALUES (?, ?, ?, ?)",
            (parent_id, node["label"], node.get("comment", ""), order))
        if node.get("children"):
            _insert_tree_nodes(con, node["children"], cur.lastrowid)

def _read_tree_children(con, parent_id) -> list[dict]:
    rows = con.execute(
        "SELECT id, label, comment FROM signal_tree "
        "WHERE parent_id IS ? ORDER BY sort_order, id",
        (parent_id,)).fetchall()
    result = []
    for row_id, label, comment in rows:
        children = _read_tree_children(con, row_id)
        result.append({"label": label, "comment": comment, "children": children})
    return result

def _get_signal_composition(obj: dict) -> str:
    """
    Generate a compact representation of signals in a composition.
    
    Example: For signals [HDI, HDO, HDI, HDO], returns "2HDI 2HDO"
    
    Args:
        obj: Composition dict with "signals" list
        
    Returns:
        String like "2HDI 1HDO" or "1AI 1AO"
    """
    # Count signals by type, respecting each signal's count field
    signal_counts = {}
    for sig in obj.get("signals", []):
        # Signals are stored with "signal_type" field in the database
        sig_type = sig.get("signal_type", "")
        if sig_type:
            sig_count = int(sig.get("count", 1) or 1)
            signal_counts[sig_type] = signal_counts.get(sig_type, 0) + sig_count
    
    # Sort by type name for consistent display
    composition_parts = []
    for sig_type in sorted(signal_counts.keys()):
        count = signal_counts[sig_type]
        composition_parts.append(f"{count}{sig_type}")
    
    return " ".join(composition_parts)  # e.g., "2HDI 1HDO"

def db_load_signal_types() -> list[dict]:
    """Return recursive tree [{label, comment, children:[...]}]; seeds defaults on first run."""
    with _db_connect() as con:
        _ensure_signal_tree_table(con)
        count = con.execute("SELECT COUNT(*) FROM signal_tree").fetchone()[0]
        if count == 0:
            # Initialize with empty tree on first run
            # (users can configure signal types via the UI)
            con.commit()
            return []
        return _read_tree_children(con, None)
    

def db_save_signal_types(tree: list[dict]) -> None:
    with _db_connect() as con:
        _ensure_signal_tree_table(con)
        con.execute("DELETE FROM signal_tree")
        _insert_tree_nodes(con, tree, None)
        con.commit()


# ---------------------------------------------------------------------------
# Signal-tree  ➜  Excel export
# ---------------------------------------------------------------------------
def export_signal_types_to_xlsx(path: str, tree: list[dict]) -> None:
    """
    Write the signal-type tree to an Excel workbook.

    Columns
    -------
    Level               : nesting depth (1 = top-level group, 2 = child, …)
    Type                : "Group" if the node has children, "Signal" if leaf
    Label               : the node label shown in the drop-down menu
    Comment/Description : the descriptive text shown alongside the label

    The file can be edited by the user, then re-saved from Excel as
    "XML Spreadsheet 2003 (*.xml)" and imported back with
    import_signal_types_from_xml().
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill  = PatternFill("solid", start_color="1F4E79")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    grp_font  = Font(name="Arial", bold=True, size=10)
    grp_fill  = PatternFill("solid", start_color="D9E1F2")
    row_font  = Font(name="Arial", size=10)
    alt_fill  = PatternFill("solid", start_color="EEF2FA")
    center_al = Alignment(horizontal="center", vertical="center")
    left_al   = Alignment(horizontal="left",   vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "Signal Types"

    # ── header row ──────────────────────────────────────────────────────────
    headers = ["Level", "Type", "Label", "Comment/Description"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
    ws.row_dimensions[1].height = 22

    # ── flatten tree recursively ─────────────────────────────────────────────
    data_rows: list[tuple] = []   # (level, type_str, label, comment)

    def _flatten(nodes: list[dict], depth: int = 1) -> None:
        for node in nodes:
            children  = node.get("children") or []
            node_type = "Group" if children else "Signal"
            data_rows.append((depth, node_type,
                               node.get("label", ""),
                               node.get("comment", "")))
            if children:
                _flatten(children, depth + 1)

    _flatten(tree)

    # ── write data rows ──────────────────────────────────────────────────────
    for ri, (level, node_type, label, comment) in enumerate(data_rows, start=2):
        is_group = node_type == "Group"
        fill = grp_fill if is_group else (alt_fill if ri % 2 == 0 else PatternFill())
        font = grp_font if is_group else row_font

        ws.cell(row=ri, column=1, value=level).font      = font
        ws.cell(row=ri, column=1).fill                   = fill
        ws.cell(row=ri, column=1).alignment               = center_al

        ws.cell(row=ri, column=2, value=node_type).font  = font
        ws.cell(row=ri, column=2).fill                   = fill
        ws.cell(row=ri, column=2).alignment               = center_al

        indent = "    " * (level - 1)
        ws.cell(row=ri, column=3, value=indent + label).font = font
        ws.cell(row=ri, column=3).fill                        = fill
        ws.cell(row=ri, column=3).alignment                   = left_al

        ws.cell(row=ri, column=4, value=comment).font = row_font
        ws.cell(row=ri, column=4).fill                = fill
        ws.cell(row=ri, column=4).alignment           = left_al

    # ── column widths ────────────────────────────────────────────────────────
    for ci, w in enumerate([8, 10, 36, 50], start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A2"

    # ── instructions sheet ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ("HOW TO EDIT AND IMPORT THIS FILE", True),
        ("", False),
        ("1.  Edit the 'Signal Types' sheet:", True),
        ("    • Level  : depth in the tree (1 = top group, 2 = child, …)", False),
        ("    • Type   : 'Group' for nodes that contain children, 'Signal' for leaf items", False),
        ("    • Label  : the text shown in the drop-down menu", False),
        ("    • Comment/Description : optional description shown next to the label", False),
        ("", False),
        ("2.  Rules:", True),
        ("    • A 'Group' row must be followed by rows with Level = (Group Level + 1)", False),
        ("    • 'Signal' rows are leaf items — they must NOT be followed by deeper levels", False),
        ("    • Do NOT change or remove the header row", False),
        ("    • The Level and Type columns drive the tree structure — keep them consistent", False),
        ("", False),
        ("3.  Save the file in one of two ways to import:", True),
        ("    a) Keep as .xlsx and use 'Import Signal Types from Excel (.xlsx)…'", False),
        ("    b) In Excel: File → Save As → 'XML Spreadsheet 2003 (*.xml)'", False),
        ("       Then use 'Import Signal Types from XML…'", False),
        ("", False),
        ("4.  In the application: Edit menu → Import Signal Types from XML…", True),
        ("    (or from Excel…) and select your saved file.", False),
    ]
    for ri, (text, bold) in enumerate(instructions, start=1):
        cell = ws2.cell(row=ri, column=1, value=text)
        cell.font = Font(name="Arial", bold=bold, size=10)
    ws2.column_dimensions["A"].width = 80

    wb.save(path)
    wb.close()


# ---------------------------------------------------------------------------
# Signal-tree  ←  XML import  (Excel "XML Spreadsheet 2003" format)
# ---------------------------------------------------------------------------
def import_signal_types_from_xml(path: str) -> list[dict]:
    """
    Parse a file saved from Excel as "XML Spreadsheet 2003 (*.xml)".
    Expects the same four columns produced by export_signal_types_to_xlsx:
      Level | Type | Label | Comment/Description

    Returns a recursive tree  [{label, comment, children:[…]}]
    ready to be passed to db_save_signal_types().

    Raises ValueError with a descriptive message on structural problems.
    """
    import xml.etree.ElementTree as ET

    ns_ss  = "urn:schemas-microsoft-com:office:spreadsheet"
    ns_map = {
        "ss": ns_ss,
        "o":  "urn:schemas-microsoft-com:office:office",
        "x":  "urn:schemas-microsoft-com:office:excel",
    }

    tree_xml = ET.parse(path)
    root     = tree_xml.getroot()

    # ── locate the "Signal Types" worksheet ─────────────────────────────────
    worksheet = None
    for ws in root.iter(f"{{{ns_ss}}}Worksheet"):
        name = ws.get(f"{{{ns_ss}}}Name", "")
        if name.strip().lower() == "signal types":
            worksheet = ws
            break
    if worksheet is None:
        # Fall back to first worksheet
        worksheet = next(root.iter(f"{{{ns_ss}}}Worksheet"), None)
    if worksheet is None:
        raise ValueError("No worksheet found in the XML file.")

    table = worksheet.find(f"{{{ns_ss}}}Table")
    if table is None:
        raise ValueError("No Table element found in the worksheet.")

    # ── extract rows ─────────────────────────────────────────────────────────
    rows_data: list[list[str]] = []
    for row_el in table.findall(f"{{{ns_ss}}}Row"):
        cells = []
        for cell_el in row_el.findall(f"{{{ns_ss}}}Cell"):
            data_el = cell_el.find(f"{{{ns_ss}}}Data")
            cells.append(data_el.text.strip() if data_el is not None and data_el.text else "")
        rows_data.append(cells)

    if not rows_data:
        raise ValueError("The worksheet is empty.")

    # Skip header row (first row)
    data_rows = rows_data[1:]

    # ── parse rows into flat list ─────────────────────────────────────────────
    flat: list[dict] = []
    for i, row in enumerate(data_rows, start=2):   # i = Excel row number for errors
        if not any(row):
            continue   # skip blank rows
        try:
            level    = int(row[0]) if len(row) > 0 and row[0] else 1
            label    = row[2].strip() if len(row) > 2 else ""
            comment  = row[3].strip() if len(row) > 3 else ""
        except (ValueError, IndexError) as exc:
            raise ValueError(f"Row {i}: cannot parse Level value — {exc}") from exc

        # Strip leading indent spaces that were added on export
        label = label.lstrip()

        if not label:
            raise ValueError(f"Row {i}: Label is empty. Every row must have a label.")

        flat.append({"level": level, "label": label, "comment": comment})

    if not flat:
        raise ValueError("No data rows found after the header.")

    # ── rebuild recursive tree from flat list ─────────────────────────────────
    def _build_tree(rows: list[dict], start: int, parent_level: int) -> tuple[list[dict], int]:
        """
        Recursively consume rows from `start` while their level > parent_level.
        Returns (nodes, next_index).
        """
        nodes = []
        i = start
        while i < len(rows):
            row = rows[i]
            lvl = row["level"]
            if lvl <= parent_level:
                break
            if lvl > parent_level + 1:
                raise ValueError(
                    f"Row {i + 2}: Level jumps from {parent_level} to {lvl} "
                    f"(label: '{row['label']}'). Levels must increase by 1.")
            # Consume this node
            node = {"label": row["label"], "comment": row["comment"], "children": []}
            i += 1
            # Collect children (rows with level = lvl + 1)
            children, i = _build_tree(rows, i, lvl)
            node["children"] = children
            nodes.append(node)
        return nodes, i

    result, _ = _build_tree(flat, 0, 0)

    if not result:
        raise ValueError("Could not build a tree from the imported data.")

    return result


# ---------------------------------------------------------------------------
# Signal-tree  ←  xlsx import  (openpyxl — same sheet layout as export)
# ---------------------------------------------------------------------------
def import_signal_types_from_xlsx(path: str) -> list[dict]:
    """
    Import a signal-type tree from an .xlsx file that matches the layout
    produced by export_signal_types_to_xlsx().

    Columns expected (row 1 = header, skipped):
      A: Level  |  B: Type  |  C: Label  |  D: Comment/Description

    Returns a recursive tree [{label, comment, children:[…]}].
    Raises ValueError with a descriptive message on structural problems.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)

    # Prefer "Signal Types" sheet, fall back to first sheet
    ws = wb["Signal Types"] if "Signal Types" in wb.sheetnames else wb.active

    flat: list[dict] = []
    for ri, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip fully blank rows
        if not any(c for c in row if c is not None):
            continue

        raw_level   = row[0] if len(row) > 0 else None
        raw_label   = row[2] if len(row) > 2 else None
        raw_comment = row[3] if len(row) > 3 else None

        try:
            level = int(raw_level) if raw_level is not None else 1
        except (ValueError, TypeError) as exc:
            raise ValueError(f"Row {ri}: cannot parse Level value '{raw_level}' — {exc}") from exc

        label   = str(raw_label).strip().lstrip() if raw_label is not None else ""
        comment = str(raw_comment).strip()        if raw_comment is not None else ""

        if not label:
            raise ValueError(f"Row {ri}: Label is empty. Every row must have a label.")

        flat.append({"level": level, "label": label, "comment": comment})

    wb.close()

    if not flat:
        raise ValueError("No data rows found in the 'Signal Types' sheet.")

    # Rebuild recursive tree (same logic as xml importer)
    def _build_tree(rows: list[dict], start: int, parent_level: int) -> tuple[list[dict], int]:
        nodes = []
        i = start
        while i < len(rows):
            lvl = rows[i]["level"]
            if lvl <= parent_level:
                break
            if lvl > parent_level + 1:
                raise ValueError(
                    f"Row {i + 2}: Level jumps from {parent_level} to {lvl} "
                    f"(label: '{rows[i]['label']}'). Levels must increase by 1.")
            node = {"label": rows[i]["label"], "comment": rows[i]["comment"], "children": []}
            i += 1
            children, i = _build_tree(rows, i, lvl)
            node["children"] = children
            nodes.append(node)
        return nodes, i

    result, _ = _build_tree(flat, 0, 0)
    if not result:
        raise ValueError("Could not build a tree from the imported data.")
    return result


# ---------------------------------------------------------------------------
# Signal-tree configuration dialog — QTreeWidget with drag-drop
# ---------------------------------------------------------------------------
import copy as _copy_mod

class SignalTypeConfigDialog(QDialog):
    """
    Unlimited-depth tree editor for the signal type hierarchy.

    Left: QTreeWidget showing the full tree, drag-drop to reorder/reparent.
    Right: form to edit label & comment of the selected node.
    Toolbar: Add Child, Add Sibling, Delete, Expand All, Collapse All.

    Rules:
      - Any node can have children (becomes a submenu) or be a leaf (signal type).
      - Drag a leaf onto a branch to move it; drag a branch to reorganise groups.
      - On OK the current tree state is returned via signal_types().
    """
    _LABEL_COL   = 0
    _COMMENT_COL = 1

    def __init__(self, current_tree: list[dict], parent=None):
        super().__init__(parent)
        self.setWindowTitle("Configure Signal Types — Tree Editor")
        self.setMinimumWidth(720)
        self.setMinimumHeight(520)

        # ── tree widget ──────────────────────────────────────────────────
        self.tree = QTreeWidget()
        self.tree.setColumnCount(2)
        self.tree.setHeaderLabels(["Label", "Comment / Description"])
        self.tree.header().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.tree.header().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.tree.setDragDropMode(QAbstractItemView.DragDropMode.InternalMove)
        self.tree.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.tree.setAlternatingRowColors(True)
        self.tree.setEditTriggers(QAbstractItemView.EditTrigger.DoubleClicked |
                                   QAbstractItemView.EditTrigger.EditKeyPressed)
        self.tree.setItemsExpandable(True)
        self.tree.setIndentation(20)

        self._populate_tree(current_tree)
        self.tree.expandAll()

        # ── toolbar buttons ──────────────────────────────────────────────
        self.add_root_btn    = QPushButton("➕ Add Group")
        self.add_child_btn   = QPushButton("➕ Add Child")
        self.add_sibling_btn = QPushButton("➕ Add Sibling")
        self.delete_btn      = QPushButton("🗑 Delete")
        self.expand_btn      = QPushButton("⊞ Expand All")
        self.collapse_btn    = QPushButton("⊟ Collapse All")
        self.export_btn      = QPushButton("📤 Export to Excel…")
        self.import_btn      = QPushButton("📥 Import from XML/Excel…")
        self.delete_btn.setStyleSheet("color:#B71C1C;font-weight:bold;")
        self.add_root_btn.setStyleSheet("font-weight:bold;")
        self.export_btn.setStyleSheet(
            "QPushButton { background:#1F4E79; color:white; border-radius:4px;"
            "padding:3px 8px; font-weight:bold; }"
            "QPushButton:hover { background:#2E75B6; }")
        self.import_btn.setStyleSheet(
            "QPushButton { background:#1B5E20; color:white; border-radius:4px;"
            "padding:3px 8px; font-weight:bold; }"
            "QPushButton:hover { background:#2E7D32; }")

        toolbar = QHBoxLayout()
        for btn in (self.add_root_btn, self.add_child_btn, self.add_sibling_btn,
                    self.delete_btn, self.expand_btn, self.collapse_btn):
            toolbar.addWidget(btn)
        toolbar.addStretch()
        toolbar.addWidget(self.export_btn)
        toolbar.addWidget(self.import_btn)

        hint = QLabel(
            "💡 Tree structure:  "
            "<b>Group</b> (e.g. Digital)  →  "
            "<b>Signal Type</b> (e.g. DI)  →  "
            "<b>Sub-signal</b> (e.g. HDI, SDI)  →  "
            "<b>Comment</b> column for descriptions.  "
            "Nodes with children become sub-menus; leaf nodes appear as selectable signal types."
        )
        hint.setStyleSheet("color:#999999;font-size: 8pt;")
        hint.setWordWrap(True)

        ok_cancel = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok |
            QDialogButtonBox.StandardButton.Cancel)
        ok_cancel.accepted.connect(self._on_accept)
        ok_cancel.rejected.connect(self.reject)

        root = QVBoxLayout()
        root.addLayout(toolbar)
        root.addWidget(self.tree)
        root.addWidget(hint)
        root.addWidget(ok_cancel)
        self.setLayout(root)

        self.add_root_btn.clicked.connect(self._add_root)
        self.add_child_btn.clicked.connect(self._add_child)
        self.add_sibling_btn.clicked.connect(self._add_sibling)
        self.delete_btn.clicked.connect(self._delete_node)
        self.expand_btn.clicked.connect(self.tree.expandAll)
        self.collapse_btn.clicked.connect(self.tree.collapseAll)
        self.export_btn.clicked.connect(self._export_to_excel)
        self.import_btn.clicked.connect(self._import_from_file)

    # ── tree population ──────────────────────────────────────────────────
    def _populate_tree(self, nodes: list[dict],
                        parent: "QTreeWidgetItem | QTreeWidget | None" = None) -> None:
        parent = parent if parent is not None else self.tree
        for node in nodes:
            item = QTreeWidgetItem(parent)
            item.setText(self._LABEL_COL,   node.get("label", ""))
            item.setText(self._COMMENT_COL, node.get("comment", ""))
            item.setFlags(item.flags()
                          | Qt.ItemFlag.ItemIsEditable
                          | Qt.ItemFlag.ItemIsDragEnabled
                          | Qt.ItemFlag.ItemIsDropEnabled)
            if node.get("children"):
                self._populate_tree(node["children"], item)

    # ── add / delete ─────────────────────────────────────────────────────
    def _make_item(self, label="New Item", comment="") -> QTreeWidgetItem:
        item = QTreeWidgetItem()
        item.setText(self._LABEL_COL,   label)
        item.setText(self._COMMENT_COL, comment)
        item.setFlags(item.flags()
                      | Qt.ItemFlag.ItemIsEditable
                      | Qt.ItemFlag.ItemIsDragEnabled
                      | Qt.ItemFlag.ItemIsDropEnabled)
        return item

    def _add_root(self):
        """Add a new top-level group at the end of the tree."""
        text, ok = QInputDialog.getText(self, "Add Group", "Group name:")
        text = text.strip()
        if not ok or not text:
            return
        item = self._make_item(text)
        self.tree.addTopLevelItem(item)
        self.tree.setCurrentItem(item)
        self.tree.scrollToItem(item)

    def _add_child(self):
        sel = self.tree.currentItem()
        if sel is None:
            # Add at root level if nothing selected
            item = self._make_item("New Group")
            self.tree.addTopLevelItem(item)
        else:
            item = self._make_item()
            sel.addChild(item)
            sel.setExpanded(True)
        self.tree.setCurrentItem(item)
        self.tree.editItem(item, self._LABEL_COL)

    def _add_sibling(self):
        sel = self.tree.currentItem()
        if sel is None:
            self._add_child()
            return
        parent = sel.parent()
        item = self._make_item()
        if parent:
            idx = parent.indexOfChild(sel)
            parent.insertChild(idx + 1, item)
        else:
            idx = self.tree.indexOfTopLevelItem(sel)
            self.tree.insertTopLevelItem(idx + 1, item)
        self.tree.setCurrentItem(item)
        self.tree.editItem(item, self._LABEL_COL)

    def _delete_node(self):
        sel = self.tree.currentItem()
        if sel is None:
            return
        label = sel.text(self._LABEL_COL) or "this item"
        has_children = sel.childCount() > 0
        msg = f'Delete "{label}"'
        if has_children:
            msg += f" and its {sel.childCount()} child(ren)?"
        else:
            msg += "?"
        msg += "\n\nExisting markers are not affected."
        ans = QMessageBox.question(self, "Delete", msg,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if ans != QMessageBox.StandardButton.Yes:
            return
        parent = sel.parent()
        if parent:
            parent.removeChild(sel)
        else:
            self.tree.takeTopLevelItem(self.tree.indexOfTopLevelItem(sel))

    # ── serialise tree → dict list ────────────────────────────────────────
    def _item_to_dict(self, item: QTreeWidgetItem) -> dict:
        children = [self._item_to_dict(item.child(i))
                    for i in range(item.childCount())]
        return {
            "label":    item.text(self._LABEL_COL).strip(),
            "comment":  item.text(self._COMMENT_COL).strip(),
            "children": children,
        }

    def _collect_tree(self) -> list[dict]:
        return [self._item_to_dict(self.tree.topLevelItem(i))
                for i in range(self.tree.topLevelItemCount())]

    # ── accept ────────────────────────────────────────────────────────────
    def _on_accept(self):
        tree = self._collect_tree()
        if not tree:
            QMessageBox.warning(self, "Empty", "Add at least one node.")
            return
        # Check all labels are non-empty
        def check(nodes):
            for n in nodes:
                if not n["label"]:
                    return False
                if not check(n["children"]):
                    return False
            return True
        if not check(tree):
            QMessageBox.warning(self, "Empty label",
                                "All nodes must have a label.")
            return
        self.accept()

    def signal_types(self) -> list[dict]:
        return self._collect_tree()

    # ── Excel export ──────────────────────────────────────────────────────
    def _export_to_excel(self) -> None:
        """Export the current tree (as shown in the editor) to an .xlsx file."""
        current_tree = self._collect_tree()
        if not current_tree:
            QMessageBox.warning(self, "Nothing to export",
                                "The tree is empty — add some items first.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Signal Types to Excel",
            "signal_types.xlsx", "Excel Files (*.xlsx)")
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        try:
            export_signal_types_to_xlsx(path, current_tree)
            QMessageBox.information(
                self, "Export successful",
                f"Signal types exported to:\n{path}\n\n"
                f"Edit the file, then import it back via\n"
                f"'Import from XML/Excel…'")
        except Exception as exc:
            QMessageBox.critical(self, "Export failed", str(exc))

    # ── Import from XML or xlsx ───────────────────────────────────────────
    def _import_from_file(self) -> None:
        """
        Import a signal-type tree from an xlsx or XML Spreadsheet file.
        On success, replaces the current tree in the editor (undo is available
        by cancelling the dialog without pressing OK).
        """
        path, _ = QFileDialog.getOpenFileName(
            self, "Import Signal Types",
            "", "Excel / XML Files (*.xlsx *.xml);;All Files (*)")
        if not path:
            return

        try:
            ext = os.path.splitext(path)[1].lower()
            if ext == ".xml":
                new_tree = import_signal_types_from_xml(path)
            elif ext == ".xlsx":
                new_tree = import_signal_types_from_xlsx(path)
            else:
                QMessageBox.warning(self, "Unsupported format",
                                    "Please select an .xlsx or .xml file.")
                return
        except Exception as exc:
            QMessageBox.critical(self, "Import failed",
                                 f"Could not parse the file:\n\n{exc}")
            return

        if not new_tree:
            QMessageBox.warning(self, "Empty tree",
                                "The imported file contained no signal-type data.")
            return

        ans = QMessageBox.question(
            self, "Replace current tree?",
            f"This will replace the {self.tree.topLevelItemCount()} current "
            f"top-level group(s) with {len(new_tree)} imported group(s).\n\n"
            f"Continue? (You can still Cancel the dialog to discard the change.)",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if ans != QMessageBox.StandardButton.Yes:
            return

        # Replace tree content
        self.tree.clear()
        self._populate_tree(new_tree)
        self.tree.expandAll()
        QMessageBox.information(
            self, "Import successful",
            f"Imported {len(new_tree)} top-level group(s).\n\n"
            f"Review the tree, then press OK to save.")



# ---------------------------------------------------------------------------
# Excel export — DB helpers
# ---------------------------------------------------------------------------
# Available data sources a column can pull from each marker
EXPORT_SOURCES = [
    ("Signal Type",   "type"),
    ("Description",   "comment"),
    ("Count",         "count"),
    ("Notes",         "description"),
    ("Page Number",   "page"),
    ("Annotation Kind", "kind"),
    ("Fixed Value",   "__fixed__"),
]
EXPORT_SOURCE_LABELS = [s[0] for s in EXPORT_SOURCES]
EXPORT_SOURCE_KEYS   = {s[0]: s[1] for s in EXPORT_SOURCES}

def _ensure_export_tables(con: sqlite3.Connection) -> None:
    con.execute("""
        CREATE TABLE IF NOT EXISTS export_columns (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            col_name    TEXT NOT NULL,
            source      TEXT NOT NULL,
            fixed_value TEXT NOT NULL DEFAULT '',
            sort_order  INTEGER NOT NULL DEFAULT 0
        )
    """)
    con.execute("""
        CREATE TABLE IF NOT EXISTS export_metadata (
            pdf_path   TEXT PRIMARY KEY,
            identifier TEXT NOT NULL DEFAULT '',
            drw_name   TEXT NOT NULL DEFAULT '',
            area       TEXT NOT NULL DEFAULT ''
        )
    """)
    con.commit()

def db_load_export_columns() -> list[dict]:
    """Return [{col_name, source, fixed_value}]; seeds sensible defaults on first run."""
    with _db_connect() as con:
        _ensure_export_tables(con)
        rows = con.execute(
            "SELECT col_name, source, fixed_value "
            "FROM export_columns ORDER BY sort_order, id").fetchall()
        if not rows:
            defaults = [
                ("Signal Type", "type",     ""),
                ("Description", "comment",  ""),
                ("Page Number", "page",     ""),
            ]
            for i, (cn, src, fv) in enumerate(defaults):
                con.execute(
                    "INSERT INTO export_columns (col_name, source, fixed_value, sort_order) "
                    "VALUES (?,?,?,?)", (cn, src, fv, i))
            con.commit()
            return [{"col_name": cn, "source": src, "fixed_value": fv}
                    for cn, src, fv in defaults]
        return [{"col_name": r[0], "source": r[1], "fixed_value": r[2]} for r in rows]

def db_save_export_columns(cols: list[dict]) -> None:
    with _db_connect() as con:
        _ensure_export_tables(con)
        con.execute("DELETE FROM export_columns")
        for i, c in enumerate(cols):
            con.execute(
                "INSERT INTO export_columns (col_name, source, fixed_value, sort_order) "
                "VALUES (?,?,?,?)",
                (c["col_name"], c["source"], c.get("fixed_value", ""), i))
        con.commit()

def db_load_export_metadata(pdf_path: str) -> dict:
    with _db_connect() as con:
        _ensure_export_tables(con)
        row = con.execute(
            "SELECT identifier, drw_name, area FROM export_metadata WHERE pdf_path=?",
            (pdf_path,)).fetchone()
    if row:
        return {"identifier": row[0], "drw_name": row[1], "area": row[2]}
    return {"identifier": "", "drw_name": "", "area": ""}

def db_save_export_metadata(pdf_path: str, meta: dict) -> None:
    with _db_connect() as con:
        _ensure_export_tables(con)
        con.execute("""
            INSERT INTO export_metadata (pdf_path, identifier, drw_name, area)
            VALUES (?,?,?,?)
            ON CONFLICT(pdf_path) DO UPDATE SET
                identifier=excluded.identifier,
                drw_name=excluded.drw_name,
                area=excluded.area
        """, (pdf_path, meta["identifier"], meta["drw_name"], meta["area"]))
        con.commit()


# ---------------------------------------------------------------------------
# Excel export — Metadata dialog  (Step 1 of export)
# ---------------------------------------------------------------------------
class ExportMetadataDialog(QDialog):
    """Collects the three fixed header fields before export."""
    def __init__(self, meta: dict, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Export to Excel — Drawing Information")
        self.setFixedWidth(420)

        form = QFormLayout()
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        form.setVerticalSpacing(10)
        form.setContentsMargins(16, 16, 16, 8)

        self.id_edit   = QLineEdit(meta.get("identifier", ""))
        self.name_edit = QLineEdit(meta.get("drw_name",   ""))
        self.area_edit = QLineEdit(meta.get("area",       ""))

        self.id_edit.setPlaceholderText("e.g. P-001-A")
        self.name_edit.setPlaceholderText("e.g. Cooling Water System")
        self.area_edit.setPlaceholderText("e.g. Utility")

        form.addRow("<b>Technical Drawing Identifier:</b>", self.id_edit)
        form.addRow("<b>Technical Drawing Name:</b>",       self.name_edit)
        form.addRow("<b>Area / System:</b>",                self.area_edit)

        hint = QLabel("These values will appear in the first three columns of every row.")
        hint.setStyleSheet("color:#999999;font-size: 8pt;")
        hint.setWordWrap(True)

        bb = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok |
                              QDialogButtonBox.StandardButton.Cancel)
        bb.accepted.connect(self._on_accept)
        bb.rejected.connect(self.reject)

        lay = QVBoxLayout()
        lay.addLayout(form)
        lay.addWidget(hint)
        lay.addWidget(bb)
        self.setLayout(lay)

    def _on_accept(self):
        if not self.id_edit.text().strip():
            QMessageBox.warning(self, "Required",
                                "Technical Drawing Identifier cannot be empty.")
            return
        self.accept()

    @property
    def metadata(self) -> dict:
        return {
            "identifier": self.id_edit.text().strip(),
            "drw_name":   self.name_edit.text().strip(),
            "area":       self.area_edit.text().strip(),
        }


# ---------------------------------------------------------------------------
# Excel export — Column configurator dialog
# ---------------------------------------------------------------------------
class ExportColumnConfigDialog(QDialog):
    """
    Manage additional export columns (beyond the three fixed metadata fields).
    Each column has a name and a data source (or fixed value).
    """
    _SRC_COL  = 1
    _NAME_COL = 0
    _FV_COL   = 2

    def __init__(self, current_cols: list[dict], parent=None):
        super().__init__(parent)
        self.setWindowTitle("Configure Export Columns")
        self.setMinimumWidth(620)
        self.setMinimumHeight(400)

        # Table: Column Name | Data Source | Fixed Value
        self.table = QTableWidget(0, 3)
        self.table.setHorizontalHeaderLabels(
            ["Column Name", "Data Source", "Fixed Value (if applicable)"])
        self.table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.ResizeMode.Stretch)
        self.table.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(
            QAbstractItemView.SelectionMode.SingleSelection)
        self.table.verticalHeader().setVisible(False)

        for col in current_cols:
            self._append_row(col["col_name"], col["source"],
                             col.get("fixed_value", ""))

        # Side buttons
        self.add_btn  = QPushButton("➕  Add Column")
        self.del_btn  = QPushButton("🗑  Delete")
        self.up_btn   = QPushButton("▲  Up")
        self.down_btn = QPushButton("▼  Down")
        self.del_btn.setStyleSheet("color:#B71C1C;font-weight:bold;")

        side = QVBoxLayout()
        for b in (self.add_btn, self.del_btn, self.up_btn, self.down_btn):
            side.addWidget(b)
        side.addStretch()

        body = QHBoxLayout()
        body.addWidget(self.table, stretch=1)
        body.addLayout(side)

        note = QLabel(
            "Fixed metadata columns (Identifier, Name, Area/System) are always "
            "included as the first three columns and are not listed here.\n"
            "Set 'Data Source' to 'Fixed Value' and fill the last column to embed a "
            "constant (e.g. revision, approver).")
        note.setStyleSheet("color:#AAAAAA;font-size: 8pt;")
        note.setWordWrap(True)

        bb = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok |
                              QDialogButtonBox.StandardButton.Cancel)
        bb.accepted.connect(self._on_accept)
        bb.rejected.connect(self.reject)

        root = QVBoxLayout()
        root.addLayout(body)
        root.addWidget(note)
        root.addWidget(bb)
        self.setLayout(root)

        self.add_btn.clicked.connect(self._add)
        self.del_btn.clicked.connect(self._delete)
        self.up_btn.clicked.connect(self._move_up)
        self.down_btn.clicked.connect(self._move_down)
        self.table.cellChanged.connect(self._on_cell_changed)

    # ── helpers ──────────────────────────────────────────────────────────
    def _append_row(self, col_name: str, source: str, fixed_value: str) -> None:
        r = self.table.rowCount()
        self.table.insertRow(r)

        # Name cell — editable text
        self.table.setItem(r, self._NAME_COL, QTableWidgetItem(col_name))

        # Source cell — dropdown
        combo = QComboBox()
        combo.addItems(EXPORT_SOURCE_LABELS)
        # Find matching label for this source key
        label = next((lbl for lbl, key in EXPORT_SOURCE_KEYS.items()
                      if key == source), EXPORT_SOURCE_LABELS[0])
        combo.setCurrentText(label)
        combo.currentTextChanged.connect(
            lambda text, row=r: self._on_source_changed(row, text))
        self.table.setCellWidget(r, self._SRC_COL, combo)

        # Fixed value cell
        fv_item = QTableWidgetItem(fixed_value)
        self.table.setItem(r, self._FV_COL, fv_item)
        self._update_fv_cell(r, label)

    def _on_source_changed(self, row: int, label: str) -> None:
        self._update_fv_cell(row, label)

    def _update_fv_cell(self, row: int, source_label: str) -> None:
        """Grey out / enable the Fixed Value cell depending on source."""
        item = self.table.item(row, self._FV_COL)
        if item is None:
            item = QTableWidgetItem("")
            self.table.setItem(row, self._FV_COL, item)
        is_fixed = (source_label == "Fixed Value")
        flags = (Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsEditable |
                 Qt.ItemFlag.ItemIsSelectable) if is_fixed else Qt.ItemFlag.NoItemFlags
        item.setFlags(flags)
        item.setBackground(
            QColor("#FFFFFF") if is_fixed else QColor("#F0F0F0"))

    def _on_cell_changed(self, row, col):
        pass  # placeholder for future validation

    def _add(self):
        self.table.blockSignals(True)
        self._append_row("New Column", "type", "")
        self.table.blockSignals(False)
        new_row = self.table.rowCount() - 1
        self.table.setCurrentCell(new_row, self._NAME_COL)
        self.table.editItem(self.table.item(new_row, self._NAME_COL))

    def _delete(self):
        row = self.table.currentRow()
        if row < 0:
            return
        name = (self.table.item(row, self._NAME_COL) or
                QTableWidgetItem("this row")).text()
        ans = QMessageBox.question(
            self, "Delete Column", f'Delete column "{name}"?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if ans == QMessageBox.StandardButton.Yes:
            self.table.removeRow(row)

    def _swap(self, a: int, b: int) -> None:
        # Swap name and fixed-value items; re-create combos
        for col in (self._NAME_COL, self._FV_COL):
            ta = self.table.takeItem(a, col)
            tb = self.table.takeItem(b, col)
            self.table.setItem(a, col, tb)
            self.table.setItem(b, col, ta)
        # Swap combo selections
        ca = self.table.cellWidget(a, self._SRC_COL)
        cb = self.table.cellWidget(b, self._SRC_COL)
        va, vb = ca.currentText(), cb.currentText()
        ca.setCurrentText(vb)
        cb.setCurrentText(va)

    def _move_up(self):
        row = self.table.currentRow()
        if row > 0:
            self._swap(row, row - 1)
            self.table.setCurrentCell(row - 1, self._NAME_COL)

    def _move_down(self):
        row = self.table.currentRow()
        if 0 <= row < self.table.rowCount() - 1:
            self._swap(row, row + 1)
            self.table.setCurrentCell(row + 1, self._NAME_COL)

    def _on_accept(self):
        cols = self.export_columns()
        for c in cols:
            if not c["col_name"].strip():
                QMessageBox.warning(self, "Empty name",
                                    "All columns must have a name.")
                return
        self.accept()

    def export_columns(self) -> list[dict]:
        result = []
        for r in range(self.table.rowCount()):
            name  = (self.table.item(r, self._NAME_COL) or
                     QTableWidgetItem("")).text().strip()
            combo = self.table.cellWidget(r, self._SRC_COL)
            label = combo.currentText() if combo else EXPORT_SOURCE_LABELS[0]
            src   = EXPORT_SOURCE_KEYS.get(label, "type")
            fv    = (self.table.item(r, self._FV_COL) or
                     QTableWidgetItem("")).text().strip()
            result.append({"col_name": name, "source": src, "fixed_value": fv})
        return result

def _get_signal_composition(complex_obj: dict) -> str:
    """
    Generate a compact representation of signals in a complex object.
    
    Example: For signals [HDI, HDO, HDI, HDO], returns "2HDI 2HDO"
    
    Args:
        complex_obj: Complex object dict with "signals" list
        
    Returns:
        String like "2HDI 1HDO" or "1AI 1AO"
    """
    # Count signals by type, respecting each signal's count field
    signal_counts = {}
    for sig in complex_obj.get("signals", []):
        sig_type = sig.get("signal_type", "")  # e.g., "HDI", "HDO", "AI", "AO"
        if sig_type:
            sig_count = int(sig.get("count", 1) or 1)
            signal_counts[sig_type] = signal_counts.get(sig_type, 0) + sig_count
    
    # Sort by type name for consistent display
    composition_parts = []
    for sig_type in sorted(signal_counts.keys()):
        count = signal_counts[sig_type]
        composition_parts.append(f"{count}{sig_type}")
    
    return " ".join(composition_parts)  # e.g., "2HDI 1HDO"

# ─────────────────────────────────────────────────────────────────────
# Helper: read the PDF embedded title from file metadata
# ─────────────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────
# Helper: Expand markers to the new 6-column export format
# ─────────────────────────────────────────────────────────────────────
def _expand_markers_for_excel(markers: list[dict]) -> list[dict]:
    """
    Expand each marker into one or more row dicts with keys:
        name, type, desc1, desc2, tag_name, comments, page, file_name

    For composition markers the output is:
        1. One row for the Control Module (if present, i.e. not "NA")
        2. One row for the Field Device   (if present, i.e. not "NA")
        3. One row per signal × marker count
           (e.g. composition "2HDI 1HDO" placed with count=2 → 4 HDI + 2 HDO rows)

    Description 2 (desc2) is the composition title (first field in the Signal
    Configuration menu) and is repeated on every row of the same composition.
    Comments is the composition description and is also repeated.
    Tag Name for signals is  (prefix or "NA")-(signal_type)-(suffix or "NA").
    Tag Name for CM/FD rows is always "NA".

    For plain (non-composition) markers a single row is produced using the
    marker's own type / comment / description.
    """
    expanded = []

    for m in markers:
        if m.get("is_composition") and m.get("composition_id"):
            composition = db_load_signal_composition(m["composition_id"])
            if not composition:
                continue

            tag_parts = m.get("tag_parts") or {}
            prefix    = tag_parts.get("prefix", "") or ""
            suffix    = tag_parts.get("suffix", "") or ""
            # count stored in tag_parts["count"] by CompositionPlacementDialog
            count     = int(tag_parts.get("count", m.get("count", 1)) or 1)

            # Description 2 = composition title (first field in Signal Config menu)
            desc2    = composition.get("title", "")
            # Description/comments: prefer per-instance override from tag_parts
            comments = (tag_parts.get("description") or
                        composition.get("description", ""))
            file_nm  = m.get("file_name", "")
            drw_nm   = m.get("drw_name", "")
            drw_num  = m.get("drw_number", "")
            file_pt  = m.get("file_path", "")
            page     = m.get("page", 0)

            # Per-signal overrides saved by MarkerEditDialog / CompositionPlacementDialog
            signal_overrides = tag_parts.get("signal_overrides", [])

            # ── Control Module row (not multiplied by count) ─────────
            cm_name = composition.get("control_module", "NA") or "NA"
            if cm_name and cm_name.strip().upper() != "NA":
                expanded.append({
                    "name":       cm_name,
                    "type":       tag_parts.get("cm_type") or composition.get("cm_type", "") or "",
                    "desc1":      tag_parts.get("cm_description") or composition.get("cm_description", "") or "",
                    "desc2":      desc2,
                    "tag_name":   "NA",
                    "comments":   comments,
                    "page":       page,
                    "file_name":  file_nm,
                    "file_path":  file_pt,
                    "drw_name":   drw_nm,
                    "drw_number": drw_num,
                })

            # ── Field Device row (not multiplied by count) ────────────
            fd_name = composition.get("field_device", "NA") or "NA"
            if fd_name and fd_name.strip().upper() != "NA":
                expanded.append({
                    "name":       fd_name,
                    "type":       tag_parts.get("fd_type") or composition.get("fd_type", "") or "",
                    "desc1":      tag_parts.get("fd_description") or composition.get("fd_description", "") or "",
                    "desc2":      desc2,
                    "tag_name":   "NA",
                    "comments":   comments,
                    "page":       page,
                    "file_name":  file_nm,
                    "file_path":  file_pt,
                    "drw_name":   drw_nm,
                    "drw_number": drw_num,
                })

            # ── Signal rows × count ──────────────────────────────────
            extra_headers = composition.get("extra_column_headers", [])
            for sig_idx, signal in enumerate(composition.get("signals", [])):
                ov = signal_overrides[sig_idx] if sig_idx < len(signal_overrides) else {}
                sig_type = signal.get("signal_type", "")
                # Per-signal prefix/suffix override, then global, then "NA"
                sig_prefix = ov.get("prefix", prefix) or prefix
                sig_suffix = ov.get("suffix", suffix) or suffix
                tag_name   = f"{sig_prefix or 'NA'}-{sig_type}-{sig_suffix or 'NA'}"

                # Extra column values: prefer per-signal override
                ov_extras = ov.get("extra_column_values",
                                   signal.get("extra_column_values", []))
                extra_dict = {
                    f"extra_{h}": (ov_extras[i] if i < len(ov_extras) else "")
                    for i, h in enumerate(extra_headers)
                }

                row = {
                    "name":       signal.get("signal_name", ""),
                    "type":       sig_type,
                    "desc1":      ov.get("signal_description",
                                        signal.get("signal_description", "")),
                    "desc2":      desc2,
                    "tag_name":   tag_name,
                    "comments":   comments,
                    "page":       page,
                    "file_name":  file_nm,
                    "file_path":  file_pt,
                    "drw_name":   drw_nm,
                    "drw_number": drw_num,
                    **extra_dict,
                }
                for _ in range(count):
                    expanded.append(row.copy())
        else:
            # ── Plain (non-composition) marker ───────────────────────
            sig_type = (m.get("signal_type") or m.get("type") or "")
            expanded.append({
                "name":       m.get("type", ""),
                "type":       sig_type,
                "desc1":      m.get("description", "") or m.get("comment", ""),
                "desc2":      "",
                "tag_name":   m.get("complete_tag") or m.get("base_tag", ""),
                "comments":   m.get("comment", "") or m.get("signal_comment", ""),
                "page":       m.get("page", 0),
                "file_name":  m.get("file_name", ""),
                "file_path":  m.get("file_path", ""),
                "drw_name":   m.get("drw_name", ""),
                "drw_number": m.get("drw_number", ""),
            })

    return expanded
# ---------------------------------------------------------------------------
# Excel export — writer
# ---------------------------------------------------------------------------
def export_to_excel(path: str, pdf_path: str,
                    markers: list[dict], meta: dict,
                    col_config: list[dict]) -> None:
    """
    Build an Excel workbook from markers using the new 6-column format.

    Columns per row:
        Name | Type | Description 1 | Description 2 | Tag Name | Comments

    Composition markers are expanded:
        • One row for the Control Module (if present)
        • One row for the Field Device    (if present)
        • One row per signal in the composition
          (repeated signals, e.g. 2×HDI, produce two rows each)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    FIXED_HEADERS = ["Name", "Type", "Description 1",
                     "Description 2 (Signal Typical Details)",
                     "Tag Name", "Comments"]
    COL_WIDTHS    = [24, 14, 36, 36, 22, 36]

    # Expand markers into rows
    expanded_markers = _expand_markers_for_excel(markers)

    # Group by page
    pages: dict[int, list[dict]] = {}
    for m in expanded_markers:
        pages.setdefault(m["page"], []).append(m)

    # Styles
    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill  = PatternFill("solid", start_color="1F4E79")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    row_font  = Font(name="Arial", size=10)
    alt_fill  = PatternFill("solid", start_color="EEF2FA")
    center_al = Alignment(horizontal="center", vertical="center")
    left_al   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    wb = Workbook()
    wb.remove(wb.active)

    all_pages = sorted(pages.keys()) if pages else [0]

    for page_idx in all_pages:
        sheet_name = f"Page {page_idx + 1}"
        ws = wb.create_sheet(title=sheet_name)

        # Header row
        for ci, (hdr, w) in enumerate(zip(FIXED_HEADERS, COL_WIDTHS), start=1):
            cell = ws.cell(row=1, column=ci, value=hdr)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = hdr_align
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 28

        # Data rows
        page_rows = pages.get(page_idx, [])
        for ri, m in enumerate(page_rows, start=2):
            fill = alt_fill if ri % 2 == 0 else PatternFill()
            values = [
                m.get("name", ""),
                m.get("type", ""),
                m.get("desc1", ""),
                m.get("desc2", ""),
                m.get("tag_name", ""),
                m.get("comments", ""),
            ]
            for ci, val in enumerate(values, start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font      = row_font
                cell.fill      = fill
                cell.alignment = center_al if ci == 2 else left_al

        # Summary row
        if page_rows:
            summary_row = len(page_rows) + 2
            ws.cell(row=summary_row, column=1,
                    value=f"Total rows on this page: {len(page_rows)}"
                    ).font = Font(name="Arial", bold=True, italic=True,
                                  size=10, color="1F4E79")

        ws.freeze_panes = "A2"

    wb.save(path)
    wb.close()


# ---------------------------------------------------------------------------
# Project IO List — Excel export  (all IO markers across a whole project)
# ---------------------------------------------------------------------------
def export_project_io_list(path: str, project_id: int,
                           file_path: str | None = None) -> None:
    """
    Export IO signal markers to a single Excel sheet.

    Columns: Name | Type | Description 1 | Description 2 (Signal Typical Details)
             | Tag Name | Comments | Technical Drawing Number | Page Number
             | Project Name | Project Number | Project Description
             | [extra signal columns from Signal Typical config...]

    All markers from all drawings are written sequentially into one "IO List"
    sheet.  No per-drawing sub-sheets are created.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    markers = db_load_project_markers(project_id, file_path)
    meta    = db_get_project_drawing_meta(project_id)

    # Collect all extra column headers across all referenced compositions
    extra_headers_seen: list[str] = []
    comp_cache: dict[int, dict] = {}
    for m in markers:
        cid = m.get("composition_id")
        if cid and cid not in comp_cache:
            comp = db_load_signal_composition(cid)
            if comp:
                comp_cache[cid] = comp
                for h in comp.get("extra_column_headers", []):
                    if h not in extra_headers_seen:
                        extra_headers_seen.append(h)

    FIXED_COLS = [
        "Name", "Type", "Description 1",
        "Description 2 (Signal Typical Details)",
        "Tag Name", "Comments",
        "Technical Drawing Number", "Technical Drawing Description", "Page Number",
        "Project Name", "Project Number", "Project Description",
    ]
    FIXED_WIDTHS = [24, 14, 36, 36, 22, 36, 34, 28, 12, 28, 16, 36]
    EXTRA_WIDTH  = 22

    COLS       = FIXED_COLS + extra_headers_seen
    COL_WIDTHS = FIXED_WIDTHS + [EXTRA_WIDTH] * len(extra_headers_seen)

    # Styles
    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill  = PatternFill("solid", start_color="1F4E79")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    row_font  = Font(name="Arial", size=10)
    alt_fill  = PatternFill("solid", start_color="EEF2FA")
    center_al = Alignment(horizontal="center", vertical="center")
    left_al   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def _write_header(ws, start_row: int = 1):
        for ci, (hdr, w) in enumerate(zip(COLS, COL_WIDTHS), 1):
            cell = ws.cell(row=start_row, column=ci, value=hdr)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = hdr_align
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[start_row].height = 28

    def _write_row(ws, row_idx: int, r: dict, use_alt: bool):
        fill = alt_fill if use_alt else PatternFill()
        # If the user set file metadata (drw_name), use it; otherwise fall back
        # to the PDF file name stem.
        drawing_name   = r.get("drw_name", "") or os.path.splitext(r.get("file_name", ""))[0]
        drawing_number = r.get("drw_number", "")
        page_number    = r.get("page", 0) + 1
        fixed_values = [
            r.get("name", ""),
            r.get("type", ""),
            r.get("desc1", ""),
            r.get("desc2", ""),
            r.get("tag_name", ""),
            r.get("comments", ""),
            drawing_name,
            drawing_number,
            page_number,
            meta.get("name", ""),
            meta.get("number", ""),
            meta.get("description", ""),
        ]
        extra_values = [r.get(f"extra_{h}", "") for h in extra_headers_seen]
        values = fixed_values + extra_values
        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=ci, value=val)
            cell.font      = row_font
            cell.fill      = fill
            cell.alignment = center_al if ci in (2, 8) else left_al

    wb = Workbook()
    wb.remove(wb.active)

    # ── Single "IO List" sheet ─────────────────────────────────────────────
    ws_all = wb.create_sheet("IO List")

    # Row 1: project banner
    banner_val = (f"Project: {meta['name']}"
                  + (f"  ({meta['number']})" if meta["number"] else ""))
    cell = ws_all.cell(row=1, column=1, value=banner_val)
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell.fill      = PatternFill("solid", start_color="0D2B4E")
    cell.alignment = left_al
    ws_all.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=len(COLS))
    ws_all.row_dimensions[1].height = 20

    # Row 2: column headers
    _write_header(ws_all, start_row=2)
    ws_all.freeze_panes = "A3"

    # Data rows starting from row 3
    data_row   = 3
    alt_toggle = False
    total_rows = 0

    expanded_all = _expand_markers_for_excel(markers)
    for r in expanded_all:
        alt_toggle = not alt_toggle
        _write_row(ws_all, data_row, r, alt_toggle)
        data_row   += 1
        total_rows += 1

    if total_rows:
        ws_all.cell(row=data_row + 1, column=1,
                    value=f"Total rows: {total_rows}"
                    ).font = Font(name="Arial", bold=True,
                                  italic=True, size=10, color="1F4E79")

    # ── "Final Count" sheet — signal typical counts only ──────────────────
    ws_count = wb.create_sheet("Final Count")

    # Count each signal type from signal rows only (not CM/FD rows).
    # Signal rows are identified by their type being one of the 8 IO types.
    signal_type_counts: dict[str, int] = {t: 0 for t in _SIGNAL_IO_TYPES}
    for r in expanded_all:
        sig_t = r.get("type", "")
        if sig_t in signal_type_counts:
            signal_type_counts[sig_t] += 1

    # Header row
    fc_hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    fc_hdr_fill  = PatternFill("solid", start_color="1F4E79")
    fc_hdr_align = Alignment(horizontal="center", vertical="center")
    fc_row_font  = Font(name="Arial", size=11)
    fc_alt_fill  = PatternFill("solid", start_color="EEF2FA")
    fc_num_align = Alignment(horizontal="center", vertical="center")
    fc_lbl_align = Alignment(horizontal="left",   vertical="center")

    for ci, hdr in enumerate(("Signal Type", "Count"), start=1):
        c = ws_count.cell(row=1, column=ci, value=hdr)
        c.font      = fc_hdr_font
        c.fill      = fc_hdr_fill
        c.alignment = fc_hdr_align
    ws_count.column_dimensions["A"].width = 20
    ws_count.column_dimensions["B"].width = 14
    ws_count.row_dimensions[1].height = 24

    # Data rows
    grand_total = 0
    for ri, sig_type in enumerate(_SIGNAL_IO_TYPES, start=2):
        count_val = signal_type_counts[sig_type]
        grand_total += count_val
        use_alt = (ri % 2 == 0)
        row_fill = fc_alt_fill if use_alt else PatternFill()

        lbl_cell = ws_count.cell(row=ri, column=1, value=sig_type)
        lbl_cell.font      = fc_row_font
        lbl_cell.fill      = row_fill
        lbl_cell.alignment = fc_lbl_align

        cnt_cell = ws_count.cell(row=ri, column=2, value=count_val)
        cnt_cell.font      = fc_row_font
        cnt_cell.fill      = row_fill
        cnt_cell.alignment = fc_num_align

    # Grand total row
    total_row = len(_SIGNAL_IO_TYPES) + 2
    tot_font  = Font(name="Arial", bold=True, size=11, color="1F4E79")
    tot_fill  = PatternFill("solid", start_color="D9E1F2")

    lbl_tot = ws_count.cell(row=total_row, column=1, value="TOTAL")
    lbl_tot.font      = tot_font
    lbl_tot.fill      = tot_fill
    lbl_tot.alignment = fc_lbl_align

    cnt_tot = ws_count.cell(row=total_row, column=2, value=grand_total)
    cnt_tot.font      = tot_font
    cnt_tot.fill      = tot_fill
    cnt_tot.alignment = fc_num_align

    ws_count.freeze_panes = "A2"

    wb.save(path)
    wb.close()


# ---------------------------------------------------------------------------
# SessionMatchDialog — shown when a renamed/moved PDF is detected
# ---------------------------------------------------------------------------
class SessionMatchDialog(QDialog):
    """
    Presents one or more sessions that fingerprint-match the current PDF
    (but have a different path). The user picks which session to restore
    from, or starts fresh.
    """
    def __init__(self, current_path: str, matches: list[dict], parent=None):
        super().__init__(parent)
        self.setWindowTitle("Previous Session Detected")
        self.setMinimumWidth(580)
        self._chosen: dict | None = None

        cur_name = os.path.basename(current_path)
        cur_dir  = os.path.dirname(current_path)

        intro = QLabel(
            f"<b>The file you opened appears to match a previous session.</b><br>"
            f"Current file: <code>{cur_name}</code><br>"
            f"Current path: <code>{cur_dir}</code>"
        )
        intro.setWordWrap(True)
        intro.setStyleSheet("margin-bottom:8px;")

        self._table = QTableWidget(len(matches), 5)
        self._table.setHorizontalHeaderLabels(
            ["Previous Filename", "Previous Path", "Match Reason",
             "Last Opened", "Markers"])
        self._table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.ResizeToContents)
        self._table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch)
        self._table.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.ResizeMode.ResizeToContents)
        self._table.horizontalHeader().setSectionResizeMode(
            3, QHeaderView.ResizeMode.ResizeToContents)
        self._table.horizontalHeader().setSectionResizeMode(
            4, QHeaderView.ResizeMode.ResizeToContents)
        self._table.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setSelectionMode(
            QAbstractItemView.SelectionMode.SingleSelection)
        self._table.verticalHeader().setVisible(False)
        self._table.setEditTriggers(
            QAbstractItemView.EditTrigger.NoEditTriggers)

        self._matches = matches
        for r, m in enumerate(matches):
            old_dir = os.path.dirname(m["old_path"])
            reason_icon = "🔑" if "internal ID" in m["match_reason"] else "🔢"
            for c, val in enumerate([
                m["old_filename"],
                old_dir,
                f"{reason_icon}  {m['match_reason']}",
                m["last_opened"],
                str(m["marker_count"]),
            ]):
                item = QTableWidgetItem(val)
                item.setFlags(Qt.ItemFlag.ItemIsSelectable |
                              Qt.ItemFlag.ItemIsEnabled)
                self._table.setItem(r, c, item)

        if matches:
            self._table.selectRow(0)

        note = QLabel(
            "⚠️  If this is a revised version of the drawing, loading old markers "
            "may place them at incorrect positions. Review carefully after loading."
        )
        note.setWordWrap(True)
        note.setStyleSheet("color:#7B3F00; background:#FFF3E0;"
                           "padding:6px; border-radius:4px; font-size: 8pt;")

        bb = QDialogButtonBox()
        self._load_btn  = bb.addButton("✅  Load markers from selected session",
                                        QDialogButtonBox.ButtonRole.AcceptRole)
        self._fresh_btn = bb.addButton("🚫  Start fresh",
                                        QDialogButtonBox.ButtonRole.RejectRole)
        self._load_btn.clicked.connect(self._on_load)
        self._fresh_btn.clicked.connect(self.reject)

        lay = QVBoxLayout()
        lay.setContentsMargins(14, 14, 14, 10)
        lay.setSpacing(10)
        lay.addWidget(intro)
        lay.addWidget(self._table)
        lay.addWidget(note)
        lay.addWidget(bb)
        self.setLayout(lay)

    def _on_load(self):
        row = self._table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "No selection",
                                "Please select a session to load from.")
            return
        self._chosen = self._matches[row]
        self.accept()

    @property
    def chosen_match(self) -> dict | None:
        """The match dict the user selected, or None if they chose fresh."""
        return self._chosen


# ---------------------------------------------------------------------------
# ManualLinkDialog — File → "Link to Previous Session…"
# ---------------------------------------------------------------------------
class ManualLinkDialog(QDialog):
    """
    Lets the user manually pick any past session and link the current PDF to it.
    Used when auto-detection fails (e.g. the file was re-exported from CAD).
    """
    def __init__(self, current_path: str, sessions: list[dict], parent=None):
        super().__init__(parent)
        self.setWindowTitle("Link to Previous Session")
        self.setMinimumWidth(620)
        self.setMinimumHeight(360)
        self._chosen: dict | None = None
        self._sessions = sessions

        cur_name = os.path.basename(current_path)
        intro = QLabel(
            f"Manually link <b>{cur_name}</b> to a previous session.<br>"
            "All markers from the selected session will be loaded into the current file."
        )
        intro.setWordWrap(True)

        # Search bar
        self._search = QLineEdit()
        self._search.setPlaceholderText("🔍  Filter by filename…")
        self._search.textChanged.connect(self._filter)

        self._table = QTableWidget(0, 4)
        self._table.setHorizontalHeaderLabels(
            ["Filename", "Path", "Last Opened", "Markers"])
        self._table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.ResizeToContents)
        self._table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch)
        self._table.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.ResizeMode.ResizeToContents)
        self._table.horizontalHeader().setSectionResizeMode(
            3, QHeaderView.ResizeMode.ResizeToContents)
        self._table.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setSelectionMode(
            QAbstractItemView.SelectionMode.SingleSelection)
        self._table.verticalHeader().setVisible(False)
        self._table.setEditTriggers(
            QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.doubleClicked.connect(self._on_link)

        self._populate(sessions)

        note = QLabel(
            "⚠️  Existing markers on the current file (if any) will be replaced."
        )
        note.setStyleSheet("color:#FFD580; background:#3B2800;"
                           "padding:5px; border-radius:4px; font-size: 8pt;")
        note.setWordWrap(True)

        bb = QDialogButtonBox()
        link_btn   = bb.addButton("🔗  Link && Load",
                                   QDialogButtonBox.ButtonRole.AcceptRole)
        cancel_btn = bb.addButton(QDialogButtonBox.StandardButton.Cancel)
        link_btn.clicked.connect(self._on_link)
        cancel_btn.clicked.connect(self.reject)

        lay = QVBoxLayout()
        lay.setContentsMargins(14, 14, 14, 10)
        lay.setSpacing(8)
        lay.addWidget(intro)
        lay.addWidget(self._search)
        lay.addWidget(self._table)
        lay.addWidget(note)
        lay.addWidget(bb)
        self.setLayout(lay)

    def _populate(self, sessions: list[dict]) -> None:
        self._table.setRowCount(0)
        for s in sessions:
            r = self._table.rowCount()
            self._table.insertRow(r)
            for c, val in enumerate([
                s["pdf_filename"],
                os.path.dirname(s["pdf_path"]),
                s["last_opened"],
                str(s["marker_count"]),
            ]):
                item = QTableWidgetItem(val)
                item.setFlags(Qt.ItemFlag.ItemIsSelectable |
                              Qt.ItemFlag.ItemIsEnabled)
                item.setData(Qt.ItemDataRole.UserRole, s["session_id"])
                self._table.setItem(r, c, item)
        if self._table.rowCount():
            self._table.selectRow(0)

    def _filter(self, text: str) -> None:
        filtered = [s for s in self._sessions
                    if text.lower() in s["pdf_filename"].lower()]
        self._populate(filtered)

    def _on_link(self):
        row = self._table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "No selection",
                                "Please select a session to link to.")
            return
        sid = self._table.item(row, 0).data(Qt.ItemDataRole.UserRole)
        self._chosen = next(
            (s for s in self._sessions if s["session_id"] == sid), None)
        if self._chosen:
            self.accept()

    @property
    def chosen_session(self) -> dict | None:
        return self._chosen


# ---------------------------------------------------------------------------
# Project DB helpers
# ---------------------------------------------------------------------------
def _ensure_project_tables(con: sqlite3.Connection) -> None:
    con.execute("""
        CREATE TABLE IF NOT EXISTS projects (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            name        TEXT    NOT NULL,
            number      TEXT    NOT NULL DEFAULT '',
            description TEXT    NOT NULL DEFAULT '',
            created     TEXT    NOT NULL,
            modified    TEXT    NOT NULL
        )
    """)
    con.execute("""
        CREATE TABLE IF NOT EXISTS project_files (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id  INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
            file_path   TEXT    NOT NULL,
            file_name   TEXT    NOT NULL,
            sort_order  INTEGER NOT NULL DEFAULT 0
        )
    """)
    # Add per-file drawing metadata columns (migration-safe)
    for _col in ("drw_name", "drw_number"):
        try:
            con.execute(f"ALTER TABLE project_files ADD COLUMN {_col} TEXT NOT NULL DEFAULT ''")
        except Exception:
            pass
    con.commit()

def db_create_project(name: str, number: str, description: str) -> int:
    from datetime import datetime
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    with _db_connect() as con:
        _ensure_project_tables(con)
        cur = con.execute(
            "INSERT INTO projects (name, number, description, created, modified) "
            "VALUES (?,?,?,?,?)", (name, number, description, now, now))
        con.commit()
        return cur.lastrowid

def db_update_project(project_id: int, name: str, number: str, description: str) -> None:
    from datetime import datetime
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    with _db_connect() as con:
        _ensure_project_tables(con)
        con.execute(
            "UPDATE projects SET name=?, number=?, description=?, modified=? "
            "WHERE id=?", (name, number, description, now, project_id))
        con.commit()

def db_delete_project(project_id: int) -> None:
    with _db_connect() as con:
        _ensure_project_tables(con)
        con.execute("DELETE FROM projects WHERE id=?", (project_id,))
        con.commit()

def db_load_projects() -> list[dict]:
    with _db_connect() as con:
        _ensure_project_tables(con)
        rows = con.execute(
            "SELECT id, name, number, description, created, modified "
            "FROM projects ORDER BY name").fetchall()
    return [{"id": r[0], "name": r[1], "number": r[2],
             "description": r[3], "created": r[4], "modified": r[5]}
            for r in rows]

def db_add_project_file(project_id: int, file_path: str) -> int:
    with _db_connect() as con:
        _ensure_project_tables(con)
        # Avoid duplicates within same project
        exists = con.execute(
            "SELECT id FROM project_files WHERE project_id=? AND file_path=?",
            (project_id, file_path)).fetchone()
        if exists:
            return exists[0]
        order = (con.execute(
            "SELECT COALESCE(MAX(sort_order),0) FROM project_files "
            "WHERE project_id=?", (project_id,)).fetchone()[0]) + 1
        cur = con.execute(
            "INSERT INTO project_files (project_id, file_path, file_name, sort_order) "
            "VALUES (?,?,?,?)",
            (project_id, file_path, os.path.basename(file_path), order))
        con.commit()
        return cur.lastrowid

def db_remove_project_file(file_id: int) -> None:
    with _db_connect() as con:
        _ensure_project_tables(con)
        con.execute("DELETE FROM project_files WHERE id=?", (file_id,))
        con.commit()

def db_load_project_files(project_id: int) -> list[dict]:
    with _db_connect() as con:
        _ensure_project_tables(con)
        rows = con.execute(
            "SELECT id, file_path, file_name, "
            "COALESCE(drw_name,''), COALESCE(drw_number,'') "
            "FROM project_files "
            "WHERE project_id=? ORDER BY sort_order, id",
            (project_id,)).fetchall()
    return [{"id": r[0], "file_path": r[1], "file_name": r[2],
             "drw_name": r[3], "drw_number": r[4]} for r in rows]


def db_get_file_metadata(file_id: int) -> dict:
    """Return per-file drawing metadata (drw_name, drw_number)."""
    with _db_connect() as con:
        _ensure_project_tables(con)
        row = con.execute(
            "SELECT COALESCE(drw_name,''), COALESCE(drw_number,'') "
            "FROM project_files WHERE id=?", (file_id,)).fetchone()
    if row:
        return {"drw_name": row[0], "drw_number": row[1]}
    return {"drw_name": "", "drw_number": ""}


def db_save_file_metadata(file_id: int, drw_name: str, drw_number: str) -> None:
    """Persist per-file drawing metadata."""
    with _db_connect() as con:
        _ensure_project_tables(con)
        con.execute(
            "UPDATE project_files SET drw_name=?, drw_number=? WHERE id=?",
            (drw_name, drw_number, file_id))
        con.commit()

def db_find_pdf_in_projects(file_path: str) -> list[dict]:
    """
    Return a list of projects that contain file_path.
    Each entry: {project_id, project_name, project_number}
    Returns [] if the file is not in any project.
    """
    with _db_connect() as con:
        _ensure_project_tables(con)
        rows = con.execute(
            "SELECT p.id, p.name, p.number "
            "FROM projects p "
            "JOIN project_files pf ON pf.project_id = p.id "
            "WHERE pf.file_path = ?",
            (file_path,)).fetchall()
    return [{"project_id": r[0], "project_name": r[1],
             "project_number": r[2]} for r in rows]

def _get_projects_for_pdf(pdf_path: str) -> list[dict]:
    """
    Return all projects that contain this PDF.
    Each entry: {project_id, project_name, owner_id}
    """
    projects = db_find_pdf_in_projects(pdf_path)
    result = []
    for p in projects:
        owner_id = db_get_or_create_project_owner(p["project_id"])
        result.append({
            "project_id":   p["project_id"],
            "project_name": p["project_name"],
            "owner_id":     owner_id,
        })
    return result

# ─────────────────────────────────────────────────────────────────────
# CompositionPlacementDialog — dialog for placing a signal composition
# ─────────────────────────────────────────────────────────────────────
class CompositionPlacementDialog(QDialog):
    """
    Dialog shown when placing a Signal Typical on the PDF.
    Displays the full typical configuration.

    Read-only: title, Control Module Name, Field Device Name,
               Signal Name, Signal Type, Signal Description,
               Count, and Resulting Signal columns.
    Editable:  description, CM type/desc, FD type/desc,
               Prefix, Suffix, and any extra columns.
    """

    @staticmethod
    def _ro_colors() -> tuple:
        """Return (bg QColor, fg QColor) for read-only cells, respecting the current theme."""
        palette = QApplication.instance().palette()
        is_dark = palette.color(QPalette.ColorRole.Window).lightness() < 128
        if is_dark:
            return QColor("#2A2A2A"), QColor("#888888")
        return QColor("#E0E0E0"), QColor("#555555")

    # Fixed columns that mirror the Signal Typical configuration dialog.
    # Columns: Signal Name | Signal Type | Signal Description | Count |
    #          Prefix | Suffix | Resulting Signal | [extra columns…]
    _FIXED_COL_COUNT = 7

    def __init__(self, composition: dict, parent=None, tag_parts: dict = None):
        super().__init__(parent)
        self.setWindowTitle(f"Place {composition['title']}")
        self.setMinimumSize(750, 560)
        self._composition = composition
        tp = tag_parts or {}

        lay = QVBoxLayout()
        lay.setContentsMargins(16, 14, 16, 10)
        lay.setSpacing(10)

        # ── Title ─────────────────────────────────────────────────────────
        composition_text = _get_signal_composition(composition)
        title_lbl = QLabel(
            f"<b>{composition['title']}</b>"
            + (f"  <span style='color:#7EC8F0;'>{composition_text}</span>"
               if composition_text else ""))
        title_lbl.setStyleSheet("font-size: 11pt;")
        lay.addWidget(title_lbl)

        # ── Description ───────────────────────────────────────────────────
        desc_row = QHBoxLayout()
        desc_row.addWidget(QLabel("<b>Description:</b>"))
        self.desc_edit = QLineEdit(tp.get("description", composition.get("description", "")))
        desc_row.addWidget(self.desc_edit)
        lay.addLayout(desc_row)

        # ── Control Module + Field Device (side-by-side) ───────────────────
        cm_fd_lay = QHBoxLayout()

        cm_group = QGroupBox("Control Module")
        cm_group.setStyleSheet(
            "QGroupBox{border:1px solid #555;border-radius:4px;margin-top:6px;"
            "color:#F0F0F0;font-weight:bold;}"
            "QGroupBox::title{subcontrol-origin:margin;left:8px;padding:0 4px;}")
        cm_form = QFormLayout(cm_group)
        cm_form.setContentsMargins(8, 8, 8, 4)
        cm_form.setSpacing(4)
        cm_name_lbl = QLabel(composition.get("control_module", "NA") or "NA")
        cm_name_lbl.setStyleSheet("color:#AAAAAA;")
        cm_form.addRow("Name:", cm_name_lbl)
        self.cm_type_edit = QLineEdit(tp.get("cm_type", composition.get("cm_type", "NA") or "NA"))
        self.cm_desc_edit = QLineEdit(tp.get("cm_description", composition.get("cm_description", "NA") or "NA"))
        cm_form.addRow("Type:", self.cm_type_edit)
        cm_form.addRow("Description:", self.cm_desc_edit)
        cm_fd_lay.addWidget(cm_group)

        fd_group = QGroupBox("Field Device")
        fd_group.setStyleSheet(
            "QGroupBox{border:1px solid #555;border-radius:4px;margin-top:6px;"
            "color:#F0F0F0;font-weight:bold;}"
            "QGroupBox::title{subcontrol-origin:margin;left:8px;padding:0 4px;}")
        fd_form = QFormLayout(fd_group)
        fd_form.setContentsMargins(8, 8, 8, 4)
        fd_form.setSpacing(4)
        fd_name_lbl = QLabel(composition.get("field_device", "NA") or "NA")
        fd_name_lbl.setStyleSheet("color:#AAAAAA;")
        fd_form.addRow("Name:", fd_name_lbl)
        self.fd_type_edit = QLineEdit(tp.get("fd_type", composition.get("fd_type", "NA") or "NA"))
        self.fd_desc_edit = QLineEdit(tp.get("fd_description", composition.get("fd_description", "NA") or "NA"))
        fd_form.addRow("Type:", self.fd_type_edit)
        fd_form.addRow("Description:", self.fd_desc_edit)
        cm_fd_lay.addWidget(fd_group)
        lay.addLayout(cm_fd_lay)

        # ── Signals table (mirrors Signal Typical configuration columns) ──
        extra_headers = composition.get("extra_column_headers", [])
        fixed_labels = [
            "Signal Name", "Signal Type", "Signal Description",
            "Count", "Prefix", "Suffix", "Resulting Signal",
        ]
        all_headers = fixed_labels + extra_headers
        lay.addWidget(QLabel("<b>Signals:</b>"))
        self.signals_table = QTableWidget(0, len(all_headers))
        self.signals_table.setHorizontalHeaderLabels(all_headers)
        hdr = self.signals_table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        hdr.setSectionResizeMode(3, QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(4, QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(5, QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(6, QHeaderView.ResizeMode.Stretch)
        for ci in range(self._FIXED_COL_COUNT, len(all_headers)):
            hdr.setSectionResizeMode(ci, QHeaderView.ResizeMode.Interactive)
        self.signals_table.setColumnWidth(0, 110)
        self.signals_table.setColumnWidth(1, 90)
        self.signals_table.setColumnWidth(3, 60)
        self.signals_table.setColumnWidth(4, 65)
        self.signals_table.setColumnWidth(5, 65)
        self.signals_table.verticalHeader().setVisible(False)
        self.signals_table.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows)
        self.signals_table.setEditTriggers(
            QAbstractItemView.EditTrigger.DoubleClicked |
            QAbstractItemView.EditTrigger.SelectedClicked)

        overrides = tp.get("signal_overrides", [])
        signals = composition.get("signals", [])
        ro_bg, ro_fg = self._ro_colors()
        for i, sig in enumerate(signals):
            ov = overrides[i] if i < len(overrides) else {}
            r = self.signals_table.rowCount()
            self.signals_table.insertRow(r)

            sig_type = sig.get("signal_type", "")
            count    = int(sig.get("count", 1) or 1)

            # Read-only columns: Signal Name (0), Signal Type (1),
            #                    Signal Description (2), Count (3),
            #                    Resulting Signal (6)
            ro_values = [
                sig.get("signal_name", ""),        # 0
                sig_type,                           # 1
                sig.get("signal_description", ""),  # 2
                str(count),                         # 3
            ]
            for ci, val in enumerate(ro_values):
                item = QTableWidgetItem(val)
                item.setFlags(Qt.ItemFlag.ItemIsEnabled)
                item.setBackground(QBrush(ro_bg))
                item.setForeground(QBrush(ro_fg))
                self.signals_table.setItem(r, ci, item)

            # Editable: Prefix (4), Suffix (5)
            self.signals_table.setItem(r, 4, QTableWidgetItem(
                ov.get("prefix", sig.get("prefix", "NA"))))
            self.signals_table.setItem(r, 5, QTableWidgetItem(
                ov.get("suffix", sig.get("suffix", "NA"))))

            # Read-only: Resulting Signal (6) — computed from type × count
            resulting = sig_type if count <= 1 else f"{count}{sig_type}"
            res_item = QTableWidgetItem(resulting)
            res_item.setFlags(Qt.ItemFlag.ItemIsEnabled)
            res_item.setBackground(QBrush(ro_bg))
            res_item.setForeground(QBrush(ro_fg))
            self.signals_table.setItem(r, 6, res_item)

            # Editable extra columns (7+)
            ov_extras = ov.get("extra_column_values",
                               sig.get("extra_column_values", []))
            for ec in range(len(extra_headers)):
                val = ov_extras[ec] if ec < len(ov_extras) else ""
                self.signals_table.setItem(r, self._FIXED_COL_COUNT + ec, QTableWidgetItem(val))

        lay.addWidget(self.signals_table, stretch=1)

        # ── OK / Cancel ───────────────────────────────────────────────────
        bb = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok |
                              QDialogButtonBox.StandardButton.Cancel)
        bb.accepted.connect(self.accept)
        bb.rejected.connect(self.reject)
        lay.addWidget(bb)

        self.setLayout(lay)

    @property
    def tag_parts(self) -> dict:
        """Return all placement overrides."""
        extra_headers = self._composition.get("extra_column_headers", [])
        overrides = []
        for r in range(self.signals_table.rowCount()):
            ev = [
                (self.signals_table.item(r, self._FIXED_COL_COUNT + ec) or QTableWidgetItem()).text()
                for ec in range(len(extra_headers))
            ]
            overrides.append({
                "signal_description": (self.signals_table.item(r, 2) or QTableWidgetItem()).text(),
                "prefix":             (self.signals_table.item(r, 4) or QTableWidgetItem()).text(),
                "suffix":             (self.signals_table.item(r, 5) or QTableWidgetItem()).text(),
                "extra_column_values": ev,
            })
        return {
            "description":    self.desc_edit.text().strip(),
            "cm_type":        self.cm_type_edit.text().strip(),
            "cm_description": self.cm_desc_edit.text().strip(),
            "fd_type":        self.fd_type_edit.text().strip(),
            "fd_description": self.fd_desc_edit.text().strip(),
            "signal_overrides": overrides,
        }
            
# ---------------------------------------------------------------------------
# SignalCompositionTemplateDialog — manage templates
# ---------------------------------------------------------------------------
class SignalCompositionTemplateDialog(QDialog):
    """
    Manage signal composition templates that can be copied/pasted between projects.
    """
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Signal Typical Templates")
        self.setMinimumSize(1100, 720)
        self.resize(1200, 780)
        self._templates = {}
        self._current_template_id = None
        self._build_ui()
        self._load_default_templates()
    
    def _build_ui(self):
        lay = QVBoxLayout()
        lay.setContentsMargins(16, 14, 16, 10)
        lay.setSpacing(10)

        lay.addWidget(QLabel(
            "<b style='font-size:11pt;'>Signal Typical Templates</b><br>"
            "<span style='color:#AAAAAA;font-size:9pt;'>"
            "Create and manage reusable signal typical templates. "
            "Use Export / Import to share templates across machines.</span>"))

        # ── Left panel: template list ──────────────────────────────────────
        left_lay = QVBoxLayout()
        lbl_templates = QLabel("<b>Templates</b>")
        lbl_templates.setStyleSheet("font-size:10pt;")
        left_lay.addWidget(lbl_templates)

        self.template_list = QListWidget()
        self.template_list.setMinimumWidth(220)
        self.template_list.itemSelectionChanged.connect(self._on_template_selected)
        left_lay.addWidget(self.template_list, stretch=1)

        left_btn_lay = QHBoxLayout()
        new_btn = QPushButton("➕ New")
        del_btn = QPushButton("🗑 Delete")
        new_btn.setMinimumHeight(28)
        del_btn.setMinimumHeight(28)
        new_btn.clicked.connect(self._new_template)
        del_btn.clicked.connect(self._delete_template)
        left_btn_lay.addWidget(new_btn)
        left_btn_lay.addWidget(del_btn)
        left_lay.addLayout(left_btn_lay)

        left_widget = QWidget()
        left_widget.setLayout(left_lay)

        # ── Right panel: template editor ───────────────────────────────────
        right_lay = QVBoxLayout()
        right_lay.setSpacing(8)

        lbl_details = QLabel("<b>Template Details</b>")
        lbl_details.setStyleSheet("font-size:10pt;")
        right_lay.addWidget(lbl_details)

        # Form: name + description
        form_lay = QFormLayout()
        form_lay.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        form_lay.setHorizontalSpacing(10)
        form_lay.setVerticalSpacing(6)

        self.title_edit = QLineEdit()
        self.title_edit.setPlaceholderText("Enter template name…")
        self.title_edit.setMinimumHeight(28)
        form_lay.addRow("<b>Name:</b>", self.title_edit)

        self.desc_edit = QPlainTextEdit()
        self.desc_edit.setPlaceholderText("Enter template description…")
        self.desc_edit.setFixedHeight(72)
        form_lay.addRow("<b>Description:</b>", self.desc_edit)

        right_lay.addLayout(form_lay)

        lbl_signals = QLabel("<b>Signals</b>")
        lbl_signals.setStyleSheet("font-size:10pt;")
        right_lay.addWidget(lbl_signals)

        self.signals_table = QTableWidget(0, 3)
        self.signals_table.setHorizontalHeaderLabels(
            ["Signal Name", "Signal Type", "Description"])
        self.signals_table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.ResizeToContents)
        self.signals_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.ResizeToContents)
        self.signals_table.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.ResizeMode.Stretch)
        self.signals_table.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows)
        self.signals_table.setSelectionMode(
            QAbstractItemView.SelectionMode.SingleSelection)
        self.signals_table.verticalHeader().setVisible(False)
        self.signals_table.setMinimumHeight(200)
        right_lay.addWidget(self.signals_table, stretch=1)

        # Signal row buttons
        sig_btn_lay = QHBoxLayout()
        add_sig_btn = QPushButton("➕ Add Signal")
        rem_sig_btn = QPushButton("🗑 Remove Signal")
        add_sig_btn.setMinimumHeight(28)
        rem_sig_btn.setMinimumHeight(28)
        add_sig_btn.clicked.connect(self._add_signal_row)
        rem_sig_btn.clicked.connect(self._remove_signal_row)
        sig_btn_lay.addWidget(add_sig_btn)
        sig_btn_lay.addWidget(rem_sig_btn)
        sig_btn_lay.addStretch()
        right_lay.addLayout(sig_btn_lay)

        # Save / Clear buttons
        save_btn_lay = QHBoxLayout()
        save_btn = QPushButton("💾 Save Template")
        clear_btn = QPushButton("🔄 Clear")
        save_btn.setMinimumHeight(30)
        clear_btn.setMinimumHeight(30)
        save_btn.clicked.connect(self._save_template)
        clear_btn.clicked.connect(self._clear_form)
        save_btn_lay.addWidget(save_btn)
        save_btn_lay.addWidget(clear_btn)
        save_btn_lay.addStretch()
        right_lay.addLayout(save_btn_lay)

        right_widget = QWidget()
        right_widget.setLayout(right_lay)

        # ── Splitter (resizable left / right) ─────────────────────────────
        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.addWidget(left_widget)
        splitter.addWidget(right_widget)
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 3)
        splitter.setSizes([280, 820])
        lay.addWidget(splitter, stretch=1)

        # ── Export / Import row ────────────────────────────────────────────
        io_lay = QHBoxLayout()
        export_btn = QPushButton("📤  Export Templates to Excel…")
        import_btn = QPushButton("📥  Import Templates from Excel…")
        export_btn.setMinimumHeight(30)
        import_btn.setMinimumHeight(30)
        export_btn.setToolTip(
            "Export all templates to an Excel file.\n"
            "Each row represents one signal; template title and description\n"
            "are repeated across all rows belonging to the same template.")
        import_btn.setToolTip(
            "Import templates from a previously exported (or manually filled) Excel file.\n"
            "Columns: Template Title | Template Description | Signal Name | Signal Type | Signal Description")
        export_btn.clicked.connect(self._export_templates_xlsx)
        import_btn.clicked.connect(self._import_templates_xlsx)
        io_lay.addWidget(export_btn)
        io_lay.addWidget(import_btn)
        io_lay.addStretch()
        lay.addLayout(io_lay)

        # ── Dialog button box ──────────────────────────────────────────────
        bb = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
        bb.accepted.connect(self.accept)
        lay.addWidget(bb)

        self.setLayout(lay)
    
    def _load_default_templates(self):
        """Load templates from the database. Seed three built-in defaults if empty."""
        self._templates = db_load_all_templates()
        if not self._templates:
            defaults = [
                {
                    "title": "On-Off Valve",
                    "description": "2/2 Solenoid Valve",
                    "signals": [
                        {"signal_name": "XS",  "signal_type": "HDO",
                         "signal_description": "Valve position"},
                        {"signal_name": "ZSH", "signal_type": "HDI",
                         "signal_description": "Open limit"},
                        {"signal_name": "ZSL", "signal_type": "HDI",
                         "signal_description": "Closed limit"},
                    ],
                },
                {
                    "title": "Check Valve",
                    "description": "Inline Check Valve",
                    "signals": [
                        {"signal_name": "P_in",  "signal_type": "AI",
                         "signal_description": "Inlet Pressure"},
                        {"signal_name": "P_out", "signal_type": "AI",
                         "signal_description": "Outlet Pressure"},
                    ],
                },
                {
                    "title": "Flow Meter",
                    "description": "Flow Rate Meter",
                    "signals": [
                        {"signal_name": "FLOW",  "signal_type": "AI",
                         "signal_description": "Flow rate"},
                        {"signal_name": "PULSE", "signal_type": "HDI",
                         "signal_description": "Pulse output"},
                    ],
                },
            ]
            for d in defaults:
                tid = db_save_new_template(d["title"], d["description"], d["signals"])
                self._templates[tid] = {"id": tid, **d}
        self._refresh_template_list()
    
    def _refresh_template_list(self):
        """Refresh the template list widget."""
        self.template_list.clear()
        for template in self._templates.values():
            item = QListWidgetItem(template["title"])
            item.setData(Qt.ItemDataRole.UserRole, template["id"])
            self.template_list.addItem(item)
    
    def _on_template_selected(self):
        """Load selected template into editor."""
        items = self.template_list.selectedItems()
        if not items:
            self._clear_form()
            return
        
        template_id = items[0].data(Qt.ItemDataRole.UserRole)
        template = self._templates.get(template_id)
        
        if template:
            self._current_template_id = template_id
            self.title_edit.setText(template["title"])
            self.desc_edit.setPlainText(template["description"])
            self._populate_signals_table(template["signals"])
    
    def _populate_signals_table(self, signals: list[dict]):
        """Fill the signals table."""
        self.signals_table.setRowCount(0)
        for sig in signals:
            r = self.signals_table.rowCount()
            self.signals_table.insertRow(r)

            self.signals_table.setItem(r, 0, QTableWidgetItem(sig["signal_name"]))
            cb = QComboBox()
            cb.addItems(_SIGNAL_IO_TYPES)
            idx = cb.findText(sig.get("signal_type", "").strip())
            if idx >= 0:
                cb.setCurrentIndex(idx)
            self.signals_table.setCellWidget(r, 1, cb)
            self.signals_table.setItem(r, 2, QTableWidgetItem(sig.get("signal_description", "")))

    def _add_signal_row(self):
        """Add empty signal row."""
        r = self.signals_table.rowCount()
        self.signals_table.insertRow(r)
        self.signals_table.setItem(r, 0, QTableWidgetItem(""))
        cb = QComboBox()
        cb.addItems(_SIGNAL_IO_TYPES)
        self.signals_table.setCellWidget(r, 1, cb)
        self.signals_table.setItem(r, 2, QTableWidgetItem(""))

    def _remove_signal_row(self):
        """Remove selected signal row."""
        r = self.signals_table.currentRow()
        if r >= 0:
            self.signals_table.removeRow(r)
    
    def _clear_form(self):
        """Clear all form fields."""
        self._current_template_id = None
        self.title_edit.clear()
        self.desc_edit.clear()
        self.signals_table.setRowCount(0)
    
    def _new_template(self):
        """Create a new blank template."""
        self._clear_form()
        self.title_edit.setFocus()
    
    def _save_template(self):
        """Save current template to the database."""
        title = self.title_edit.text().strip()
        if not title:
            QMessageBox.warning(self, "Required", "Template name is required.")
            return

        # Collect signals
        signals = []
        for r in range(self.signals_table.rowCount()):
            sig_name = (self.signals_table.item(r, 0) or QTableWidgetItem()).text().strip()
            cb = self.signals_table.cellWidget(r, 1)
            sig_type = cb.currentText() if isinstance(cb, QComboBox) else (
                self.signals_table.item(r, 1) or QTableWidgetItem()).text().strip()
            sig_desc = (self.signals_table.item(r, 2) or QTableWidgetItem()).text().strip()

            if sig_name and sig_type:
                signals.append({
                    "signal_name": sig_name,
                    "signal_type": sig_type,
                    "signal_description": sig_desc,
                })

        if not signals:
            QMessageBox.warning(self, "Required", "At least one signal is required.")
            return

        description = self.desc_edit.toPlainText().strip()

        if self._current_template_id is None:
            # New template — persist to DB
            new_id = db_save_new_template(title, description, signals)
        else:
            # Update existing — persist to DB
            new_id = self._current_template_id
            db_update_template(new_id, title, description, signals)

        self._templates[new_id] = {
            "id":          new_id,
            "title":       title,
            "description": description,
            "signals":     signals,
        }
        self._current_template_id = new_id
        self._refresh_template_list()
        QMessageBox.information(self, "Saved", f"Template '{title}' saved successfully.")
    
    def _delete_template(self):
        """Delete selected template from the database."""
        items = self.template_list.selectedItems()
        if not items:
            QMessageBox.warning(self, "No Selection", "Please select a template to delete.")
            return

        template_id = items[0].data(Qt.ItemDataRole.UserRole)
        title = self._templates[template_id]["title"]

        ans = QMessageBox.question(
            self, "Delete Template",
            f"Delete template '{title}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)

        if ans == QMessageBox.StandardButton.Yes:
            db_delete_template(template_id)
            del self._templates[template_id]
            if self._current_template_id == template_id:
                self._clear_form()
            self._refresh_template_list()
    
    def _export_templates_xlsx(self):
        """Export all templates to an Excel file.

        Format: one row per signal.  Template Title and Template Description
        are repeated on every row that belongs to the same template so that
        the file can be round-tripped through the importer without any
        manual editing.

        Columns (A-E):
            A  Template Title
            B  Template Description
            C  Signal Name
            D  Signal Type
            E  Signal Description
        """
        if not self._templates:
            QMessageBox.information(self, "Nothing to Export",
                                    "No templates to export.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Signal Typical Templates",
            "signal_typical_templates.xlsx",
            "Excel Files (*.xlsx)")
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            wb = Workbook()
            ws = wb.active
            ws.title = "Templates"
            ws.freeze_panes = "A2"
            hdr_font  = Font(bold=True, color="FFFFFF")
            hdr_fill  = PatternFill("solid", start_color="1F4E79")
            hdr_align = Alignment(horizontal="center", vertical="center")
            COLS = [
                ("Template Title",       30),
                ("Template Description", 36),
                ("Signal Name",          20),
                ("Signal Type",          18),
                ("Signal Description",   36),
            ]
            for ci, (hdr, width) in enumerate(COLS, 1):
                cell = ws.cell(row=1, column=ci, value=hdr)
                cell.font  = hdr_font
                cell.fill  = hdr_fill
                cell.alignment = hdr_align
                ws.column_dimensions[get_column_letter(ci)].width = width
            row = 2
            for tmpl in self._templates.values():
                signals = tmpl.get("signals", [])
                if not signals:
                    # Export a placeholder row even when a template has no
                    # signals so the template name / description is preserved.
                    ws.cell(row=row, column=1, value=tmpl.get("title", ""))
                    ws.cell(row=row, column=2, value=tmpl.get("description", ""))
                    ws.cell(row=row, column=3, value="")
                    ws.cell(row=row, column=4, value="")
                    ws.cell(row=row, column=5, value="")
                    row += 1
                else:
                    for sig in signals:
                        ws.cell(row=row, column=1, value=tmpl.get("title", ""))
                        ws.cell(row=row, column=2, value=tmpl.get("description", ""))
                        ws.cell(row=row, column=3, value=sig.get("signal_name", ""))
                        ws.cell(row=row, column=4, value=sig.get("signal_type", ""))
                        ws.cell(row=row, column=5, value=sig.get("signal_description", ""))
                        row += 1
            wb.save(path)
            wb.close()
            QMessageBox.information(
                self, "Export Successful",
                f"Exported {len(self._templates)} template(s) to:\n{path}\n\n"
                "Note: each row represents one signal. Template Title and\n"
                "Description are repeated for every signal in that template.")
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", str(e))

    def _import_templates_xlsx(self):
        """Import templates from an Excel file with duplicate-title handling.

        Expected column order:
            A  Template Title       (required)
            B  Template Description (optional)
            C  Signal Name          (optional – template may have no signals)
            D  Signal Type          (optional – defaults to empty string)
            E  Signal Description   (optional)

        Row 1 is treated as a header row if cell A1 looks like a column heading
        (e.g. "Template Title", "Title", etc.).  If cell A1 contains real data
        (e.g. an actual template name) the row is treated as data so that files
        created manually without a header still import correctly.

        Multiple rows with the same Template Title are grouped into a single
        template; Title and Description are taken from the first row of each
        group (subsequent rows may repeat the same values or leave them blank).
        """
        path, _ = QFileDialog.getOpenFileName(
            self, "Import Signal Typical Templates",
            "", "Excel Files (*.xlsx)")
        if not path:
            return
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active

            # ── Read all rows starting at row 1 ───────────────────────────
            all_rows = list(ws.iter_rows(min_row=1, values_only=True))
            wb.close()
            if not all_rows:
                QMessageBox.warning(self, "Empty File",
                                    "No data rows found in the Excel file.")
                return

            # ── Detect whether row 1 is a header ──────────────────────────
            # If cell A1 contains a known header keyword we skip that row;
            # otherwise we treat row 1 as real data (manually created files).
            _HEADER_KEYWORDS = {
                "template title", "template name", "title", "name",
                "signal name", "signal type", "signal description",
                "description",
            }
            first_cell_val = all_rows[0][0] if all_rows[0] else None
            first_cell = str(first_cell_val).strip().lower() if first_cell_val is not None else ""
            is_header_row = first_cell in _HEADER_KEYWORDS or first_cell.startswith("template")
            rows = all_rows[1:] if is_header_row else all_rows

            if not rows:
                QMessageBox.warning(self, "Empty File",
                                    "No data rows found in the Excel file.")
                return

            # ── Parse rows → dict keyed by title ──────────────────────────
            imported: dict[str, dict] = {}
            for row in rows:
                if not row:
                    continue
                # Pad to 5 columns so index access is always safe
                padded = list(row) + [None] * 5
                title    = str(padded[0]).strip() if padded[0] is not None else ""
                desc     = str(padded[1]).strip() if padded[1] is not None else ""
                sig_name = str(padded[2]).strip() if padded[2] is not None else ""
                sig_type = str(padded[3]).strip() if padded[3] is not None else ""
                sig_desc = str(padded[4]).strip() if padded[4] is not None else ""

                if not title:
                    continue  # skip blank-title rows

                if title not in imported:
                    imported[title] = {
                        "title":       title,
                        "description": desc,
                        "signals":     [],
                    }
                # Add a signal row whenever at least the Signal Name is present.
                # Signal Type and Description default to empty string if omitted.
                if sig_name:
                    imported[title]["signals"].append({
                        "signal_name":        sig_name,
                        "signal_type":        sig_type,
                        "signal_description": sig_desc,
                    })

            if not imported:
                QMessageBox.warning(self, "No Templates",
                                    "Could not find any valid template data.\n\n"
                                    "Make sure the file uses the following column order "
                                    "(a header row is optional):\n"
                                    "A: Template Title  B: Template Description\n"
                                    "C: Signal Name     D: Signal Type  E: Signal Description")
                return

            # ── Detect conflicts ───────────────────────────────────────────
            existing_by_title: dict[str, dict] = {
                t["title"]: t for t in self._templates.values()
            }
            conflicts = [t for t in imported if t in existing_by_title]

            # ── Ask once for bulk strategy when there are conflicts ────────
            # Options:  "Overwrite All"  |  "Add All as New"  |  "Skip All"
            #           |  "Decide per template"
            bulk_action = None   # None means: decide per template
            if conflicts:
                mb = QMessageBox(self)
                mb.setWindowTitle("Duplicate Templates Detected")
                mb.setText(
                    f"<b>{len(conflicts)}</b> template(s) in the file already exist "
                    f"in the system:<br><br>"
                    + "<br>".join(f"• {c}" for c in conflicts[:10])
                    + ("<br>…" if len(conflicts) > 10 else "")
                    + "<br><br>How would you like to handle <b>all</b> conflicts?")
                ow_all_btn  = mb.addButton(
                    "Overwrite All",     QMessageBox.ButtonRole.AcceptRole)
                new_all_btn = mb.addButton(
                    "Add All as New",    QMessageBox.ButtonRole.YesRole)
                skip_all_btn = mb.addButton(
                    "Skip All",          QMessageBox.ButtonRole.NoRole)
                per_btn     = mb.addButton(
                    "Decide per Template", QMessageBox.ButtonRole.ResetRole)
                mb.setDefaultButton(per_btn)
                mb.exec()
                clicked = mb.clickedButton()
                if clicked == ow_all_btn:
                    bulk_action = "overwrite"
                elif clicked == new_all_btn:
                    bulk_action = "add_new"
                elif clicked == skip_all_btn:
                    bulk_action = "skip"
                else:
                    bulk_action = None  # decide per template

            # ── Process each imported template ─────────────────────────────
            # Remember the template that is currently open in the editor so
            # we can reload it after the list refresh (template_list.clear()
            # triggers _on_template_selected → _clear_form(), blanking the
            # editor even when the overwrite updated its signals).
            editing_id_before = self._current_template_id

            added = 0
            overwritten = 0
            skipped = 0

            for tmpl in imported.values():
                title       = tmpl["title"]
                signals     = tmpl["signals"]
                description = tmpl["description"]

                if title in existing_by_title:
                    action = bulk_action
                    if action is None:
                        # Per-template dialog
                        mb2 = QMessageBox(self)
                        mb2.setWindowTitle("Duplicate Template")
                        mb2.setText(
                            f"A template named <b>{title}</b> already exists.<br><br>"
                            "What would you like to do?")
                        ow_btn   = mb2.addButton(
                            "Overwrite existing",  QMessageBox.ButtonRole.AcceptRole)
                        new_btn2 = mb2.addButton(
                            "Add as new template", QMessageBox.ButtonRole.YesRole)
                        sk_btn   = mb2.addButton(
                            "Skip",                QMessageBox.ButtonRole.RejectRole)
                        mb2.setDefaultButton(sk_btn)
                        mb2.exec()
                        c2 = mb2.clickedButton()
                        if c2 == ow_btn:
                            action = "overwrite"
                        elif c2 == new_btn2:
                            action = "add_new"
                        else:
                            action = "skip"

                    if action == "overwrite":
                        existing = existing_by_title[title]
                        db_update_template(
                            existing["id"], title, description, signals)
                        existing["description"] = description
                        existing["signals"]     = signals
                        overwritten += 1
                    elif action == "add_new":
                        new_id = db_save_new_template(title, description, signals)
                        self._templates[new_id] = {
                            "id":          new_id,
                            "title":       title,
                            "description": description,
                            "signals":     signals,
                        }
                        added += 1
                    else:
                        skipped += 1
                else:
                    # No conflict — insert directly
                    new_id = db_save_new_template(title, description, signals)
                    self._templates[new_id] = {
                        "id":          new_id,
                        "title":       title,
                        "description": description,
                        "signals":     signals,
                    }
                    added += 1

            # ── Reload from DB + refresh list ──────────────────────────────
            # Reloading from the DB guarantees the in-memory state mirrors
            # exactly what was written, regardless of any edge case in the
            # in-memory patch above.
            self._templates = db_load_all_templates()
            self._refresh_template_list()

            # If the template that was open in the editor was affected by the
            # import, reload its form so the user immediately sees the updated
            # signals (without having to re-click the item in the list).
            if editing_id_before is not None and editing_id_before in self._templates:
                reloaded = self._templates[editing_id_before]
                self._current_template_id = editing_id_before
                self.title_edit.setText(reloaded["title"])
                self.desc_edit.setPlainText(reloaded["description"])
                self._populate_signals_table(reloaded["signals"])
                # Re-select the item in the list so it stays highlighted.
                # Block signals to avoid triggering _on_template_selected
                # again (which would overwrite the text we just set).
                self.template_list.blockSignals(True)
                for i in range(self.template_list.count()):
                    item = self.template_list.item(i)
                    if item.data(Qt.ItemDataRole.UserRole) == editing_id_before:
                        self.template_list.setCurrentItem(item)
                        break
                self.template_list.blockSignals(False)

            parts = []
            if added:
                parts.append(f"{added} added")
            if overwritten:
                parts.append(f"{overwritten} overwritten")
            if skipped:
                parts.append(f"{skipped} skipped")
            QMessageBox.information(
                self, "Import Complete",
                f"Import finished: {', '.join(parts) or 'nothing changed'}.\n"
                f"Source: {path}")
        except Exception as e:
            QMessageBox.critical(self, "Import Failed", str(e))

    def get_templates(self) -> dict:
        """Return all templates."""
        return self._templates

# ---------------------------------------------------------------------------
# NoProjectDialog — shown when trying to configure without a project
# ---------------------------------------------------------------------------
class NoProjectDialog(QDialog):
    """
    Shown when user tries to configure signal compositions without opening a project.
    Provides guidance on how to set up a project.
    """
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Signal Typicals Configuration")
        self.setMinimumWidth(500)
        self.setMinimumHeight(380)
        lay = QVBoxLayout()
        lay.setContentsMargins(24, 24, 24, 24)
        lay.setSpacing(16)
        
        # Icon/emoji
        title = QLabel("⚙️")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet("font-size: 48pt;")
        lay.addWidget(title)
        
        # Main message
        msg = QLabel("<b>Signal Typicals Configuration</b>")
        msg.setAlignment(Qt.AlignmentFlag.AlignCenter)
        msg.setStyleSheet("font-size: 14pt; color: #7EC8F0;")
        lay.addWidget(msg)
        
        # Instructions
        instructions = QLabel(
            "Signal typicals are configured on a <b>per-project basis</b>.<br><br>"
            "To configure signal typicals:<br><br>"
            "<span style='color: #90CAF9;'><b>1. Open or Create a Project</b></span><br>"
            "   • Click the <b>Project Panel</b> (📁) in the toolbar<br>"
            "   • Create a new project or open an existing one<br><br>"
            "<span style='color: #90CAF9;'><b>2. Add PDF to Project</b></span><br>"
            "   • Drag and drop PDF files onto the project<br>"
            "   • Or right-click project → Add PDF file(s)<br><br>"
            "<span style='color: #90CAF9;'><b>3. Open PDF and Configure</b></span><br>"
            "   • Open the PDF file from the project<br>"
            "   • Go to Edit → Configure Signal Typicals<br><br>"
            "Signal typicals will then be available when placing markers<br>"
            "on that project's PDFs."
        )
        instructions.setWordWrap(True)
        instructions.setStyleSheet("line-height: 1.6;")
        lay.addWidget(instructions)
    
        lay.addStretch()
        
        # Close button
        close_btn = QPushButton("Close")
        close_btn.setMaximumWidth(100)
        close_btn.clicked.connect(self.accept)
        btn_lay = QHBoxLayout()
        btn_lay.addStretch()
        btn_lay.addWidget(close_btn)
        lay.addLayout(btn_lay)
        
        self.setLayout(lay)
        
# ---------------------------------------------------------------------------
# Allowed IO signal types (used in both config and template dialogs)
# ---------------------------------------------------------------------------
_SIGNAL_IO_TYPES = ("HDI", "HDO", "HAI", "HAO", "SDI", "SDO", "SAI", "SAO")

# ---------------------------------------------------------------------------
# SignalCompositionConfigDialog — manage signal compositions (project-specific)
# ---------------------------------------------------------------------------
class SignalCompositionConfigDialog(QDialog):
    """
    Manage signal compositions for a specific project.
    No longer uses Default owner - all compositions are project-specific.
    """
    
    _FIXED_COL_COUNT = 7  # Signal Name, Type, Desc, Count, Prefix, Suffix, Resulting Signal
    
    def __init__(self, owner_id: int, owner_name: str, parent=None):
        """
        Args:
            owner_id: ID from composition_owners table (project owner)
            owner_name: Display name (project name)
        """
        super().__init__(parent)
        self.setWindowTitle(f"Signal Typicals - {owner_name}")
        self.setMinimumSize(950, 650)
        self._owner_id = owner_id
        self._owner_name = owner_name
        self._current_comp_id = None
        self._compositions = db_load_compositions_by_owner(owner_id)
        self._build_ui()
    
    def _build_ui(self):
        lay = QVBoxLayout()
        lay.setContentsMargins(16, 14, 16, 10)
        lay.setSpacing(10)
        
        lay.addWidget(QLabel(
            "<span style='color:#AAAAAA;font-size:8pt;'>"
            "Define signal typicals with detailed signal information.</span>"))
        
        # ── Left Panel: Typical Tree ──────────────────────────────────────
        left_lay = QVBoxLayout()
        left_lay.addWidget(QLabel("<b>Typicals:</b>"))

        self.comp_tree = QTreeWidget()
        self.comp_tree.setHeaderHidden(True)
        self.comp_tree.setRootIsDecorated(True)
        self.comp_tree.setSortingEnabled(False)
        self.comp_tree.itemSelectionChanged.connect(self._on_comp_selected)
        # Enable drag-and-drop between categories
        self.comp_tree.setDragEnabled(True)
        self.comp_tree.setAcceptDrops(True)
        self.comp_tree.setDragDropMode(QAbstractItemView.DragDropMode.InternalMove)
        self.comp_tree.setDefaultDropAction(Qt.DropAction.MoveAction)
        self.comp_tree.model().rowsInserted.connect(self._schedule_tree_sync)
        left_lay.addWidget(self.comp_tree, stretch=1)

        # Typical buttons (new / delete)
        left_btn_lay = QHBoxLayout()
        new_comp_btn = QPushButton("➕ New")
        del_comp_btn = QPushButton("🗑 Delete")
        new_comp_btn.clicked.connect(self._new_composition)
        del_comp_btn.clicked.connect(self._delete_composition)
        left_btn_lay.addWidget(new_comp_btn)
        left_btn_lay.addWidget(del_comp_btn)
        left_lay.addLayout(left_btn_lay)

        # Category management buttons
        left_lay.addWidget(QLabel("<b>Categories:</b>"))
        cat_btn_lay = QHBoxLayout()
        add_cat_btn    = QPushButton("➕ Add")
        rename_cat_btn = QPushButton("✏️ Rename")
        del_cat_btn    = QPushButton("🗑 Delete")
        add_cat_btn.setToolTip("Add a new category to the tree")
        rename_cat_btn.setToolTip("Rename the selected category")
        del_cat_btn.setToolTip("Delete empty category (or reassign its typicals to General)")
        add_cat_btn.clicked.connect(self._add_category)
        rename_cat_btn.clicked.connect(self._rename_category)
        del_cat_btn.clicked.connect(self._delete_category)
        cat_btn_lay.addWidget(add_cat_btn)
        cat_btn_lay.addWidget(rename_cat_btn)
        cat_btn_lay.addWidget(del_cat_btn)
        left_lay.addLayout(cat_btn_lay)
        
        left_widget = QWidget()
        left_widget.setLayout(left_lay)
        left_widget.setMaximumWidth(280)
        
        # ── Right Panel: Composition Editor ───────────────────────────────
        right_lay = QVBoxLayout()

        # Title and description
        right_lay.addWidget(QLabel("<b>Typical Details:</b>"))

        # Category field
        cat_row = QHBoxLayout()
        cat_row.addWidget(QLabel("Category:"))
        self.cat_edit = QLineEdit()
        self.cat_edit.setPlaceholderText("e.g., HVAC, Valves (leave blank for General)")
        self.cat_edit.setToolTip("Group this typical under a category in the tree")
        cat_row.addWidget(self.cat_edit)
        right_lay.addLayout(cat_row)

        self.title_edit = QLineEdit()
        self.title_edit.setPlaceholderText("e.g., On-Off Valve")
        right_lay.addWidget(self.title_edit)

        # Description fields — first field always present; more via button
        right_lay.addWidget(QLabel("Description:"))
        self._desc_fields: list = []
        self._desc_buttons: list = []
        self._desc_container = QWidget()
        self._desc_layout = QVBoxLayout(self._desc_container)
        self._desc_layout.setContentsMargins(0, 0, 0, 0)
        self._desc_layout.setSpacing(4)
        self._add_desc_field()  # always show at least one field
        right_lay.addWidget(self._desc_container)

        self._add_desc_btn = QPushButton("➕ Add Description")
        self._add_desc_btn.clicked.connect(lambda: self._add_desc_field())
        right_lay.addWidget(self._add_desc_btn)

        # Control Module and Field Device as horizontal single-row tables
        cm_fd_outer = QHBoxLayout()

        # Control Module table
        cm_widget = QWidget()
        cm_vlay   = QVBoxLayout(cm_widget)
        cm_vlay.setContentsMargins(0, 0, 0, 0)
        cm_vlay.setSpacing(2)
        cm_vlay.addWidget(QLabel("<b>Control Module:</b>"))
        self.cm_table = QTableWidget(1, 3)
        self.cm_table.setHorizontalHeaderLabels(["Name", "Type", "Description"])
        self.cm_table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.Interactive)
        self.cm_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Interactive)
        self.cm_table.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.ResizeMode.Stretch)
        self.cm_table.verticalHeader().setVisible(False)
        self.cm_table.setFixedHeight(56)
        self.cm_table.setColumnWidth(0, 100)
        self.cm_table.setColumnWidth(1, 90)
        # Name cell — editable (plain text; read-only lock comes after load)
        self.cm_name_edit = QTableWidgetItem("NA")
        self.cm_table.setItem(0, 0, self.cm_name_edit)
        self.cm_type_item = QTableWidgetItem("CM")
        self.cm_type_item.setFlags(Qt.ItemFlag.ItemIsEnabled)  # read-only: always "CM"
        self.cm_table.setItem(0, 1, self.cm_type_item)
        self.cm_desc_item = QTableWidgetItem("NA")
        self.cm_table.setItem(0, 2, self.cm_desc_item)
        cm_vlay.addWidget(self.cm_table)
        cm_fd_outer.addWidget(cm_widget)

        # Field Device table
        fd_widget = QWidget()
        fd_vlay   = QVBoxLayout(fd_widget)
        fd_vlay.setContentsMargins(0, 0, 0, 0)
        fd_vlay.setSpacing(2)
        fd_vlay.addWidget(QLabel("<b>Field Device:</b>"))
        self.fd_table = QTableWidget(1, 3)
        self.fd_table.setHorizontalHeaderLabels(["Name", "Type", "Description"])
        self.fd_table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.Interactive)
        self.fd_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Interactive)
        self.fd_table.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.ResizeMode.Stretch)
        self.fd_table.verticalHeader().setVisible(False)
        self.fd_table.setFixedHeight(56)
        self.fd_table.setColumnWidth(0, 100)
        self.fd_table.setColumnWidth(1, 90)
        self.fd_name_item = QTableWidgetItem("NA")
        self.fd_table.setItem(0, 0, self.fd_name_item)
        self.fd_type_item = QTableWidgetItem("NA")
        self.fd_table.setItem(0, 1, self.fd_type_item)
        self.fd_desc_item = QTableWidgetItem("NA")
        self.fd_table.setItem(0, 2, self.fd_desc_item)
        fd_vlay.addWidget(self.fd_table)
        cm_fd_outer.addWidget(fd_widget)

        right_lay.addLayout(cm_fd_outer)

        # Back-compat aliases so the rest of the code that reads .text() works
        # We override the read helpers below instead.

        # Composition display
        right_lay.addWidget(QLabel("<b>Typical:</b>"))
        self.composition_display = QLabel("")
        self.composition_display.setStyleSheet("color: #7EC8F0; font-weight: bold; font-size: 11pt;")
        right_lay.addWidget(self.composition_display)
        
        # Signals table with 7 columns
        right_lay.addWidget(QLabel("<b>Signals Configuration:</b>"))
        
        self.signals_table = QTableWidget(0, 7)
        self.signals_table.setHorizontalHeaderLabels(
            ["Signal Name*", "Signal Type*", "Signal Description*",
             "Count*", "Prefix*", "Suffix*", "Resulting Signal on PDF"])
        hdr = self.signals_table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        hdr.setSectionResizeMode(3, QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(4, QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(5, QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(6, QHeaderView.ResizeMode.Stretch)
        # Set initial widths large enough to show the full header text
        self.signals_table.setColumnWidth(0, 110)  # "Signal Name*"
        self.signals_table.setColumnWidth(1, 105)  # "Signal Type*"
        self.signals_table.setColumnWidth(3, 65)   # "Count*"
        self.signals_table.setColumnWidth(4, 65)   # "Prefix*"
        self.signals_table.setColumnWidth(5, 65)   # "Suffix*"
        self.signals_table.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows)
        self.signals_table.setSelectionMode(
            QAbstractItemView.SelectionMode.SingleSelection)
        self.signals_table.verticalHeader().setVisible(False)
        self.signals_table.itemChanged.connect(self._update_resulting_signal)
        right_lay.addWidget(self.signals_table, stretch=1)
        
        # Signal row and column buttons
        sig_btn_lay = QHBoxLayout()
        add_sig_btn = QPushButton("➕ Add Signal")
        rem_sig_btn = QPushButton("🗑 Remove Signal")
        add_col_btn = QPushButton("➕ Add Column")
        rem_col_btn = QPushButton("🗑 Remove Column")
        add_sig_btn.clicked.connect(self._add_signal_row)
        rem_sig_btn.clicked.connect(self._remove_signal_row)
        add_col_btn.clicked.connect(self._add_extra_column)
        rem_col_btn.clicked.connect(self._remove_extra_column)
        add_col_btn.setToolTip("Add a custom description/comment column to the right of the table")
        rem_col_btn.setToolTip("Remove a custom description/comment column")
        sig_btn_lay.addWidget(add_sig_btn)
        sig_btn_lay.addWidget(rem_sig_btn)
        sig_btn_lay.addSpacing(16)
        sig_btn_lay.addWidget(add_col_btn)
        sig_btn_lay.addWidget(rem_col_btn)
        sig_btn_lay.addStretch()
        right_lay.addLayout(sig_btn_lay)
        
        right_widget = QWidget()
        right_widget.setLayout(right_lay)
        
        # ── Main split layout ─────────────────────────────────────────────
        main_lay = QHBoxLayout()
        main_lay.addWidget(left_widget, stretch=0)
        main_lay.addWidget(right_widget, stretch=1)
        lay.addLayout(main_lay, stretch=1)
        
        # Template buttons
        template_lay = QHBoxLayout()
        template_btn = QPushButton("📋 Use Template")
        template_btn.clicked.connect(self._use_template)
        template_lay.addWidget(template_btn)
        template_lay.addStretch()
        lay.addLayout(template_lay)
        
        # Dialog buttons – OK, Apply (save without closing), Cancel
        bb = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok |
                              QDialogButtonBox.StandardButton.Apply |
                              QDialogButtonBox.StandardButton.Cancel)
        bb.accepted.connect(self._on_accept)
        bb.button(QDialogButtonBox.StandardButton.Apply).clicked.connect(self._on_apply)
        bb.rejected.connect(self.reject)
        lay.addWidget(bb)
        
        self.setLayout(lay)
        self._populate_tree()

    def _populate_tree(self):
        """Populate the tree of typicals grouped by category."""
        # Disconnect rowsInserted for the duration of this programmatic rebuild.
        # rowsInserted fires for every item inserted, which would otherwise schedule
        # a _sync_categories_from_tree timer after each item.  Because the timer fires
        # *after* this method returns (next event-loop iteration), the re-entrancy flag
        # alone cannot break the resulting timer chain.  Disconnecting the signal means
        # only genuine user drag-drop operations trigger the sync, not our own rebuilds.
        try:
            self.comp_tree.model().rowsInserted.disconnect(self._schedule_tree_sync)
        except RuntimeError:
            pass  # already disconnected; safe to ignore
        try:
            self.comp_tree.setUpdatesEnabled(False)
            self.comp_tree.clear()
            # Group by category
            groups: dict[str, list] = {}
            for comp in self._compositions:
                cat = comp.get("category", "").strip() or "No group assigned"
                groups.setdefault(cat, []).append(comp)

            # Also include pending (empty) categories created with "Add Category"
            for pending in getattr(self, "_pending_new_categories", set()):
                if pending not in groups:
                    groups[pending] = []

            # Sort: "No group assigned" first, then alphabetical
            sorted_cats = sorted(groups.keys(),
                                 key=lambda c: ("" if c == "No group assigned" else c.lower()))
            for cat in sorted_cats:
                cat_item = QTreeWidgetItem(self.comp_tree, [cat])
                cat_item.setData(0, Qt.ItemDataRole.UserRole, None)  # not a composition
                font = cat_item.font(0)
                font.setBold(True)
                cat_item.setFont(0, font)
                # Make category items selectable (for rename/delete) and droppable
                # so that typicals can be dragged onto them.
                cat_item.setFlags(
                    Qt.ItemFlag.ItemIsEnabled
                    | Qt.ItemFlag.ItemIsSelectable
                    | Qt.ItemFlag.ItemIsDropEnabled)
                for comp in groups[cat]:
                    child = QTreeWidgetItem(cat_item, [comp["title"]])
                    child.setData(0, Qt.ItemDataRole.UserRole, comp["id"])
                    # Leaf items must be draggable so they can be moved between categories.
                    child.setFlags(
                        Qt.ItemFlag.ItemIsEnabled
                        | Qt.ItemFlag.ItemIsSelectable
                        | Qt.ItemFlag.ItemIsDragEnabled)
                cat_item.setExpanded(True)
        finally:
            self.comp_tree.setUpdatesEnabled(True)
            self.comp_tree.model().rowsInserted.connect(self._schedule_tree_sync)
    
    def _on_comp_selected(self):
        """When user selects a typical from the tree."""
        items = self.comp_tree.selectedItems()
        if not items:
            self._clear_form()
            return

        comp_id = items[0].data(0, Qt.ItemDataRole.UserRole)
        if comp_id is None:
            # Category header selected — ignore
            return

        comp = next((c for c in self._compositions if c["id"] == comp_id), None)

        if comp:
            self._current_comp_id = comp_id
            self.cat_edit.setText(comp.get("category", ""))
            self.title_edit.setText(comp["title"])
            # Populate description fields
            desc_lines = [dl for dl in comp["description"].split("\n") if dl] if comp["description"] else []
            self._reset_desc_fields()
            if desc_lines:
                self._desc_fields[0].setText(desc_lines[0])
                for dl in desc_lines[1:]:
                    self._add_desc_field(dl)
            self._add_desc_btn.setEnabled(len(self._desc_fields) < 5)
            self.cm_table.item(0, 0).setText(comp.get("control_module", "NA") or "NA")
            self.cm_table.item(0, 1).setText("CM")  # always fixed
            self.cm_table.item(0, 2).setText(comp.get("cm_description", "NA") or "NA")
            self.fd_table.item(0, 0).setText(comp.get("field_device", "NA") or "NA")
            self.fd_table.item(0, 1).setText(comp.get("fd_type", "NA") or "NA")
            self.fd_table.item(0, 2).setText(comp.get("fd_description", "NA") or "NA")
            # Reset to fixed columns then restore any extra columns saved with this composition
            self.signals_table.blockSignals(True)
            self.signals_table.setColumnCount(self._FIXED_COL_COUNT)
            self.signals_table.setRowCount(0)
            self.signals_table.blockSignals(False)
            for header in comp.get("extra_column_headers", []):
                self._add_extra_column(header)
            self._populate_signals_table(comp["signals"])
            self._update_composition_display()
        else:
            self._clear_form()
    
    def _clear_form(self):
        """Clear all form fields and reset the signals table to fixed columns only."""
        self._current_comp_id = None
        self.cat_edit.clear()
        self.title_edit.clear()
        self._reset_desc_fields()
        self.cm_table.item(0, 0).setText("NA")
        self.cm_table.item(0, 1).setText("CM")  # always fixed
        self.cm_table.item(0, 2).setText("NA")
        self.fd_table.item(0, 0).setText("NA")
        self.fd_table.item(0, 1).setText("NA")
        self.fd_table.item(0, 2).setText("NA")
        self.signals_table.blockSignals(True)
        self.signals_table.setColumnCount(self._FIXED_COL_COUNT)
        self.signals_table.setRowCount(0)
        self.signals_table.blockSignals(False)
        self.composition_display.setText("")
    
    def _populate_signals_table(self, signals: list[dict]):
        """Fill the signals table (fixed columns + any extra columns already set up)."""
        self.signals_table.blockSignals(True)
        try:
            self.signals_table.setRowCount(0)
            extra_count = self.signals_table.columnCount() - self._FIXED_COL_COUNT
            for sig in signals:
                r = self.signals_table.rowCount()
                self.signals_table.insertRow(r)

                self.signals_table.setItem(r, 0, QTableWidgetItem(sig["signal_name"]))
                cb = self._make_signal_type_combo(sig["signal_type"])
                cb.currentTextChanged.connect(
                    lambda _txt, c=cb: self._update_resulting_signal_for_combo(c))
                self.signals_table.setCellWidget(r, 1, cb)
                self.signals_table.setItem(r, 2, QTableWidgetItem(sig.get("signal_description", "")))
                count_item = QTableWidgetItem(str(sig.get("count", 1)))
                self.signals_table.setItem(r, 3, count_item)
                self.signals_table.setItem(r, 4, QTableWidgetItem(sig.get("prefix") or "NA"))
                self.signals_table.setItem(r, 5, QTableWidgetItem(sig.get("suffix") or "NA"))
                resulting = self._calculate_resulting_signal(sig["signal_type"], sig.get("count", 1))
                result_item = QTableWidgetItem(resulting)
                result_item.setFlags(result_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.signals_table.setItem(r, 6, result_item)
                # Extra columns
                extra_values = sig.get("extra_column_values", [])
                for ec in range(extra_count):
                    val = extra_values[ec] if ec < len(extra_values) else ""
                    self.signals_table.setItem(
                        r, self._FIXED_COL_COUNT + ec, QTableWidgetItem(val))
        finally:
            self.signals_table.blockSignals(False)
        self._update_composition_display()

    def _add_signal_row(self):
        """Add an empty signal row (fills all fixed and any extra columns)."""
        r = self.signals_table.rowCount()
        self.signals_table.insertRow(r)
        self.signals_table.blockSignals(True)
        try:
            self.signals_table.setItem(r, 0, QTableWidgetItem(""))
            cb = self._make_signal_type_combo()
            cb.currentTextChanged.connect(
                lambda _txt, c=cb: self._update_resulting_signal_for_combo(c))
            self.signals_table.setCellWidget(r, 1, cb)
            self.signals_table.setItem(r, 2, QTableWidgetItem(""))
            self.signals_table.setItem(r, 3, QTableWidgetItem("1"))
            self.signals_table.setItem(r, 4, QTableWidgetItem("NA"))
            self.signals_table.setItem(r, 5, QTableWidgetItem("NA"))
            result_item = QTableWidgetItem(cb.currentText())
            result_item.setFlags(result_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.signals_table.setItem(r, 6, result_item)
            # Fill any extra columns with empty text
            for ec in range(self.signals_table.columnCount() - self._FIXED_COL_COUNT):
                self.signals_table.setItem(r, self._FIXED_COL_COUNT + ec, QTableWidgetItem(""))
        finally:
            self.signals_table.blockSignals(False)

    def _remove_signal_row(self):
        """Remove the selected signal row."""
        r = self.signals_table.currentRow()
        if r >= 0:
            self.signals_table.removeRow(r)
        self._update_composition_display()
    
    def _calculate_resulting_signal(self, signal_type: str, count: int) -> str:
        """Calculate the resulting signal (e.g., 2HDI)."""
        if count <= 1:
            return signal_type
        return f"{count}{signal_type}"

    def _make_signal_type_combo(self, current: str = "") -> "QComboBox":
        """Return a QComboBox pre-loaded with the 8 allowed IO signal types."""
        cb = QComboBox()
        cb.addItems(_SIGNAL_IO_TYPES)
        idx = cb.findText(current.strip())
        if idx >= 0:
            cb.setCurrentIndex(idx)
        return cb

    def _get_row_signal_type(self, row: int) -> str:
        """Return the signal type for the given row (combo widget or item text)."""
        widget = self.signals_table.cellWidget(row, 1)
        if isinstance(widget, QComboBox):
            return widget.currentText()
        item = self.signals_table.item(row, 1)
        return (item.text() if item else "").strip()

    def _update_resulting_signal_for_combo(self, combo: "QComboBox"):
        """Find the current row of *combo* and recompute its Resulting Signal cell."""
        for r in range(self.signals_table.rowCount()):
            if self.signals_table.cellWidget(r, 1) is combo:
                self._update_resulting_signal_for_row(r)
                return

    def _update_resulting_signal_for_row(self, row: int):
        """Recompute the Resulting Signal cell for *row*."""
        sig_type  = self._get_row_signal_type(row)
        count_item = self.signals_table.item(row, 3)
        count_str  = (count_item.text() if count_item else "1").strip()
        try:
            count = int(count_str) if count_str else 1
        except ValueError:
            count = 1
        resulting   = self._calculate_resulting_signal(sig_type, count)
        result_item = self.signals_table.item(row, 6)
        if result_item:
            result_item.setText(resulting)
        self._update_composition_display()

    def _update_resulting_signal(self, item: QTableWidgetItem):
        """Update resulting signal when count (col 3) changes."""
        if item.column() == 3:
            self._update_resulting_signal_for_row(item.row())
        else:
            self._update_composition_display()

    def _update_composition_display(self):
        """Update the composition display (e.g., "2HDI 1HDO")."""
        signal_counts: dict[str, int] = {}
        for r in range(self.signals_table.rowCount()):
            sig_type  = self._get_row_signal_type(r)
            count_item = self.signals_table.item(r, 3)
            count_str  = (count_item.text() if count_item else "1").strip()
            if sig_type:
                try:
                    count = int(count_str) if count_str else 1
                except ValueError:
                    count = 1
                signal_counts[sig_type] = signal_counts.get(sig_type, 0) + count

        parts = [f"{signal_counts[t]}{t}" for t in sorted(signal_counts)]
        self.composition_display.setText(" ".join(parts))

    def _new_composition(self):
        """Create a new blank composition."""
        self._clear_form()
        self.title_edit.setFocus()
    
    def _delete_composition(self):
        """Delete the selected typical."""
        items = self.comp_tree.selectedItems()
        if not items or items[0].data(0, Qt.ItemDataRole.UserRole) is None:
            QMessageBox.warning(self, "No Selection", "Please select a typical to delete.")
            return

        comp_name = items[0].text(0)
        ans = QMessageBox.question(
            self, "Delete Typical",
            f"Delete '{comp_name}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)

        if ans == QMessageBox.StandardButton.Yes:
            comp_id = items[0].data(0, Qt.ItemDataRole.UserRole)
            db_delete_signal_composition(comp_id)
            self._compositions = [c for c in self._compositions if c["id"] != comp_id]
            self._populate_tree()
            self._clear_form()

    # ── Category management ───────────────────────────────────────────────────
    def _selected_category(self) -> str | None:
        """Return the category name of the currently selected tree item, or None."""
        items = self.comp_tree.selectedItems()
        if not items:
            return None
        item = items[0]
        # If it's a leaf (typical), use its parent's text
        if item.data(0, Qt.ItemDataRole.UserRole) is not None:
            parent = item.parent()
            if parent:
                name = parent.text(0)
                return "" if name == "No group assigned" else name
            return None
        # It's a category header
        name = item.text(0)
        return "" if name == "No group assigned" else name

    def _all_category_names(self) -> list[str]:
        """Return sorted list of all category names currently in the tree."""
        cats = set()
        for comp in self._compositions:
            cats.add(comp.get("category", "").strip())
        return sorted(cats)

    def _add_category(self):
        name, ok = QInputDialog.getText(
            self, "Add Category", "New category name:")
        if not ok:
            return
        name = name.strip()
        if not name:
            QMessageBox.warning(self, "Invalid Name", "Category name cannot be empty.")
            return
        if name in self._all_category_names():
            QMessageBox.information(self, "Exists",
                                    f"Category '{name}' already exists.")
            return
        # Add a placeholder typical to anchor the category, OR just rebuild tree
        # with no typicals — we store category as a text field on compositions.
        # Easiest: just refresh; user can create a new typical and assign this cat.
        # To show the empty category we need a sentinel — rebuild tree with it.
        self._pending_new_categories = getattr(self, "_pending_new_categories", set())
        self._pending_new_categories.add(name)
        self._populate_tree()
        # Pre-fill Category field so the next new typical gets this category
        self.cat_edit.setText(name)

    def _rename_category(self):
        old_name = self._selected_category()
        if old_name is None:
            QMessageBox.warning(self, "No Selection",
                                "Please click a category header to rename.")
            return
        display = old_name if old_name else "No group assigned"
        new_name, ok = QInputDialog.getText(
            self, "Rename Category", f"Rename '{display}' to:",
            text=old_name)
        if not ok:
            return
        new_name = new_name.strip()
        if not new_name or new_name == old_name:
            return
        # Update all compositions in this category
        for comp in self._compositions:
            if comp.get("category", "").strip() == old_name:
                comp["category"] = new_name
        # Update cat_edit if it currently shows the old name
        if self.cat_edit.text().strip() == old_name:
            self.cat_edit.setText(new_name)
        self._populate_tree()

    def _delete_category(self):
        old_name = self._selected_category()
        if old_name is None:
            QMessageBox.warning(self, "No Selection",
                                "Please click a category header to delete.")
            return
        display = old_name if old_name else "No group assigned"
        members = [c for c in self._compositions
                   if c.get("category", "").strip() == old_name]
        if members:
            ans = QMessageBox.question(
                self, "Delete Category",
                f"Category '{display}' has {len(members)} typical(s).\n"
                "Move them to 'No group assigned' and delete the category?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if ans != QMessageBox.StandardButton.Yes:
                return
            for comp in members:
                comp["category"] = ""
        else:
            # Remove from pending set if present
            pnc = getattr(self, "_pending_new_categories", set())
            pnc.discard(old_name)
        self._populate_tree()

    def _schedule_tree_sync(self):
        """
        Schedule a category-sync for after the current drag-drop operation
        completes.  rowsInserted fires before Qt removes the item from its
        original position (InternalMove), so deferring to the next event-loop
        iteration guarantees the tree is in its final state.
        """
        QTimer.singleShot(0, self._sync_categories_from_tree)

    def _sync_categories_from_tree(self):
        """
        Walk every category item in the tree and update each composition's
        category field to match where it now lives.  Then rebuild the tree so
        that all items carry the correct drag-enable flags (Qt's InternalMove
        clones items without preserving custom flags).
        """
        root = self.comp_tree.invisibleRootItem()
        # Preserve all categories currently visible in the tree (even those that
        # become empty after a typical is dragged out of them) so that a move
        # operation does not silently delete the source category.
        pnc = getattr(self, "_pending_new_categories", set())
        for ci in range(root.childCount()):
            cat_label = root.child(ci).text(0)
            if cat_label != "No group assigned":
                pnc.add(cat_label)
        self._pending_new_categories = pnc

        for ci in range(root.childCount()):
            cat_item = root.child(ci)
            cat_label = cat_item.text(0)
            cat_value = "" if cat_label == "No group assigned" else cat_label
            for ti in range(cat_item.childCount()):
                comp_item = cat_item.child(ti)
                comp_id = comp_item.data(0, Qt.ItemDataRole.UserRole)
                if comp_id is None:
                    continue
                for comp in self._compositions:
                    if comp["id"] == comp_id:
                        comp["category"] = cat_value
                        break
        # Rebuild tree to restore correct item flags for all items.
        self._populate_tree()

    def _add_desc_field(self, text: str = "") -> None:
        """Add a description line field (max 5)."""
        if len(self._desc_fields) >= 5:
            return
        row_lay = QHBoxLayout()
        row_lay.setContentsMargins(0, 0, 0, 0)
        edit = QLineEdit(text)
        edit.setPlaceholderText(f"Description {len(self._desc_fields) + 1}…")
        rm_btn = QPushButton("✕")
        rm_btn.setFixedWidth(24)
        rm_btn.setStyleSheet(
            "QPushButton { background:#3A1010; color:#FF8A80; border:none;"
            " border-radius:3px; font-size: 8pt; }"
            "QPushButton:hover { background:#5A1A1A; }")
        rm_btn.clicked.connect(lambda: self._remove_desc_field(row_lay, edit))
        row_lay.addWidget(edit)
        row_lay.addWidget(rm_btn)
        self._desc_layout.addLayout(row_lay)
        self._desc_fields.append(edit)
        self._desc_buttons.append(rm_btn)
        # Show remove buttons on all fields only when there are 2 or more
        show = len(self._desc_fields) > 1
        for btn in self._desc_buttons:
            btn.setVisible(show)
        if hasattr(self, '_add_desc_btn'):
            self._add_desc_btn.setEnabled(len(self._desc_fields) < 5)

    def _remove_desc_field(self, row_lay: QHBoxLayout, edit: QLineEdit) -> None:
        """Remove a description line field (keep at least one)."""
        if len(self._desc_fields) <= 1:
            return
        idx = self._desc_fields.index(edit)
        self._desc_fields.remove(edit)
        self._desc_buttons.pop(idx)
        while row_lay.count():
            item = row_lay.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        self._desc_layout.removeItem(row_lay)
        for i, e in enumerate(self._desc_fields):
            e.setPlaceholderText(f"Description {i + 1}…")
        # Hide remove buttons if only one field remains
        show = len(self._desc_fields) > 1
        for btn in self._desc_buttons:
            btn.setVisible(show)
        if hasattr(self, '_add_desc_btn'):
            self._add_desc_btn.setEnabled(len(self._desc_fields) < 5)

    def _reset_desc_fields(self) -> None:
        """Completely clear and rebuild the description field area (one empty field)."""
        # Remove every row-layout and its child widgets from desc_layout
        while self._desc_layout.count():
            layout_item = self._desc_layout.takeAt(0)
            sub_lay = layout_item.layout()
            if sub_lay:
                while sub_lay.count():
                    widget_item = sub_lay.takeAt(0)
                    if widget_item.widget():
                        widget_item.widget().deleteLater()
        self._desc_fields = []
        self._desc_buttons = []
        if hasattr(self, '_add_desc_btn'):
            self._add_desc_btn.setEnabled(True)
        self._add_desc_field()  # always keep at least one slot

    def _add_extra_column(self, header: str = "") -> None:
        """Add a user-defined extra column to the right of the signals table.

        If *header* is empty the user is prompted for a column name.
        Existing rows automatically get an empty cell in the new column.
        """
        if not header:
            header, ok = QInputDialog.getText(
                self, "Add Column",
                "Enter column header (e.g., 'DI Comment', 'IO Remark'):")
            if not ok or not header.strip():
                return
            header = header.strip()
        col = self.signals_table.columnCount()
        self.signals_table.setColumnCount(col + 1)
        self.signals_table.setHorizontalHeaderItem(col, QTableWidgetItem(header))
        self.signals_table.horizontalHeader().setSectionResizeMode(
            col, QHeaderView.ResizeMode.Stretch)
        # Fill existing rows with empty text
        for r in range(self.signals_table.rowCount()):
            self.signals_table.setItem(r, col, QTableWidgetItem(""))

    def _remove_extra_column(self) -> None:
        """Remove a user-defined extra column from the signals table.

        If more than one extra column exists the user is prompted to choose which to remove.
        """
        extra_count = self.signals_table.columnCount() - self._FIXED_COL_COUNT
        if extra_count <= 0:
            QMessageBox.information(
                self, "No Extra Columns",
                "There are no extra columns to remove.\n"
                "Only the built-in fixed columns are present.")
            return

        # Collect extra column header names
        headers = []
        for c in range(self._FIXED_COL_COUNT, self.signals_table.columnCount()):
            item = self.signals_table.horizontalHeaderItem(c)
            headers.append(item.text() if item else f"Column {c - self._FIXED_COL_COUNT + 1}")

        if extra_count == 1:
            col_to_remove = self._FIXED_COL_COUNT
        else:
            choice, ok = QInputDialog.getItem(
                self, "Remove Column",
                "Select the column to remove:", headers, 0, False)
            if not ok:
                return
            col_to_remove = self._FIXED_COL_COUNT + headers.index(choice)

        self.signals_table.removeColumn(col_to_remove)

    def _use_template(self):
        """Load a template and use it as basis for new composition."""
        template_dlg = SignalCompositionTemplateDialog(self)
        if template_dlg.exec() != QDialog.DialogCode.Accepted:
            return
        
        # Ask user which template to use
        templates = template_dlg.get_templates()
        if not templates:
            QMessageBox.information(self, "No Templates", "No templates available.")
            return
        
        names = [t["title"] for t in templates.values()]
        choice, ok = QInputDialog.getItem(
            self, "Select Template",
            "Choose a template to use:", names, 0, False)
        
        if not ok:
            return
        
        # Find selected template
        template = next((t for t in templates.values() if t["title"] == choice), None)
        if template:
            # Ask if user wants to keep the same name or rename
            new_title, ok = QInputDialog.getText(
                self, "Composition Name",
                f"Template: {template['title']}\n\n"
                f"Enter name for the project-specific composition:\n"
                f"(You can use the same name as the template)",
                text=f"{template['title']}")
            
            if not ok or not new_title.strip():
                return
            
            self._clear_form()
            self.title_edit.setText(new_title.strip())
            self._desc_fields[0].setText(template.get("description", ""))

            # Populate signals from template
            self.signals_table.blockSignals(True)
            try:
                self.signals_table.setRowCount(0)
                for sig in template["signals"]:
                    r = self.signals_table.rowCount()
                    self.signals_table.insertRow(r)

                    self.signals_table.setItem(r, 0, QTableWidgetItem(sig["signal_name"]))
                    sig_type_val = sig.get("signal_type", "") or _SIGNAL_IO_TYPES[0]
                    cb = self._make_signal_type_combo(sig_type_val)
                    cb.currentTextChanged.connect(
                        lambda _txt, c=cb: self._update_resulting_signal_for_combo(c))
                    self.signals_table.setCellWidget(r, 1, cb)
                    sig_desc = sig.get("signal_description", "") or "NA"
                    self.signals_table.setItem(r, 2, QTableWidgetItem(sig_desc))
                    self.signals_table.setItem(r, 3, QTableWidgetItem("1"))
                    self.signals_table.setItem(r, 4, QTableWidgetItem("NA"))
                    self.signals_table.setItem(r, 5, QTableWidgetItem("NA"))

                    resulting = self._calculate_resulting_signal(cb.currentText(), 1)
                    result_item = QTableWidgetItem(resulting)
                    result_item.setFlags(result_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    self.signals_table.setItem(r, 6, result_item)
            finally:
                self.signals_table.blockSignals(False)

            self._current_comp_id = None  # Mark as new composition
            self._update_composition_display()

    def _on_accept(self):
        """Validate, save changes and close the dialog."""
        if self._do_save():
            self.accept()

    def _on_apply(self):
        """Validate and save changes without closing the dialog."""
        if self._do_save():
            self._compositions = db_load_compositions_by_owner(self._owner_id)
            self._populate_tree()
            QMessageBox.information(self, "Saved", "Changes saved successfully.")

    def _do_save(self) -> bool:
        """
        Validate and persist all compositions to the database.
        Returns True on success, False if validation or a DB error occurred.
        """
        # Save current composition if one is being edited
        if self._current_comp_id is not None:
            self._save_current_composition()
        elif self.title_edit.text().strip():
            # Save as new composition if title is filled but no ID
            self._save_current_composition()
        
        # Validate all compositions
        for comp in self._compositions:
            if not comp["title"].strip():
                QMessageBox.warning(
                    self, "Validation Error",
                    "All compositions must have a title.")
                return False

            if not comp.get("control_module", "").strip():
                QMessageBox.warning(
                    self, "Validation Error",
                    f"Composition '{comp['title']}': Control Module Name is mandatory. "
                    "Enter a value or use NA.")
                return False

            if not comp.get("field_device", "").strip():
                QMessageBox.warning(
                    self, "Validation Error",
                    f"Composition '{comp['title']}': Field Device Name is mandatory. "
                    "Enter a value or use NA.")
                return False

            if not comp["signals"]:
                QMessageBox.warning(
                    self, "Validation Error",
                    f"Composition '{comp['title']}' must have at least one signal.")
                return False
            
            # Check mandatory fields in each signal
            for sig in comp["signals"]:
                if not sig.get("signal_name") or not sig.get("signal_type") or not sig.get("signal_description"):
                    QMessageBox.warning(
                        self, "Validation Error",
                        f"In composition '{comp['title']}': Signal Name, Type, and Description are mandatory.")
                    return False
                if not sig.get("prefix", "").strip():
                    QMessageBox.warning(
                        self, "Validation Error",
                        f"In composition '{comp['title']}': Prefix is mandatory per signal. "
                        "Enter a value or use NA.")
                    return False
                if not sig.get("suffix", "").strip():
                    QMessageBox.warning(
                        self, "Validation Error",
                        f"In composition '{comp['title']}': Suffix is mandatory per signal. "
                        "Enter a value or use NA.")
                    return False
        
        # Save to database
        for comp in self._compositions:
            try:
                if comp["id"] is None:
                    # Create new composition
                    comp["id"] = db_save_signal_composition(
                        title=comp["title"],
                        description=comp["description"],
                        signals=comp["signals"],
                        control_module=comp.get("control_module", "NA"),
                        field_device=comp.get("field_device", "NA"),
                        extra_column_headers=comp.get("extra_column_headers", []),
                        cm_type=comp.get("cm_type", "NA"),
                        cm_description=comp.get("cm_description", "NA"),
                        fd_type=comp.get("fd_type", "NA"),
                        fd_description=comp.get("fd_description", "NA"),
                        category=comp.get("category", "")
                    )
                    # Assign to owner
                    db_assign_composition_to_owner(comp["id"], self._owner_id)
                else:
                    # Update existing composition
                    db_update_signal_composition(
                        composition_id=comp["id"],
                        title=comp["title"],
                        description=comp["description"],
                        signals=comp["signals"],
                        control_module=comp.get("control_module", "NA"),
                        field_device=comp.get("field_device", "NA"),
                        extra_column_headers=comp.get("extra_column_headers", []),
                        cm_type=comp.get("cm_type", "NA"),
                        cm_description=comp.get("cm_description", "NA"),
                        fd_type=comp.get("fd_type", "NA"),
                        fd_description=comp.get("fd_description", "NA"),
                        category=comp.get("category", "")
                    )
            except Exception as e:
                QMessageBox.warning(self, "Error", str(e))
                return False

        return True
        
    def _save_current_composition(self):
        """Save the currently edited composition into self._compositions."""
        title = self.title_edit.text().strip()
        if not title:
            return
        
        # Collect extra column headers from the table
        extra_column_headers = []
        for ec in range(self._FIXED_COL_COUNT, self.signals_table.columnCount()):
            hdr_item = self.signals_table.horizontalHeaderItem(ec)
            extra_column_headers.append(hdr_item.text() if hdr_item else "")
        
        # Collect signals (fixed fields + per-row extra column values)
        signals = []
        for r in range(self.signals_table.rowCount()):
            sig_name = (self.signals_table.item(r, 0) or QTableWidgetItem()).text().strip()
            sig_type = self._get_row_signal_type(r)
            sig_desc = (self.signals_table.item(r, 2) or QTableWidgetItem()).text().strip()
            count_str = (self.signals_table.item(r, 3) or QTableWidgetItem()).text().strip()
            prefix = (self.signals_table.item(r, 4) or QTableWidgetItem()).text().strip() or "NA"
            suffix = (self.signals_table.item(r, 5) or QTableWidgetItem()).text().strip() or "NA"
            extra_column_values = [
                (self.signals_table.item(r, self._FIXED_COL_COUNT + ec) or QTableWidgetItem()).text().strip()
                for ec in range(len(extra_column_headers))
            ]
            
            if sig_name and sig_type:
                try:
                    count = int(count_str) if count_str else 1
                except ValueError:
                    count = 1
                
                signals.append({
                    "signal_name": sig_name,
                    "signal_type": sig_type,
                    "signal_description": sig_desc,  # may be empty; validated separately
                    "count": count,
                    "prefix": prefix,
                    "suffix": suffix,
                    "extra_column_values": extra_column_values,
                })
        
        if not signals:
            return
        
        # Find the composition to update
        comp_to_update = None
        
        # Check if we're editing an existing composition (has an ID)
        if self._current_comp_id is not None:
            comp_to_update = next(
                (c for c in self._compositions if c["id"] == self._current_comp_id),
                None
            )
        
        # If not found, create a new entry
        if comp_to_update is None:
            comp_to_update = {
                "id": None,  # New composition - no ID yet
                "title": title,
                "category": "",
                "description": "",
                "control_module": "NA",
                "cm_type": "NA",
                "cm_description": "NA",
                "field_device": "NA",
                "fd_type": "NA",
                "fd_description": "NA",
                "extra_column_headers": [],
                "signals": []
            }
            self._compositions.append(comp_to_update)
        
        # Update the composition
        comp_to_update["title"] = title
        comp_to_update["category"] = self.cat_edit.text().strip()
        description = "\n".join(
            e.text().strip() for e in self._desc_fields if e.text().strip())
        comp_to_update["description"] = description
        comp_to_update["control_module"] = (self.cm_table.item(0, 0).text().strip() or "NA")
        comp_to_update["cm_type"]        = "CM"  # always fixed
        comp_to_update["cm_description"] = (self.cm_table.item(0, 2).text().strip() or "NA")
        comp_to_update["field_device"]    = (self.fd_table.item(0, 0).text().strip() or "NA")
        comp_to_update["fd_type"]        = (self.fd_table.item(0, 1).text().strip() or "NA")
        comp_to_update["fd_description"] = (self.fd_table.item(0, 2).text().strip() or "NA")
        comp_to_update["extra_column_headers"] = extra_column_headers
        comp_to_update["signals"] = signals

        self._current_comp_id = comp_to_update["id"]
# ---------------------------------------------------------------------------
# ProjectMetadataDialog — create or edit a project
# ---------------------------------------------------------------------------
class ProjectMetadataDialog(QDialog):
    def __init__(self, project: dict | None = None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("New Project" if project is None else "Edit Project")
        self.setMinimumWidth(420)

        p = project or {}
        self._name_edit = QLineEdit(p.get("name", ""))
        self._num_edit  = QLineEdit(p.get("number", ""))
        self._desc_edit = QTextEdit(p.get("description", ""))
        self._desc_edit.setFixedHeight(90)
        self._name_edit.setPlaceholderText("e.g. Cooling Water System")
        self._num_edit.setPlaceholderText("e.g. PRJ-001")
        self._desc_edit.setPlaceholderText("Optional description…")

        form = QFormLayout()
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        form.setVerticalSpacing(8)
        form.addRow("<b>Project Name:</b>",   self._name_edit)
        form.addRow("<b>Project Number:</b>", self._num_edit)
        form.addRow("<b>Description:</b>",    self._desc_edit)

        if project:
            info = QLabel(
                f"Created: {p.get('created','')}   |   "
                f"Modified: {p.get('modified','')}")
            info.setStyleSheet("color:#AAAAAA; font-size: 8pt;")
            form.addRow("", info)

        bb = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok |
                              QDialogButtonBox.StandardButton.Cancel)
        bb.accepted.connect(self._on_accept)
        bb.rejected.connect(self.reject)

        lay = QVBoxLayout()
        lay.setContentsMargins(14, 14, 14, 10)
        lay.addLayout(form)
        lay.addWidget(bb)
        self.setLayout(lay)

    def _on_accept(self):
        if not self._name_edit.text().strip():
            QMessageBox.warning(self, "Required", "Project name cannot be empty.")
            return
        self.accept()

    @property
    def values(self) -> dict:
        return {
            "name":        self._name_edit.text().strip(),
            "number":      self._num_edit.text().strip(),
            "description": self._desc_edit.toPlainText().strip(),
        }

# ---------------------------------------------------------------------------
# FileMetadataDialog — edit per-file Technical Drawing metadata
# ---------------------------------------------------------------------------
class FileMetadataDialog(QDialog):
    """Dialog to set Technical Drawing Number and Description for a project file."""

    def __init__(self, file_id: int, file_name: str, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"File Metadata — {file_name}")
        self.setMinimumWidth(420)
        self._file_id = file_id

        meta = db_get_file_metadata(file_id)

        # Default the drawing number to the PDF filename (without extension)
        default_number = meta.get("drw_name", "") or os.path.splitext(file_name)[0]
        self._name_edit   = QLineEdit(default_number)
        self._number_edit = QLineEdit(meta.get("drw_number", ""))
        self._name_edit.setPlaceholderText("e.g. DWG-001")
        self._number_edit.setPlaceholderText("e.g. Cooling Water System P&ID")

        form = QFormLayout()
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        form.setVerticalSpacing(8)
        form.addRow("<b>Technical Drawing Number:</b>", self._name_edit)
        form.addRow("<b>Technical Drawing Description:</b>", self._number_edit)

        bb = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok |
                              QDialogButtonBox.StandardButton.Cancel)
        bb.accepted.connect(self._on_accept)
        bb.rejected.connect(self.reject)

        lay = QVBoxLayout()
        lay.setContentsMargins(14, 14, 14, 10)
        lay.addLayout(form)
        lay.addWidget(bb)
        self.setLayout(lay)

    def _on_accept(self):
        db_save_file_metadata(
            self._file_id,
            self._name_edit.text().strip(),
            self._number_edit.text().strip(),
        )
        self.accept()

# ---------------------------------------------------------------------------
# ProjectPanel — the sidebar tree widget
# ---------------------------------------------------------------------------
class ProjectPanel(QWidget):
    # Emitted when the user double-clicks a PDF node
    open_file_requested = Signal(str)
    # Emitted when user picks "Export Project IO List" from project context menu
    export_project_io_requested = Signal(int)  # project_id

    # Role used to store data on tree items
    _ROLE_KIND     = Qt.ItemDataRole.UserRole        # "project"|"file"|"unassigned_header"|"unassigned_file"
    _ROLE_ID       = Qt.ItemDataRole.UserRole + 1   # DB id
    _ROLE_PATH     = Qt.ItemDataRole.UserRole + 2   # file_path (files only)
    _ROLE_MISSING  = Qt.ItemDataRole.UserRole + 3   # bool

    # Tracks PDFs opened outside any project (shown in Unassigned section)
    _unassigned_paths: list = []

    def __init__(self, parent=None):
        super().__init__(parent)
        self._theme = "dark"
        self._build_ui()
        self._validation_timer = QTimer(self)
        self._validation_timer.setSingleShot(True)
        self._validation_timer.timeout.connect(self._validate_files)
        self.refresh()

    def set_theme(self, theme: str) -> None:
        """Update the active theme and refresh tree colours."""
        self._theme = theme
        self.refresh()

    def _file_text_color(self) -> QColor:
        """Return the appropriate text colour for file items in the current theme."""
        return QColor("#F0F0F0") if self._theme == "dark" else QColor("#1A1A1A")

    def _build_ui(self):
        # Toolbar buttons
        self._btn_new = QPushButton("＋ New Project")
        self._btn_new.setToolTip("Create a new empty project")
        self._btn_new.setStyleSheet(
            "QPushButton { background:#1F4E79; color:#FFFFFF; border-radius:4px;"
            "padding:4px 10px; font-weight:bold; border:none; }"
            "QPushButton:hover { background:#2E75B6; }")

        self._btn_import = QPushButton("📁 Import Folder")
        self._btn_import.setToolTip("Import a folder as a new project")
        self._btn_import.setStyleSheet(
            "QPushButton { background:#1B5E20; color:#FFFFFF; border-radius:4px;"
            "padding:4px 10px; font-weight:bold; border:none; }"
            "QPushButton:hover { background:#2E7D32; }")

        self._tree = QTreeWidget()
        self._tree.setHeaderHidden(True)
        self._tree.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self._tree.setSelectionMode(
            QAbstractItemView.SelectionMode.SingleSelection)
        self._tree.setAnimated(True)
        self._tree.setIndentation(16)

        # Use NoDragDrop on the tree so Qt never auto-reparents items.
        # We start drags manually via a viewport event filter.
        self._tree.setDragEnabled(False)
        self._tree.setDragDropMode(QAbstractItemView.DragDropMode.NoDragDrop)
        self.setAcceptDrops(True)
        self._tree.setAcceptDrops(True)
        self._tree.viewport().setAcceptDrops(True)

        # Tracks which unassigned_file item is currently being dragged
        self._dragged_item: "QTreeWidgetItem | None" = None
        self._drag_start_pos = None

        # Install event filter on viewport to manually start drags
        self._tree.viewport().installEventFilter(self)

        # Drop-hint label shown at the bottom while dragging
        self._drop_hint = QLabel("  📂  Drop PDF(s) onto a project to add them")
        self._drop_hint.setStyleSheet(
            "color:#7EC8F0; background:#1A3A5C; font-size: 8pt;"
            "padding:4px 8px; border-radius:3px;")
        self._drop_hint.setVisible(False)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(4)
        btn_row.addWidget(self._btn_new)
        btn_row.addWidget(self._btn_import)

        lay = QVBoxLayout()
        lay.setContentsMargins(4, 4, 4, 4)
        lay.setSpacing(4)
        lay.addLayout(btn_row)
        lay.addWidget(self._tree, stretch=1)
        lay.addWidget(self._drop_hint)
        self.setLayout(lay)

        self._btn_new.clicked.connect(self._on_new_project)
        self._btn_import.clicked.connect(self._on_import_project_folder)
        self._tree.itemDoubleClicked.connect(self._on_item_double_clicked)
        self._tree.customContextMenuRequested.connect(self._on_context_menu)

    # ── Drag-and-drop ─────────────────────────────────────────────────────
    def eventFilter(self, obj, event):
        """
        Intercept mouse events on the tree viewport to manually start a
        QDrag for unassigned_file OR project file items.
        Two custom MIME types are used:
          application/x-unassigned-pdf  — path only (file not in any project)
          application/x-project-pdf     — "<file_id>|<path>" (file in a project)
        """
        from PySide6.QtCore import QMimeData, QByteArray
        from PySide6.QtGui import QDrag
        from PySide6.QtCore import Qt as _Qt

        if obj is not self._tree.viewport():
            return super().eventFilter(obj, event)

        if event.type() == event.Type.MouseButtonPress:
            if event.button() == _Qt.MouseButton.LeftButton:
                item = self._tree.itemAt(event.position().toPoint())
                if item is not None and item.data(0, self._ROLE_KIND) in (
                        "unassigned_file", "file"):
                    self._drag_start_pos = event.position().toPoint()
                    self._dragged_item   = item
                else:
                    self._drag_start_pos = None
                    self._dragged_item   = None

        elif event.type() == event.Type.MouseMove:
            if (self._drag_start_pos is not None
                    and self._dragged_item is not None
                    and event.buttons() & _Qt.MouseButton.LeftButton):
                dist = (event.position().toPoint()
                        - self._drag_start_pos).manhattanLength()
                if dist >= 8:
                    mime = QMimeData()
                    kind = self._dragged_item.data(0, self._ROLE_KIND)
                    if kind == "unassigned_file":
                        path = self._dragged_item.data(0, self._ROLE_PATH)
                        mime.setData(
                            "application/x-unassigned-pdf",
                            QByteArray(path.encode("utf-8")))
                    else:  # "file" — belongs to a project
                        file_id = self._dragged_item.data(0, self._ROLE_ID)
                        path    = self._dragged_item.data(0, self._ROLE_PATH)
                        payload = f"{file_id}|{path}"
                        mime.setData(
                            "application/x-project-pdf",
                            QByteArray(payload.encode("utf-8")))
                    drag = QDrag(self._tree)
                    drag.setMimeData(mime)
                    self._drag_start_pos = None
                    drag.exec(_Qt.DropAction.MoveAction)
                    return True

        elif event.type() == event.Type.MouseButtonRelease:
            self._drag_start_pos = None

        return super().eventFilter(obj, event)

    def _pdf_urls_from_event(self, event) -> list[str]:
        """Return a list of local PDF file paths from a drag event, or []."""
        if not event.mimeData().hasUrls():
            return []
        return [
            u.toLocalFile() for u in event.mimeData().urls()
            if u.isLocalFile() and u.toLocalFile().lower().endswith(".pdf")
        ]

    def _is_internal_drag(self, event) -> bool:
        """True when this is our custom unassigned-file drag."""
        return event.mimeData().hasFormat("application/x-unassigned-pdf")

    def _path_from_internal_drag(self, event) -> str:
        """Extract the path from an unassigned-file drag."""
        return event.mimeData().data(
            "application/x-unassigned-pdf").toStdString()

    def _is_project_pdf_drag(self, event) -> bool:
        """True when this is a project file being dragged between projects."""
        return event.mimeData().hasFormat("application/x-project-pdf")

    def _data_from_project_pdf_drag(self, event) -> tuple[int, str]:
        """Return (file_id, path) from a project-pdf drag."""
        raw = event.mimeData().data(
            "application/x-project-pdf").toStdString()
        file_id_str, path = raw.split("|", 1)
        return int(file_id_str), path

    def _any_drag(self, event) -> bool:
        """True if the event carries any kind of drag this panel handles."""
        return (self._is_internal_drag(event)
                or self._is_project_pdf_drag(event)
                or bool(self._pdf_urls_from_event(event)))

    def _project_item_at(self, pos) -> "QTreeWidgetItem | None":
        """
        Return the project-level QTreeWidgetItem at `pos` in the tree viewport,
        or the parent project item if pos is over a file child, or None.
        """
        item = self._tree.itemAt(pos)
        if item is None:
            return None
        if item.data(0, self._ROLE_KIND) == "project":
            return item
        if item.data(0, self._ROLE_KIND) == "file":
            return item.parent()   # bubble up to the project row
        return None

    def dragEnterEvent(self, event):
        if self._any_drag(event):
            event.acceptProposedAction()
            if self._is_project_pdf_drag(event):
                self._drop_hint.setText(
                    "  🔀  Drop onto a project to move the file there")
            else:
                self._drop_hint.setText(
                    "  📂  Drop PDF(s) onto a project to add them")
            self._drop_hint.setVisible(True)
        else:
            event.ignore()

    def dragMoveEvent(self, event):
        if not self._any_drag(event):
            event.ignore()
            return
        tree_pos = self._tree.viewport().mapFromGlobal(
            self.mapToGlobal(event.position().toPoint()))
        target = self._project_item_at(tree_pos)
        from PySide6.QtCore import QItemSelectionModel
        if target:
            self._tree.selectionModel().clearSelection()
            self._tree.selectionModel().select(
                self._tree.indexFromItem(target),
                QItemSelectionModel.SelectionFlag.Select)
        else:
            self._tree.selectionModel().clearSelection()
        event.acceptProposedAction()

    def dragLeaveEvent(self, event):
        self._drop_hint.setVisible(False)
        self._dragged_item = None
        self._tree.selectionModel().clearSelection()

    def dropEvent(self, event):
        self._drop_hint.setVisible(False)
        self._tree.selectionModel().clearSelection()

        tree_pos = self._tree.viewport().mapFromGlobal(
            self.mapToGlobal(event.position().toPoint()))
        target_item = self._project_item_at(tree_pos)

        # ── Case 1: unassigned_file dragged from the Unassigned section ─────
        if self._is_internal_drag(event):
            self._dragged_item = None
            if target_item is None:
                event.ignore()
                return
            path = self._path_from_internal_drag(event)
            if not path:
                event.ignore()
                return
            event.acceptProposedAction()
            self._assign_to_project(
                [path],
                target_item.data(0, self._ROLE_ID),
                target_item.text(0).lstrip("📁").strip())
            return

        # ── Case 2: project file dragged from one project to another ─────
        if self._is_project_pdf_drag(event):
            self._dragged_item = None
            if target_item is None:
                event.ignore()
                return
            file_id, path = self._data_from_project_pdf_drag(event)
            target_project_id = target_item.data(0, self._ROLE_ID)
            # Find which project currently owns this file_id
            source_project_id = None
            for p in db_load_projects():
                for f in db_load_project_files(p["id"]):
                    if f["id"] == file_id:
                        source_project_id = p["id"]
                        break
                if source_project_id is not None:
                    break
            # Ignore drop onto the same project
            if source_project_id == target_project_id:
                event.ignore()
                return
            event.acceptProposedAction()
            target_name = target_item.text(0).lstrip("📁").strip()
            # Remove from source project and add to target
            db_remove_project_file(file_id)
            db_add_project_file(target_project_id, path)
            self.refresh()
            QMessageBox.information(
                self, "File moved",
                f"<b>{os.path.basename(path)}</b> moved to "
                f"project <b>{target_name}</b>.")
            return

        # ── Case 3: external PDF file(s) from Explorer / Finder ──────────
        self._dragged_item = None
        pdfs = self._pdf_urls_from_event(event)
        if not pdfs:
            event.ignore()
            return

        event.acceptProposedAction()

        if target_item is not None:
            project_id   = target_item.data(0, self._ROLE_ID)
            project_name = target_item.text(0).lstrip("📁").strip()
        else:
            projects = db_load_projects()
            if not projects:
                QMessageBox.information(
                    self, "No projects",
                    "Create a project first before adding files.")
                return
            if len(projects) == 1:
                project_id   = projects[0]["id"]
                project_name = projects[0]["name"]
            else:
                names = [p["name"] + (f"  ({p['number']})" if p["number"] else "")
                         for p in projects]
                choice, ok = QInputDialog.getItem(
                    self, "Choose project",
                    f"Which project should the {len(pdfs)} file(s) be added to?",
                    names, 0, False)
                if not ok:
                    return
                idx = names.index(choice)
                project_id   = projects[idx]["id"]
                project_name = projects[idx]["name"]

        self._assign_to_project(
            [os.path.normpath(p) for p in pdfs],
            project_id, project_name)

    def _assign_to_project(self, paths: list, project_id: int,
                           project_name: str) -> None:
        """Add `paths` to `project_id`, remove from Unassigned, refresh tree."""
        added, skipped = [], []
        for path in paths:
            path = os.path.normpath(path)
            existing = [
                os.path.normpath(f["file_path"])
                for f in db_load_project_files(project_id)
            ]
            if path in existing:
                skipped.append(os.path.basename(path))
            else:
                db_add_project_file(project_id, path)
                added.append(os.path.basename(path))
            if path in self._unassigned_paths:
                self._unassigned_paths.remove(path)

        self.refresh()

        parts = []
        if added:
            parts.append(f"{len(added)} file(s) added to '{project_name}'")
        if skipped:
            parts.append(f"{len(skipped)} already present (skipped)")
        if parts:
            QMessageBox.information(self, "Files added", "\n".join(parts))

    # ── Public query helpers ──────────────────────────────────────────────
    def get_all_project_file_paths(self) -> list[dict]:
        """
        Return all file paths registered in any project as a flat list of
        dicts: [{path, project_name, project_id}].
        Used by PDFViewer.open_pdf to check if a file is already in a project.
        Reads directly from the DB so it is always up-to-date.
        """
        results = []
        for p in db_load_projects():
            for f in db_load_project_files(p["id"]):
                results.append({
                    "path":         f["file_path"],
                    "project_name": p["name"],
                    "project_id":   p["id"],
                })
        return results

    def highlight_file(self, path: str) -> None:
        """
        Expand and select the tree item matching `path`, making it visible.
        Called after the user chooses to open a file via the project tree.
        """
        root = self._tree.invisibleRootItem()
        for i in range(root.childCount()):
            p_item = root.child(i)
            for j in range(p_item.childCount()):
                f_item = p_item.child(j)
                if f_item.data(0, self._ROLE_PATH) == path:
                    self._tree.setCurrentItem(f_item)
                    self._tree.scrollToItem(f_item)
                    return

    # ── Populate tree ─────────────────────────────────────────────────────
    def refresh(self):
        """Reload all projects and files from DB, then rebuild the Unassigned section."""
        self._tree.clear()

        # ── Unassigned section (top) ─────────────────────────────────────
        # Show PDFs that were opened outside any project so the user can
        # drag-and-drop them into a project below.
        if self._unassigned_paths:
            hdr = QTreeWidgetItem()
            hdr.setText(0, "📂  Unassigned (drag into a project)")
            hdr.setData(0, self._ROLE_KIND, "unassigned_header")
            hdr.setFlags(Qt.ItemFlag.ItemIsEnabled)   # not selectable / draggable
            font = QFont("Arial", 9)
            font.setItalic(True)
            hdr.setFont(0, font)
            hdr.setForeground(0, QColor("#7EC8F0"))
            self._tree.addTopLevelItem(hdr)

            for path in self._unassigned_paths:
                f_item = QTreeWidgetItem(hdr)
                f_item.setText(0, f"📄  {os.path.basename(path)}")
                f_item.setData(0, self._ROLE_KIND, "unassigned_file")
                f_item.setData(0, self._ROLE_PATH, path)
                f_item.setForeground(0, self._file_text_color())
                f_item.setToolTip(0, path)
                # Draggable but never shows a spurious expand arrow
                f_item.setFlags(
                    Qt.ItemFlag.ItemIsEnabled
                    | Qt.ItemFlag.ItemIsSelectable
                    | Qt.ItemFlag.ItemIsDragEnabled)
                f_item.setChildIndicatorPolicy(
                    QTreeWidgetItem.ChildIndicatorPolicy.DontShowIndicator)
            hdr.setExpanded(True)

        # ── Project sections ─────────────────────────────────────────────
        projects = db_load_projects()
        for p in projects:
            p_item = QTreeWidgetItem()
            p_item.setText(0, f"📁  {p['name']}"
                           + (f"  ({p['number']})" if p['number'] else ""))
            p_item.setData(0, self._ROLE_KIND, "project")
            p_item.setData(0, self._ROLE_ID, p["id"])
            p_item.setToolTip(0, p.get("description") or p["name"])
            font = QFont("Arial", 9)
            font.setBold(True)
            p_item.setFont(0, font)
            # Allow drops onto project items
            p_item.setFlags(p_item.flags() | Qt.ItemFlag.ItemIsDropEnabled)
            self._tree.addTopLevelItem(p_item)

            files = db_load_project_files(p["id"])
            for f in files:
                f_item = QTreeWidgetItem(p_item)
                f_item.setData(0, self._ROLE_KIND, "file")
                f_item.setData(0, self._ROLE_ID, f["id"])
                f_item.setData(0, self._ROLE_PATH, f["file_path"])
                f_item.setChildIndicatorPolicy(
                    QTreeWidgetItem.ChildIndicatorPolicy.DontShowIndicator)
                self._set_file_item_display(f_item, f["file_name"],
                                             f["file_path"])
            p_item.setExpanded(True)
        # Schedule file-existence check after tree is drawn
        self._validation_timer.start(200)

    # ── Unassigned file management ────────────────────────────────────────
    def show_unassigned_file(self, path: str) -> None:
        """
        Add `path` to the Unassigned section at the top of the tree.
        Called by PDFViewer when a PDF is opened outside any project.
        """
        path = os.path.normpath(path)
        if path not in self._unassigned_paths:
            self._unassigned_paths.append(path)
        # Make the panel visible and rebuild
        self.refresh()

    def remove_unassigned_file(self, path: str) -> None:
        """
        Remove `path` from the Unassigned section (called after it is
        successfully dropped into a project, or when the tab is closed).
        """
        path = os.path.normpath(path)
        if path in self._unassigned_paths:
            self._unassigned_paths.remove(path)
        self.refresh()

    def _set_file_item_display(self, item: QTreeWidgetItem,
                                name: str, path: str):
        exists = os.path.isfile(path)
        item.setData(0, self._ROLE_MISSING, not exists)
        if exists:
            item.setText(0, f"📄  {name}")
            item.setForeground(0, self._file_text_color())
            item.setToolTip(0, path)
        else:
            item.setText(0, f"⚠️  {name}")
            item.setForeground(0, QColor("#FF8A50"))
            item.setToolTip(0, f"File not found:\n{path}")

    def _validate_files(self):
        """Re-check file existence for all file nodes."""
        root = self._tree.invisibleRootItem()
        for i in range(root.childCount()):
            p_item = root.child(i)
            for j in range(p_item.childCount()):
                f_item = p_item.child(j)
                path = f_item.data(0, self._ROLE_PATH)
                name = os.path.basename(path) if path else ""
                self._set_file_item_display(f_item, name, path)

    # ── Context menu ──────────────────────────────────────────────────────
    def _on_context_menu(self, pos):
        item = self._tree.itemAt(pos)
        global_pos = self._tree.viewport().mapToGlobal(pos)
        menu = QMenu(self)

        if item is None:
            # Clicked on empty space
            act_new    = menu.addAction("＋  New Project")
            act_import = menu.addAction("📁  Import Project Folder…")
            act_new.triggered.connect(self._on_new_project)
            act_import.triggered.connect(self._on_import_project_folder)

        elif item.data(0, self._ROLE_KIND) == "project":
            pid = item.data(0, self._ROLE_ID)
            act_edit = menu.addAction("✏️   Edit metadata…")
            menu.addSeparator()
            act_add_file   = menu.addAction("📄   Add PDF file…")
            act_add_folder = menu.addAction("📁   Add folder…")
            menu.addSeparator()
            act_export_io  = menu.addAction("📊   Export Project IO List…")
            menu.addSeparator()
            act_del = menu.addAction("🗑   Delete project")
            act_edit.triggered.connect(lambda: self._on_edit_project(pid))
            act_add_file.triggered.connect(lambda: self._on_add_file(pid))
            act_add_folder.triggered.connect(lambda: self._on_add_folder(pid))
            act_export_io.triggered.connect(lambda: self.export_project_io_requested.emit(pid))
            act_del.triggered.connect(lambda: self._on_delete_project(pid))

        elif item.data(0, self._ROLE_KIND) == "file":
            fid  = item.data(0, self._ROLE_ID)
            path = item.data(0, self._ROLE_PATH)
            missing = item.data(0, self._ROLE_MISSING)
            fname = item.text(0)
            act_open = menu.addAction("📂   Open in new tab")
            if missing:
                act_relink = menu.addAction("🔗   Re-link file…")
                act_relink.triggered.connect(lambda: self._on_relink_file(fid, item))
            menu.addSeparator()
            act_meta = menu.addAction("📋   Edit file metadata…")
            act_meta.triggered.connect(lambda: self._on_edit_file_metadata(fid, fname))
            menu.addSeparator()
            act_remove = menu.addAction("🗑   Remove from project")
            act_open.triggered.connect(lambda: self.open_file_requested.emit(path))
            act_remove.triggered.connect(lambda: self._on_remove_file(fid))

        menu.exec(global_pos)

    def _on_item_double_clicked(self, item: QTreeWidgetItem, _col: int):
        if item.data(0, self._ROLE_KIND) == "file":
            if not item.data(0, self._ROLE_MISSING):
                self.open_file_requested.emit(item.data(0, self._ROLE_PATH))
            else:
                QMessageBox.warning(
                    self, "File not found",
                    f"Cannot open — file not found:\n"
                    f"{item.data(0, self._ROLE_PATH)}\n\n"
                    f"Right-click → Re-link file… to update the path.")

    # ── Project actions ───────────────────────────────────────────────────
    def _on_new_project(self):
        dlg = ProjectMetadataDialog(parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            v = dlg.values
            db_create_project(v["name"], v["number"], v["description"])
            self.refresh()

    def _on_import_project_folder(self):
        """
        Let the user pick a folder on disk.  All PDFs found recursively are
        previewed, then a new project is created using the folder name as the
        default name.  The metadata dialog opens pre-filled so the user can
        review / adjust before saving.
        """
        folder = QFileDialog.getExistingDirectory(
            self, "Import Project Folder — select a folder")
        if not folder:
            return

        # Scan recursively for PDFs
        found = []
        for root_dir, _dirs, files in os.walk(folder):
            for fname in sorted(files):
                if fname.lower().endswith(".pdf"):
                    found.append(os.path.normpath(
                        os.path.join(root_dir, fname)))

        if not found:
            QMessageBox.information(
                self, "No PDFs found",
                f"No PDF files were found in:\n{folder}\n\n"
                f"Nothing was imported.")
            return

        # ── Preview dialog ────────────────────────────────────────────────
        preview_dlg = QDialog(self)
        preview_dlg.setStyleSheet(
            "* { background-color: #1E1E1E; color: #F0F0F0; }"
            "QDialog { background: #1E1E1E; }"
            "QLabel { color: #F0F0F0; background: transparent; }"
            "QListWidget { background: #252525; color: #F0F0F0;"
            " border: 1px solid #3A3A3A; font-size: 8pt; }"
            "QPushButton { background: #3A3A3A; color: #F0F0F0;"
            " border: 1px solid #555555; border-radius: 4px; padding: 4px 10px; }"
            "QPushButton:hover { background: #4A4A4A; }"
        )
        preview_dlg.setWindowTitle("Import Project Folder — Preview")
        preview_dlg.setMinimumWidth(560)
        preview_dlg.setMinimumHeight(380)

        lbl = QLabel(
            f"Found <b>{len(found)}</b> PDF file(s) in:<br>"
            f"<code>{folder}</code><br><br>"
            f"Click <b>Continue</b> to set the project metadata and import.")
        lbl.setWordWrap(True)
        lbl.setStyleSheet("color:#F0F0F0; margin-bottom:6px;")

        list_widget = QListWidget()
        for path in found:
            # Show path relative to the chosen folder for readability
            try:
                display = os.path.relpath(path, folder)
            except ValueError:
                display = path
            list_widget.addItem(display)
        list_widget.setToolTip("These PDFs will be added to the new project")

        bb = QDialogButtonBox()
        continue_btn = bb.addButton(
            "Continue…", QDialogButtonBox.ButtonRole.AcceptRole)
        cancel_btn   = bb.addButton(QDialogButtonBox.StandardButton.Cancel)
        continue_btn.setStyleSheet(
            "QPushButton { background:#1F4E79; color:#FFFFFF; font-weight:bold;"
            " border:none; border-radius:4px; padding:4px 14px; }"
            "QPushButton:hover { background:#2E75B6; }")
        bb.accepted.connect(preview_dlg.accept)
        bb.rejected.connect(preview_dlg.reject)

        lay = QVBoxLayout()
        lay.setContentsMargins(14, 14, 14, 10)
        lay.setSpacing(8)
        lay.addWidget(lbl)
        lay.addWidget(list_widget, stretch=1)
        lay.addWidget(bb)
        preview_dlg.setLayout(lay)

        if preview_dlg.exec() != QDialog.DialogCode.Accepted:
            return

        # ── Metadata dialog — pre-filled with folder name ─────────────────
        folder_name = os.path.basename(os.path.normpath(folder))
        pre = {
            "name":        folder_name,
            "number":      "",
            "description": f"Imported from: {folder}",
        }
        meta_dlg = ProjectMetadataDialog(project=pre, parent=self)
        # Override the window title to make the context clear
        meta_dlg.setWindowTitle("Import Project Folder — Set Metadata")

        if meta_dlg.exec() != QDialog.DialogCode.Accepted:
            return

        v = meta_dlg.values

        # ── Create project and add all PDFs ───────────────────────────────
        project_id = db_create_project(v["name"], v["number"], v["description"])
        for path in found:
            db_add_project_file(project_id, path)

        self.refresh()

        QMessageBox.information(
            self, "Import complete",
            f"Project <b>{v['name']}</b> created with "
            f"{len(found)} PDF file(s).\n\n"
            f"Source folder: {folder}")

    def _on_edit_project(self, project_id: int):
        projects = db_load_projects()
        p = next((x for x in projects if x["id"] == project_id), None)
        if not p:
            return
        dlg = ProjectMetadataDialog(project=p, parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            v = dlg.values
            db_update_project(project_id, v["name"], v["number"], v["description"])
            self.refresh()

    def _on_delete_project(self, project_id: int):
        projects = db_load_projects()
        p = next((x for x in projects if x["id"] == project_id), None)
        name = p["name"] if p else "this project"
        ans = QMessageBox.question(
            self, "Delete Project",
            f"Delete project <b>{name}</b> and all its file links?\n"
            f"(PDF files on disk are not deleted.)",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if ans == QMessageBox.StandardButton.Yes:
            db_delete_project(project_id)
            self.refresh()

    # ── File actions ──────────────────────────────────────────────────────
    def _on_add_file(self, project_id: int):
        paths, _ = QFileDialog.getOpenFileNames(
            self, "Add PDF files", "", "PDF Files (*.pdf)")
        for path in paths:
            db_add_project_file(project_id, path)
        if paths:
            self.refresh()

    def _on_add_folder(self, project_id: int):
        folder = QFileDialog.getExistingDirectory(
            self, "Add folder — all PDFs will be added recursively")
        if not folder:
            return
        found = []
        for root, _dirs, files in os.walk(folder):
            for fname in sorted(files):
                if fname.lower().endswith(".pdf"):
                    found.append(os.path.join(root, fname))
        if not found:
            QMessageBox.information(self, "No PDFs found",
                                    "No PDF files found in the selected folder.")
            return
        # Preview dialog with scrollable file list
        preview_dlg = QDialog(self)
        preview_dlg.setWindowTitle("Add PDFs from folder")
        preview_dlg.setMinimumWidth(500)
        preview_dlg.setMinimumHeight(340)

        lbl = QLabel(f"Found <b>{len(found)}</b> PDF file(s) in:<br>"
                     f"<code>{folder}</code><br><br>Add them all to the project?")
        lbl.setWordWrap(True)

        list_widget = QListWidget()
        for path in found:
            list_widget.addItem(path)
        list_widget.setStyleSheet("font-size: 8pt;")

        bb = QDialogButtonBox(QDialogButtonBox.StandardButton.Yes |
                              QDialogButtonBox.StandardButton.No)
        bb.accepted.connect(preview_dlg.accept)
        bb.rejected.connect(preview_dlg.reject)

        lay = QVBoxLayout()
        lay.setContentsMargins(14, 14, 14, 10)
        lay.addWidget(lbl)
        lay.addWidget(list_widget, stretch=1)
        lay.addWidget(bb)
        preview_dlg.setLayout(lay)

        if preview_dlg.exec() == QDialog.DialogCode.Accepted:
            for path in found:
                db_add_project_file(project_id, path)
            self.refresh()

    def _on_remove_file(self, file_id: int):
        db_remove_project_file(file_id)
        self.refresh()

    def _on_edit_file_metadata(self, file_id: int, file_name: str):
        dlg = FileMetadataDialog(file_id, file_name, parent=self)
        dlg.exec()

    def _on_relink_file(self, file_id: int, item: QTreeWidgetItem):
        path, _ = QFileDialog.getOpenFileName(
            self, "Re-link file — choose new location", "", "PDF Files (*.pdf)")
        if not path:
            return
        with _db_connect() as con:
            _ensure_project_tables(con)
            con.execute(
                "UPDATE project_files SET file_path=?, file_name=? WHERE id=?",
                (path, os.path.basename(path), file_id))
            con.commit()
        self.refresh()


# ---------------------------------------------------------------------------
# TabState — all per-tab data bundled together
# ---------------------------------------------------------------------------
@dataclass
class TabState:
    pdf_path:     str
    pdf_document: object          # QPdfDocument
    pdf_view:     object          # DraggablePdfView
    io_list:      list = field(default_factory=list)
    undo_stack:   list = field(default_factory=list)
    redo_stack:   list = field(default_factory=list)
    container:    object = None   # QWidget holding pdf_view + zoom panel
    zoom_panel:   object = None   # floating zoom widget
    zoom_pct_lbl: object = None   # QLabel inside zoom panel

# ---------------------------------------------------------------------------
# MarkerInfoDialog — view marker information (read-only)
# ---------------------------------------------------------------------------
class MarkerInfoDialog(QDialog):
    """
    Read-only dialog showing all information about a marker.
    Opened when user double-clicks a marker.
    """
    
    def __init__(self, marker: dict, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Marker Information")
        self.setMinimumWidth(500)
        self._marker = marker
        self._build_ui()
    
    def _build_ui(self):
        lay = QVBoxLayout()
        lay.setContentsMargins(16, 14, 16, 10)
        lay.setSpacing(12)
        
        m = self._marker
        
        # ── Marker type header ────────────────────────────────────────────
        if m.get("is_composition"):
            composition = db_load_signal_composition(m.get("composition_id"))
            if composition:
                composition_text = _get_signal_composition(composition)
                title = QLabel(
                    f"<b style='font-size:12pt;color:#7EC8F0;'>"
                    f"{composition['title']}</b><br>"
                    f"<span style='color:#AAAAAA;font-size:9pt;'>"
                    f"{composition_text}</span>")
                if composition.get("description"):
                    title.setText(
                        f"<b style='font-size:12pt;color:#7EC8F0;'>"
                        f"{composition['title']}</b><br>"
                        f"<span style='color:#AAAAAA;font-size:9pt;'>"
                        f"{composition_text}</span><br>"
                        f"<span style='color:#AAAAAA;font-size:8pt;'>"
                        f"{composition['description']}</span>")
                lay.addWidget(title)
        else:
            # Simple custom marker
            title_text = m.get("type", "Unknown")
            title = QLabel(f"<b style='font-size:12pt;'>{title_text}</b>")
            lay.addWidget(title)
        
        lay.addWidget(self._divider())
        
        # ── Marker-specific information ───────────────────────────────────
        if m.get("is_composition"):
            composition = db_load_signal_composition(m.get("composition_id"))
            if composition:
                self._add_composition_info(lay, m, composition)
        else:
            self._add_custom_marker_info(lay, m)
        
        # ── Close button ──────────────────────────────────────────────────
        bb = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
        bb.accepted.connect(self.accept)
        lay.addWidget(bb)
        
        self.setLayout(lay)
    
    def _divider(self) -> QWidget:
        """Return a horizontal divider line."""
        div = QWidget()
        div.setStyleSheet("background: #3A3A3A;")
        div.setFixedHeight(1)
        return div
    
    def _add_composition_info(self, lay: QVBoxLayout, marker: dict, composition: dict):
        """Add composition-specific information to the dialog."""
        
        # ── Composition metadata ─────────────────────────────────────────
        lay.addWidget(QLabel("<b>Configuration:</b>"))

        tag_parts = marker.get("tag_parts", {})
        count = tag_parts.get("count", 1)

        config_table = QTableWidget(7, 2)
        config_table.setColumnWidth(0, 160)
        config_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch)
        config_table.horizontalHeader().setVisible(False)
        config_table.verticalHeader().setVisible(False)
        config_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)

        config_table.setItem(0, 0, QTableWidgetItem("Count (multiplier):"))
        config_table.setItem(0, 1, QTableWidgetItem(str(count)))

        config_table.setItem(1, 0, QTableWidgetItem("Control Module Name:"))
        config_table.setItem(1, 1, QTableWidgetItem(
            composition.get("control_module", "NA") or "NA"))

        config_table.setItem(2, 0, QTableWidgetItem("Control Module Type:"))
        config_table.setItem(2, 1, QTableWidgetItem(
            composition.get("cm_type", "NA") or "NA"))

        config_table.setItem(3, 0, QTableWidgetItem("Control Module Desc.:"))
        config_table.setItem(3, 1, QTableWidgetItem(
            composition.get("cm_description", "NA") or "NA"))

        config_table.setItem(4, 0, QTableWidgetItem("Field Device Name:"))
        config_table.setItem(4, 1, QTableWidgetItem(
            composition.get("field_device", "NA") or "NA"))

        config_table.setItem(5, 0, QTableWidgetItem("Field Device Type:"))
        config_table.setItem(5, 1, QTableWidgetItem(
            composition.get("fd_type", "NA") or "NA"))

        config_table.setItem(6, 0, QTableWidgetItem("F Desc.:"))
        config_table.setItem(6, 1, QTableWidgetItem(
            composition.get("fd_description", "NA") or "NA"))

        config_table.setMaximumHeight(200)
        lay.addWidget(config_table)
        
        lay.addWidget(self._divider())
        
        # ── Signals table ────────────────────────────────────────────────
        lay.addWidget(QLabel("<b>Signals in Composition:</b>"))
        
        signals_table = QTableWidget(len(composition["signals"]), 5)
        signals_table.setHorizontalHeaderLabels(
            ["Signal Name", "Signal Type", "Description", "Prefix", "Suffix"])
        signals_table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.ResizeToContents)
        signals_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.ResizeToContents)
        signals_table.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.ResizeMode.Stretch)
        signals_table.horizontalHeader().setSectionResizeMode(
            3, QHeaderView.ResizeMode.ResizeToContents)
        signals_table.horizontalHeader().setSectionResizeMode(
            4, QHeaderView.ResizeMode.ResizeToContents)
        signals_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        signals_table.verticalHeader().setVisible(False)
        
        for row, sig in enumerate(composition["signals"]):
            signals_table.setItem(row, 0, QTableWidgetItem(sig.get("signal_name", "")))
            signals_table.setItem(row, 1, QTableWidgetItem(sig.get("signal_type", "")))
            signals_table.setItem(row, 2, QTableWidgetItem(
                sig.get("signal_description", "")))
            signals_table.setItem(row, 3, QTableWidgetItem(
                sig.get("prefix") or "NA"))
            signals_table.setItem(row, 4, QTableWidgetItem(
                sig.get("suffix") or "NA"))
        
        signals_table.setMaximumHeight(min(200, len(composition["signals"]) * 25 + 30))
        lay.addWidget(signals_table, stretch=1)
    
    def _add_custom_marker_info(self, lay: QVBoxLayout, marker: dict):
        """Add custom marker-specific information to the dialog."""
        
        lay.addWidget(QLabel("<b>Marker Details:</b>"))
        
        details_table = QTableWidget(4, 2)
        details_table.setColumnWidth(0, 120)
        details_table.setColumnWidth(1, 200)
        details_table.horizontalHeader().setVisible(False)
        details_table.verticalHeader().setVisible(False)
        details_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        
        count = marker.get("count", 1)
        details_table.setItem(0, 0, QTableWidgetItem("Count:"))
        details_table.setItem(0, 1, QTableWidgetItem(str(count)))
        
        comment = marker.get("comment", "")
        details_table.setItem(1, 0, QTableWidgetItem("Comment:"))
        details_table.setItem(1, 1, QTableWidgetItem(comment if comment else "(none)"))
        
        description = marker.get("description", "")
        details_table.setItem(2, 0, QTableWidgetItem("Description:"))
        details_table.setItem(2, 1, QTableWidgetItem(description if description else "(none)"))
        
        signal_type = marker.get("signal_type", "")
        details_table.setItem(3, 0, QTableWidgetItem("Type:"))
        details_table.setItem(3, 1, QTableWidgetItem(signal_type if signal_type else "(none)"))
        
        details_table.setMaximumHeight(120)
        lay.addWidget(details_table, stretch=1)
        
# ---------------------------------------------------------------------------
# PDFViewer
# ---------------------------------------------------------------------------
class PDFViewer(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Technical Drawing Viewer")
        self.resize(1280, 900)

        # ── Resolve saved theme (applied after _build_ui so toolbar exists) ─
        self._current_theme = db_load_theme()

        # Shared across all tabs — signal type list
        self._signal_types: list = db_load_signal_types()

        self._build_ui()
        self._connect_signals()
        self._apply_theme(self._current_theme, _persist=False)
        self.act_configure_compositions.triggered.connect(self.open_signal_compositions_config)
        self._set_pdf_actions_enabled(False)

    # ── UI construction ───────────────────────────────────────────────────
    def _build_ui(self):
        mb = self.menuBar()

        # File menu
        file_menu = mb.addMenu("&File")
        self.act_open      = QAction("&Open PDF…",                self, shortcut=QKeySequence.StandardKey.Open)
        self.act_close_pdf = QAction("&Close Tab",                self, shortcut="Ctrl+W")
        self.act_save_fdf  = QAction("&Save FDF…",                self, shortcut="Ctrl+S")
        self.act_save_pdf  = QAction("Save &PDF with comments…",  self, shortcut="Ctrl+Shift+S")
        self.act_export_xlsx      = QAction("Export to &Excel…",           self, shortcut="Ctrl+E")
        self.act_export_project_io = QAction("Export &Project IO List…",   self)
        self.act_link_session      = QAction("🔗  &Link to Previous Session…", self)
        self.act_exit              = QAction("E&xit",                       self, shortcut=QKeySequence.StandardKey.Quit)
        file_menu.addAction(self.act_open)
        file_menu.addAction(self.act_close_pdf)
        file_menu.addSeparator()
        file_menu.addAction(self.act_save_fdf)
        file_menu.addAction(self.act_save_pdf)
        file_menu.addAction(self.act_export_project_io)
        file_menu.addSeparator()
        file_menu.addAction(self.act_link_session)
        file_menu.addSeparator()
        file_menu.addAction(self.act_exit)

        # View menu
        view_menu = mb.addMenu("&View")
        self.act_zoom_in   = QAction("Zoom &In",       self, shortcut="Ctrl+=")
        self.act_zoom_out  = QAction("Zoom &Out",      self, shortcut="Ctrl+-")
        self.act_zoom_fit  = QAction("&Reset Zoom",    self, shortcut="Ctrl+0")
        self.act_prev_page = QAction("&Previous Page", self, shortcut=QKeySequence.StandardKey.MoveToPreviousPage)
        self.act_next_page = QAction("&Next Page",     self, shortcut=QKeySequence.StandardKey.MoveToNextPage)
        self.act_show_projects = QAction("📁  &Project Panel", self,
                                          shortcut="Ctrl+B", checkable=True)
        self.act_show_projects.setChecked(False)
        view_menu.addAction(self.act_show_projects)
        view_menu.addSeparator()
        view_menu.addAction(self.act_zoom_in)
        view_menu.addAction(self.act_zoom_out)
        view_menu.addAction(self.act_zoom_fit)
        view_menu.addSeparator()
        view_menu.addAction(self.act_prev_page)
        view_menu.addAction(self.act_next_page)

        # ── Theme sub-menu ────────────────────────────────────────────────
        theme_menu = view_menu.addMenu("🎨  &Theme")
        self._theme_group = QActionGroup(self)
        self._theme_group.setExclusive(True)
        self.act_theme_dark   = QAction("🌑  &Dark",           self, checkable=True)
        self.act_theme_light  = QAction("☀️   &Light",          self, checkable=True)
        self.act_theme_system = QAction("🖥️   &System Default", self, checkable=True)
        for act in (self.act_theme_dark, self.act_theme_light, self.act_theme_system):
            self._theme_group.addAction(act)
            theme_menu.addAction(act)
        # Check the currently active theme
        _theme_acts = {"dark": self.act_theme_dark,
                       "light": self.act_theme_light,
                       "system": self.act_theme_system}
        _theme_acts.get(self._current_theme, self.act_theme_dark).setChecked(True)
        self.act_theme_dark.triggered.connect(lambda: self._apply_theme("dark"))
        self.act_theme_light.triggered.connect(lambda: self._apply_theme("light"))
        self.act_theme_system.triggered.connect(lambda: self._apply_theme("system"))

        # Edit menu
        edit_menu = mb.addMenu("&Edit")
        self.act_undo          = QAction("&Undo", self, shortcut="Ctrl+Z")
        self.act_redo          = QAction("&Redo", self, shortcut="Ctrl+Y")
        # Use platform undo/redo icons with emoji fallback
        _undo_icon = QApplication.style().standardIcon(
            QStyle.StandardPixmap.SP_ArrowBack)
        _redo_icon = QApplication.style().standardIcon(
            QStyle.StandardPixmap.SP_ArrowForward)
        self.act_undo.setIcon(_undo_icon)
        self.act_redo.setIcon(_redo_icon)
        self.act_configure     = QAction("&Configure Signal Types…",  self, shortcut="Ctrl+,")
        self.act_configure_compositions = QAction(
            "🔧 Configure Signal &Typicals…", self, shortcut="Ctrl+Shift+,")
        self.act_config_export = QAction("Configure &Export Columns…", self)
        self.act_undo.setEnabled(False)
        self.act_redo.setEnabled(False)
        edit_menu.addAction(self.act_undo)
        edit_menu.addAction(self.act_redo)
        edit_menu.addSeparator()
        edit_menu.addAction(self.act_configure_compositions)

        # Preferences menu
        pref_menu = mb.addMenu("&Preferences")
        self.act_mode_marker = QAction("📍  &Marker Mode",       self, shortcut="Ctrl+1", checkable=True)
        self.act_mode_text   = QAction("💬  &Text Comment Mode", self, shortcut="Ctrl+2", checkable=True)
        self.act_mode_marker.setChecked(True)
        pref_menu.addSection("Input Mode")
        pref_menu.addAction(self.act_mode_marker)
        pref_menu.addAction(self.act_mode_text)

        # ── Help menu ─────────────────────────────────────────────────────
        help_menu = mb.addMenu("&Help")
        self.act_about = QAction("ℹ️  &About",        self)
        self.act_help  = QAction("❓  &Help", self, shortcut="F1")
        help_menu.addAction(self.act_about)
        help_menu.addAction(self.act_help)
        self.act_about.triggered.connect(self._show_about)
        self.act_help.triggered.connect(self._show_help)

        # ── Quick-access toolbar ─────────────────────────────────────────
        tb = self.addToolBar("Quick Access")
        tb.setMovable(False)
        tb.setFloatable(False)
        tb.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextOnly)
        self._toolbar = tb  # set early so _apply_theme can find it

        def _tb_action(emoji: str, label: str, action: QAction) -> QAction:
            """Set the action text to 'emoji\\nlabel' for the toolbar button."""
            action.setText(f"{emoji}\n{label}")
            return action

        # Group 1 — File
        _tb_action("📂", "Open",       self.act_open)
        tb.addAction(self.act_open)
        _tb_action("✖",  "Close Tab",  self.act_close_pdf)
        tb.addAction(self.act_close_pdf)
        tb.addSeparator()

        # Group 2 — Save / Export
        _tb_action("💾", "Save FDF",   self.act_save_fdf)
        tb.addAction(self.act_save_fdf)
        _tb_action("📑", "Save PDF",   self.act_save_pdf)
        tb.addAction(self.act_save_pdf)
        tb.addSeparator()

        # Group 3 — Edit
        _tb_action("↩", "Undo", self.act_undo)
        tb.addAction(self.act_undo)
        _tb_action("↪", "Redo", self.act_redo)
        tb.addAction(self.act_redo)
        tb.addSeparator()

        # Group 4 — View / Navigation
        _tb_action("📁", "Projects",   self.act_show_projects)
        tb.addAction(self.act_show_projects)
        tb.addSeparator()
        _tb_action("⬆",  "Prev Page",  self.act_prev_page)
        tb.addAction(self.act_prev_page)
        _tb_action("⬇",  "Next Page",  self.act_next_page)
        tb.addAction(self.act_next_page)
        tb.addSeparator()
        _tb_action("🔍+", "Zoom In",   self.act_zoom_in)
        tb.addAction(self.act_zoom_in)
        _tb_action("⊙",  "Reset Zoom", self.act_zoom_fit)
        tb.addAction(self.act_zoom_fit)
        _tb_action("🔍−", "Zoom Out",  self.act_zoom_out)
        tb.addAction(self.act_zoom_out)
        tb.addSeparator()

        # Group 5 — Input Mode (checkable)
        _tb_action("📍", "Marker",     self.act_mode_marker)
        tb.addAction(self.act_mode_marker)
        _tb_action("💬", "Text",       self.act_mode_text)
        tb.addAction(self.act_mode_text)
        tb.addSeparator()

        # Group 6 — Export / Config shortcuts
        _tb_action("📊", "Export Excel",    self.act_export_xlsx)
        tb.addAction(self.act_export_xlsx)
        _tb_action("🔧", "Signal Typicals", self.act_configure_compositions)
        tb.addAction(self.act_configure_compositions)
        tb.addSeparator()

        # ── Tab widget ───────────────────────────────────────────────────
        self._tabs = QTabWidget()
        self._tabs.setTabsClosable(True)
        self._tabs.setMovable(True)
        self._tabs.setDocumentMode(True)

        # Home screen at index 0 (a plain widget, not closable)
        self._home_screen = self._build_home_screen()
        self._tabs.addTab(self._home_screen, "Home")
        self._tabs.tabBar().setTabButton(0, self._tabs.tabBar().ButtonPosition.RightSide, None)

        self.setCentralWidget(self._tabs)

        # ── Project panel dock ────────────────────────────────────────────
        self._project_panel = ProjectPanel()
        self._project_panel.set_theme(self._current_theme)
        self._project_dock  = QDockWidget("Projects", self)
        self._project_dock.setWidget(self._project_panel)
        self._project_dock.setAllowedAreas(
            Qt.DockWidgetArea.LeftDockWidgetArea |
            Qt.DockWidgetArea.RightDockWidgetArea)
        self._project_dock.setMinimumWidth(220)
        self._project_dock.setFeatures(
            QDockWidget.DockWidgetFeature.DockWidgetMovable |
            QDockWidget.DockWidgetFeature.DockWidgetClosable)
        self.addDockWidget(Qt.DockWidgetArea.LeftDockWidgetArea,
                           self._project_dock)
        self._project_dock.hide()

        # Status bar permanent widgets
        self.page_label = QLabel("Page: - / -")
        self.page_label.setStyleSheet("padding: 0 8px; color:#CCCCCC;")
        self._mode_label = QLabel("📍 Marker Mode")
        self._mode_label.setStyleSheet(
            "padding:2px 10px; border-radius:4px;"
            "background:#1A3A5C; color:#7EC8F0; font-weight:bold; font-size: 8pt;")
        self._version_label = QLabel(getattr(self, "_app_version", "v1.0.0"))
        self._version_label.setStyleSheet("color:#777777; font-size: 8pt; padding:0 10px;")
        self.statusBar().addPermanentWidget(self._mode_label)
        self.statusBar().addPermanentWidget(self.page_label)
        self.statusBar().addPermanentWidget(self._version_label)
        self.statusBar().showMessage("Open a PDF to begin.")

    # ── Theme ──────────────────────────────────────────────────────────────
    def _apply_theme(self, theme: str, _persist: bool = True) -> None:
        """Switch the application-wide theme to 'dark', 'light', or 'system'.

        For 'system', the OS palette is inspected to determine whether the
        desktop is currently in dark or light mode, and the matching theme
        is applied automatically.
        """
        self._current_theme = theme
        app = QApplication.instance()

        # Resolve 'system' → actual OS preference
        if theme == "system":
            bg_lightness = app.palette().color(QPalette.ColorRole.Window).lightness()
            effective = "dark" if bg_lightness < 128 else "light"
        else:
            effective = theme

        if effective == "dark":
            app.setStyleSheet(_DARK_THEME_SS)
            if hasattr(self, "_toolbar"):
                self._toolbar.setStyleSheet(_DARK_TOOLBAR_SS)
            if hasattr(self, "_project_dock"):
                self._project_dock.setStyleSheet(_DARK_DOCK_SS)
        else:  # light
            app.setStyleSheet(_LIGHT_THEME_SS)
            if hasattr(self, "_toolbar"):
                self._toolbar.setStyleSheet(_LIGHT_TOOLBAR_SS)
            if hasattr(self, "_project_dock"):
                self._project_dock.setStyleSheet(_LIGHT_DOCK_SS)

        if _persist:
            db_save_theme(theme)
        if hasattr(self, "_project_panel"):
            self._project_panel.set_theme(effective)

    # ── Help dialogs ───────────────────────────────────────────────────────
    def _show_about(self) -> None:
        """Display the About dialog."""
        version = getattr(self, "_app_version", "v1.0.0")
        msg = QMessageBox(self)
        msg.setWindowTitle("About")
        msg.setIcon(QMessageBox.Icon.Information)
        msg.setText(
            "<b>Tool Information</b>"
        )
        msg.setInformativeText(
            f"<table style='font-size:10pt; line-height:1.6;'>"
            f"<tr><td><b>Program&nbsp;Name</b></td><td>&nbsp;:&nbsp;</td><td>Technical Drawing Viewer </td></tr>"
            f"<tr><td><b>Developed&nbsp;By</b></td><td>&nbsp;:&nbsp;</td><td>Sriharan Thirumalai</td></tr>"
            f"<tr><td><b>Mentor</b></td><td>&nbsp;:&nbsp;</td><td>Carlo Lebrun</td></tr>"
            f"<tr><td><b>Scripted&nbsp;Using</b></td><td>&nbsp;:&nbsp;</td><td>Python</td></tr>"
            f"<tr><td><b>Version</b></td><td>&nbsp;:&nbsp;</td><td>{version}</td></tr>"
            f"<tr><td><b>Date</b></td><td>&nbsp;:&nbsp;</td><td>14/04/2026</td></tr>"
            f"</table>"
        )
        msg.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg.exec()

    def _show_help(self) -> None:
        """Display the Help / Contact dialog."""
        msg = QMessageBox(self)
        msg.setWindowTitle("Help")
        msg.setIcon(QMessageBox.Icon.Question)
        msg.setText("<b>Program Assistance</b>")
        msg.setInformativeText(
            "<p>For any program-related assistance, please contact <b>Sriharan Thirumalai</b>.</p>"
            "<table style='font-size:10pt; line-height:1.8;'>"
            "<tr><td><b>Email</b></td><td>&nbsp;:&nbsp;</td><td>sriharan.thirumalai@italiautomazione.com</td></tr>"
            "<tr><td><b>Phone</b></td><td>&nbsp;:&nbsp;</td><td>+39 3480380741</td></tr>"
            "</table>"
        )
        msg.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg.exec()

    def _connect_signals(self):
        self.act_open.triggered.connect(self.open_pdf)
        self.act_close_pdf.triggered.connect(self.close_current_tab)
        self.act_show_projects.triggered.connect(self._toggle_project_panel)
        self._project_dock.visibilityChanged.connect(
            lambda v: self.act_show_projects.setChecked(v))
        self._project_panel.open_file_requested.connect(
            lambda p: self.open_pdf(p, _from_project=True))
        self._project_panel.export_project_io_requested.connect(
            self._export_project_io_by_id)
        self.act_save_fdf.triggered.connect(self.save_fdf)
        self.act_save_pdf.triggered.connect(self.save_pdf)
        self.act_exit.triggered.connect(self.close)
        self.act_zoom_in.triggered.connect(lambda: self._do_zoom_in())
        self.act_zoom_out.triggered.connect(lambda: self._do_zoom_out())
        self.act_zoom_fit.triggered.connect(lambda: self._apply_zoom_reset())
        self.act_prev_page.triggered.connect(
            lambda: self._go_to_page(self._current_tab().pdf_view.pageNavigator().currentPage() - 1))
        self.act_next_page.triggered.connect(
            lambda: self._go_to_page(self._current_tab().pdf_view.pageNavigator().currentPage() + 1))
        self.act_configure.triggered.connect(self.open_config)
        self.act_config_export.triggered.connect(self.open_export_config)
        self.act_export_xlsx.triggered.connect(self.export_xlsx)
        self.act_export_project_io.triggered.connect(self.export_project_io_list)
        self.act_link_session.triggered.connect(self.open_manual_link)
        self.act_mode_marker.triggered.connect(lambda: self._set_input_mode("marker"))
        self.act_mode_text.triggered.connect(lambda: self._set_input_mode("text"))
        self.act_undo.triggered.connect(self.undo)
        self.act_redo.triggered.connect(self.redo)
        self._tabs.tabCloseRequested.connect(self._on_tab_close_requested)
        self._tabs.currentChanged.connect(self._on_tab_changed)

    # ── Tab helpers ───────────────────────────────────────────────────────
    def _current_tab(self) -> TabState | None:
        """Return the TabState for the active tab, or None if home is active."""
        idx = self._tabs.currentIndex()
        return self._tabs.currentWidget().property("tab_state") if idx > 0 else None

    def _tab_state_at(self, index: int) -> TabState | None:
        w = self._tabs.widget(index)
        return w.property("tab_state") if w else None

    def _find_tab_for_path(self, path: str) -> int:
        """Return tab index for a given pdf_path, or -1 if not open."""
        for i in range(1, self._tabs.count()):
            ts = self._tab_state_at(i)
            if ts and ts.pdf_path == path:
                return i
        return -1

    def _set_pdf_actions_enabled(self, enabled: bool):
        for act in (self.act_close_pdf, self.act_save_fdf, self.act_save_pdf,
                    self.act_link_session, self.act_export_xlsx,
                    self.act_zoom_in, self.act_zoom_out, self.act_zoom_fit,
                    self.act_prev_page, self.act_next_page):
            act.setEnabled(enabled)

    def _on_tab_changed(self, index: int):
        ts = self._tab_state_at(index)
        if ts is None:
            # Home tab
            self._set_pdf_actions_enabled(False)
            self.act_undo.setEnabled(False)
            self.act_redo.setEnabled(False)
            self.act_undo.setText("↩\nUndo")
            self.act_redo.setText("↪\nRedo")
            self.page_label.setText("Page: - / -")
            self.statusBar().showMessage("Open a PDF to begin.")
        else:
            self._set_pdf_actions_enabled(True)
            self._refresh_undo_actions(ts)
            self._update_page_label(ts)
            n = len(ts.io_list)
            self.statusBar().showMessage(
                f"{os.path.basename(ts.pdf_path)}  —  {n} marker{'s' if n!=1 else ''}")

    def _on_tab_close_requested(self, index: int):
        if index == 0:
            return   # home tab is not closable
        ts = self._tab_state_at(index)
        if ts:
            db_save_markers(ts.pdf_path, ts.io_list)
            # Remove from Unassigned section if it was there
            self._project_panel.remove_unassigned_file(ts.pdf_path)
        self._tabs.removeTab(index)
        if self._tabs.count() == 1:
            self._tabs.setCurrentIndex(0)

    # ── Per-tab widget factory ────────────────────────────────────────────
    def _make_tab_widget(self, ts: TabState) -> QWidget:
        """Build the container widget for a new tab and wire its signals."""
        container = QWidget()
        container.setLayout(QVBoxLayout())
        container.layout().setContentsMargins(0, 0, 0, 0)
        container.layout().addWidget(ts.pdf_view)
        ts.pdf_view._pdf_path = ts.pdf_path
        container.setProperty("tab_state", ts)

        # Floating zoom panel parented to the container
        zoom_panel = QWidget(container)
        zoom_panel.setObjectName("zoomPanel")
        zoom_panel.setStyleSheet("""
            QWidget#zoomPanel {
                background: rgba(40,40,40,210); border-radius:8px;
            }
            QPushButton {
                background:transparent; color:white; font-size: 15pt;
                font-weight:bold; border:none; padding:2px 10px;
            }
            QPushButton:hover   { background:rgba(255,255,255,30); border-radius:6px; }
            QPushButton:pressed { background:rgba(255,255,255,60); border-radius:6px; }
            QLabel { color:#CCCCCC; font-size: 8pt;
                     qproperty-alignment:AlignCenter; }
        """)
        btn_in  = QPushButton("＋", zoom_panel)
        btn_out = QPushButton("－", zoom_panel)
        pct_lbl = QLabel("100%", zoom_panel)
        zlay = QVBoxLayout(zoom_panel)
        zlay.setContentsMargins(6,6,6,6); zlay.setSpacing(2)
        zlay.addWidget(btn_in); zlay.addWidget(pct_lbl); zlay.addWidget(btn_out)
        zoom_panel.setFixedWidth(52)
        zoom_panel.adjustSize()
        zoom_panel.raise_()
        zoom_panel.show()

        ts.container    = container
        ts.zoom_panel   = zoom_panel
        ts.zoom_pct_lbl = pct_lbl

        btn_in.clicked.connect(lambda: self._do_zoom_in(ts))
        btn_out.clicked.connect(lambda: self._do_zoom_out(ts))

        # Reposition zoom panel on resize
        orig_resize = container.resizeEvent
        def _on_resize(ev, _ts=ts):
            self._reposition_zoom_panel(_ts)
            orig_resize(ev) if orig_resize else None
        container.resizeEvent = _on_resize

        # Wire pdf_view signals
        ts.pdf_view.io_registered.connect(
            lambda t, cp, pg, x, y, c, cnt, desc, st, stc, _ts=ts:
                self._register_io(_ts, t, cp, pg, x, y, c, cnt, desc, st, stc))
        ts.pdf_view.text_registered.connect(
            lambda txt, pg, x, y, _ts=ts: self._register_text(_ts, txt, pg, x, y))
        ts.pdf_view.composition_registered.connect(
            lambda cid, pg, x, y, cfg, tparts, _ts=ts:
                self._register_composition(_ts, cid, pg, x, y, cfg, tparts))
        ts.pdf_view.markers_changed.connect(
            lambda _ts=ts: self._on_overlay_changed(_ts))
        ts.pdf_view._overlay._push_undo_fn = lambda _ts=ts: self._push_undo(_ts)
        return container

    def _reposition_zoom_panel(self, ts: TabState):
        if ts.container is None or ts.zoom_panel is None:
            return
        c = ts.container
        zp = ts.zoom_panel
        zp.move(c.width() - zp.width() - 14,
                c.height() - zp.height() - 14)

    # ── Home screen ───────────────────────────────────────────────────────
    def _build_home_screen(self) -> QWidget:
        # ── Edit these lines ──────────────────────────────────────────────
        APP_VERSION = "v1.0.0"
        LOGO_PATH   = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo.png")
        LOGO_WIDTH  = 180

        self._app_version = APP_VERSION

        w = QWidget()
        w.setStyleSheet("background:#F5F7FA;")

        title = QLabel("Technical Drawing Viewer")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet("font-size: 21pt; font-weight:bold; color:#1F4E79; margin-bottom:6px;")

        subtitle = QLabel("Open a technical drawing to start placing markers and comments.")
        subtitle.setAlignment(Qt.AlignmentFlag.AlignCenter)
        subtitle.setStyleSheet("font-size: 10pt; color:#555; margin-bottom:28px;")

        open_btn = QPushButton("📂   Open PDF")
        open_btn.setFixedSize(200, 48)
        open_btn.setStyleSheet("""
            QPushButton { background:#1F4E79; color:white; border-radius:8px;
                          font-size: 10pt; font-weight:bold; }
            QPushButton:hover   { background:#2E75B6; }
            QPushButton:pressed { background:#16375A; }
        """)
        open_btn.clicked.connect(self.open_pdf)

        hint = QLabel("Ctrl+O  to open  ·  Ctrl+W  to close tab")
        hint.setAlignment(Qt.AlignmentFlag.AlignCenter)
        hint.setStyleSheet("font-size: 8pt; color:#999; margin-top:8px;")

        logo_label = QLabel()
        logo_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        if os.path.isfile(LOGO_PATH):
            from PySide6.QtGui import QPixmap
            pix = QPixmap(LOGO_PATH)
            logo_label.setPixmap(
                pix.scaledToWidth(LOGO_WIDTH, Qt.TransformationMode.SmoothTransformation))
        else:
            logo_label.setText("[ logo.png not found ]")
            logo_label.setStyleSheet("font-size: 8pt; color:#CCC;")

        ver_label = QLabel(APP_VERSION)
        ver_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        ver_label.setStyleSheet("font-size: 8pt; color:#AAA; margin-top:6px;")

        lay = QVBoxLayout()
        lay.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay.addStretch()
        lay.addWidget(title)
        lay.addWidget(subtitle)
        lay.addWidget(open_btn, alignment=Qt.AlignmentFlag.AlignCenter)
        lay.addWidget(hint)
        lay.addSpacing(24)
        lay.addWidget(logo_label)
        lay.addWidget(ver_label)
        lay.addStretch()
        w.setLayout(lay)
        return w

    # ── Open / close ──────────────────────────────────────────────────────
    def open_pdf(self, path: str = "", _from_project: bool = False):
        if not path:
            path, _ = QFileDialog.getOpenFileName(
                self, "Open PDF", "", "PDF Files (*.pdf)")
        if not path:
            return

        path = os.path.normpath(path)

        # If already open in a tab, just switch to it
        existing = self._find_tab_for_path(path)
        if existing >= 0:
            self._tabs.setCurrentIndex(existing)
            return

        # ── Project-panel integration ─────────────────────────────────────
        # Always ensure the project panel is visible when opening a PDF
        if not self._project_dock.isVisible():
            self._project_dock.show()
            self._project_panel.refresh()

        if not _from_project:
            # Check whether this PDF belongs to any project
            project_files = self._project_panel.get_all_project_file_paths()
            match = next(
                (f for f in project_files
                 if os.path.normpath(f["path"]) == path),
                None
            )
            if match:
                # Already in a project — highlight it in the panel
                self._project_panel.highlight_file(path)
            else:
                # Not in any project — show it in the Unassigned section
                # so the user can drag it into a project
                self._project_panel.show_unassigned_file(path)

        # ── Open the tab ──────────────────────────────────────────────────
        # Standalone (not-from-project) tabs go at index 1, above project tabs
        self._open_pdf_as_tab(path, insert_at_front=not _from_project)

    def _open_pdf_as_tab(self, path: str, insert_at_front: bool = False):
        """
        Internal: build the tab widget for a PDF and insert/append it.
        insert_at_front=True  → inserted at tab index 1 (just after Home).
        insert_at_front=False → appended at the end (normal project behaviour).
        """
        doc = QPdfDocument(self)
        doc.load(path)
        io_list: list[dict] = []
        view = DraggablePdfView(io_list, self._signal_types)
        view.setDocument(doc)
        view.setPageMode(QPdfView.PageMode.MultiPage)

        ts = TabState(pdf_path=path, pdf_document=doc,
                      pdf_view=view, io_list=io_list)
        container = self._make_tab_widget(ts)

        tab_name = os.path.basename(path)

        if insert_at_front:
            # Insert immediately after the Home tab (index 0)
            idx = self._tabs.insertTab(1, container, "📄 " + tab_name)
        else:
            idx = self._tabs.addTab(container, tab_name)

        self._tabs.setCurrentIndex(idx)

        # Fingerprint + restore markers
        pdf_id, partial_hash = _pdf_fingerprint(path)
        saved = db_load_markers(path)
        if saved:
            self._load_markers_from_list(ts, saved)
            db_upsert_session(path, pdf_id, partial_hash, len(saved))
            self.statusBar().showMessage(
                f"Loaded: {tab_name}  —  "
                f"{len(saved)} marker{'s' if len(saved)!=1 else ''} restored")
            return

        matches = db_find_matching_sessions(path, pdf_id, partial_hash)
        if matches:
            dlg = SessionMatchDialog(path, matches, self)
            if dlg.exec() == QDialog.DialogCode.Accepted and dlg.chosen_match:
                n = self._apply_session_match(
                    ts, dlg.chosen_match["old_path"], path, pdf_id, partial_hash)
                self.statusBar().showMessage(
                    f"Loaded: {tab_name}  —  "
                    f"{n} marker{'s' if n!=1 else ''} migrated from previous session")
                return

        db_upsert_session(path, pdf_id, partial_hash, 0)
        self.statusBar().showMessage(f"Loaded: {tab_name}")

    def close_current_tab(self):
        idx = self._tabs.currentIndex()
        if idx > 0:
            self._on_tab_close_requested(idx)

    def _toggle_project_panel(self, checked: bool):
        if checked:
            self._project_dock.show()
            self._project_panel.refresh()
        else:
            self._project_dock.hide()

    # ── Marker loading helpers ────────────────────────────────────────────
    def _load_markers_from_list(self, ts: TabState, markers: list[dict]):
        for m in markers:
            ts.io_list.append(m)
            ts.pdf_view.add_marker()

    def _apply_session_match(self, ts: TabState, old_path: str,
                              new_path: str, pdf_id, partial_hash) -> int:
        n = db_migrate_session(old_path, new_path, pdf_id, partial_hash)
        saved = db_load_markers(new_path)
        self._load_markers_from_list(ts, saved)
        return n

    # ── Register markers ──────────────────────────────────────────────────
    def _register_io(self, ts: TabState, io_type: str, content_point,
                     page: int, pdf_x: float, pdf_y: float,
                     comment: str, count: int = 1, description: str = "",
                     signal_type: str = "", signal_type_comment: str = ""):
        self._push_undo(ts)
        ts.io_list.append({
            "type":                 io_type,
            "comment":              comment,
            "page":                 page,
            "pdf_x":                pdf_x,
            "pdf_y":                pdf_y,
            "count":                max(1, int(count)),
            "description":          description,
            "signal_type":          signal_type,
            "signal_type_comment":  signal_type_comment,
        })
        ts.pdf_view.add_marker()
        self._on_markers_changed(ts)

        # ── Register markers ──────────────────────────────────────────────────
    def _register_io(self, ts: TabState, io_type: str, content_point,
                     page: int, pdf_x: float, pdf_y: float,
                     comment: str, count: int = 1, description: str = "",
                     signal_type: str = "", signal_type_comment: str = ""):
        self._push_undo(ts)
        ts.io_list.append({
            "type":                 io_type,
            "comment":              comment,
            "page":                 page,
            "pdf_x":                pdf_x,
            "pdf_y":                pdf_y,
            "count":                max(1, int(count)),
            "description":          description,
            "signal_type":          signal_type,
            "signal_type_comment":  signal_type_comment,
        })
        ts.pdf_view.add_marker()
        self._on_markers_changed(ts)

    def _register_text(self, ts: TabState, text: str,
                        page: int, pdf_x: float, pdf_y: float):
        self._push_undo(ts)
        ts.io_list.append({"kind": "text", "type": text, "comment": "",
                            "page": page, "pdf_x": pdf_x, "pdf_y": pdf_y})
        ts.pdf_view.add_marker()
        self._on_markers_changed(ts)
        
    def _register_composition(self, ts: TabState,
                         composition_id: int, page: int,
                         pdf_x: float, pdf_y: float,
                         composition: dict,
                         tag_parts: dict):
        """
        When a signal composition is placed, create ONE marker entry.
        """
        self._push_undo(ts)
        
        # User-level multiplier (how many times the whole composition is placed).
        # CompositionPlacementDialog does not set this, so it defaults to 1.
        multiplier = int(tag_parts.get("count", 1) or 1)
        
        # Build the composition text using each signal's configured count.
        signal_counts = {}
        for sig in composition.get("signals", []):
            sig_type = sig.get("signal_type", "")
            if sig_type:
                sig_count = int(sig.get("count", 1) or 1)
                signal_counts[sig_type] = signal_counts.get(sig_type, 0) + sig_count * multiplier
        
        # Build display text
        parts = []
        for sig_type in sorted(signal_counts.keys()):
            sig_count = signal_counts[sig_type]
            parts.append(f"{sig_count}{sig_type}")
        display_text = " ".join(parts) if parts else composition["title"]
        
        
        marker = {
            "type": display_text,
            "composition_id": composition_id,
            "is_composition": True,
            "tag_parts": tag_parts,
            "count": multiplier,
            "page": page,
            "pdf_x": pdf_x,
            "pdf_y": pdf_y,
            "kind": "marker",
        }
        
        ts.io_list.append(marker)
        ts.pdf_view.add_marker()
        self._on_markers_changed(ts)
        
    def _register_text(self, ts: TabState, text: str,
                        page: int, pdf_x: float, pdf_y: float):
        self._push_undo(ts)
        ts.io_list.append({"kind": "text", "type": text, "comment": "",
                            "page": page, "pdf_x": pdf_x, "pdf_y": pdf_y})
        ts.pdf_view.add_marker()
        self._on_markers_changed(ts)
        
    def _register_complex_object(self, ts: TabState,
                             complex_object_id: int, base_tag: str,
                             page: int, pdf_x: float, pdf_y: float,
                             complex_obj_config: dict,
                             tag_parts: dict = None):
        """
        When a complex object is placed, create ONE marker entry.
        
        Args:
            complex_object_id: ID from complex_objects table
            base_tag: Kept for backward compatibility (not used in new version)
            tag_parts: Dict with {prefix, middle_fields, suffix}
            complex_obj_config: The full complex object config dict with signals
        """
        self._push_undo(ts)
        
        if tag_parts is None:
            tag_parts = {"prefix": "", "middle_fields": [], "suffix": ""}
        
        ts.io_list.append({
            "type": complex_obj_config["label"],      # e.g., "On-Off Valve"
            "tag_parts": tag_parts,                   # {prefix, middle_fields, suffix}
            "complex_object_id": complex_object_id,   # Reference to config
            "is_complex": True,                       # Flag for complex marker
            "page": page,
            "pdf_x": pdf_x,
            "pdf_y": pdf_y,
            "kind": "marker",
        })
        ts.pdf_view.add_marker()
        self._on_markers_changed(ts)
    
    # ── Undo / redo ───────────────────────────────────────────────────────
    def _push_undo(self, ts: TabState):
        ts.undo_stack.append(copy.deepcopy(ts.io_list))
        if len(ts.undo_stack) > 50:
            ts.undo_stack.pop(0)
        ts.redo_stack.clear()
        self._refresh_undo_actions(ts)

    def _refresh_undo_actions(self, ts: TabState):
        u = len(ts.undo_stack)
        r = len(ts.redo_stack)
        self.act_undo.setEnabled(bool(u))
        self.act_redo.setEnabled(bool(r))
        self.act_undo.setText(f"↩\nUndo ({u})" if u else "↩\nUndo")
        self.act_redo.setText(f"↪\nRedo ({r})" if r else "↪\nRedo")

    def _restore_snapshot(self, ts: TabState, snapshot: list):
        ts.io_list.clear()
        ts.io_list.extend(snapshot)
        ts.pdf_view._overlay._selected_marker = None
        ts.pdf_view._overlay.update()
        self._on_markers_changed(ts)

    def undo(self):
        ts = self._current_tab()
        if not ts or not ts.undo_stack:
            return
        ts.redo_stack.append(copy.deepcopy(ts.io_list))
        self._restore_snapshot(ts, ts.undo_stack.pop())
        self._refresh_undo_actions(ts)
        self.statusBar().showMessage("Undo")

    def redo(self):
        ts = self._current_tab()
        if not ts or not ts.redo_stack:
            return
        ts.undo_stack.append(copy.deepcopy(ts.io_list))
        self._restore_snapshot(ts, ts.redo_stack.pop())
        self._refresh_undo_actions(ts)
        self.statusBar().showMessage("Redo")

    # ── Markers changed ───────────────────────────────────────────────────
    def _on_overlay_changed(self, ts: TabState):
        self._on_markers_changed(ts)

    def _on_markers_changed(self, ts: TabState):
        counts = {}
        for m in ts.io_list:
            counts[m["type"]] = counts.get(m["type"], 0) + 1
        count_str = "  |  ".join(f"{t}: {n}" for t, n in counts.items())
        self.statusBar().showMessage(
            f"Total: {len(ts.io_list)}   [{count_str}]")
        if ts.pdf_path:
            db_save_markers(ts.pdf_path, ts.io_list)
            try:
                from datetime import datetime
                with _db_connect() as con:
                    _ensure_sessions_table(con)
                    con.execute("""
                        INSERT INTO pdf_sessions
                            (pdf_path, pdf_filename, pdf_id, partial_hash,
                             last_opened, marker_count)
                        VALUES (?, ?, NULL, '', ?, ?)
                        ON CONFLICT(pdf_path) DO UPDATE SET
                            marker_count=excluded.marker_count,
                            last_opened=excluded.last_opened
                    """, (ts.pdf_path, os.path.basename(ts.pdf_path),
                          datetime.now().strftime("%Y-%m-%d %H:%M"),
                          len(ts.io_list)))
                    con.commit()
            except Exception:
                pass
        # Update tab title with marker count
        idx = self._tabs.indexOf(ts.container)
        if idx >= 0:
            self._tabs.setTabText(idx,
                f"{os.path.basename(ts.pdf_path)}  [{len(ts.io_list)}]")

    # ── Zoom ──────────────────────────────────────────────────────────────
    def _do_zoom_in(self, ts: TabState = None):
        ts = ts or self._current_tab()
        if ts:
            self._apply_zoom(ts, ZOOM_STEP)

    def _do_zoom_out(self, ts: TabState = None):
        ts = ts or self._current_tab()
        if ts:
            self._apply_zoom(ts, 1.0 / ZOOM_STEP)

    def _apply_zoom(self, ts: TabState, factor: float):
        ts.pdf_view.apply_zoom(factor)
        self._update_zoom_label(ts)

    def _apply_zoom_reset(self):
        ts = self._current_tab()
        if not ts:
            return
        current = ts.pdf_view._zoom
        if current != 1.0:
            ts.pdf_view.apply_zoom(1.0 / current)
        self._update_zoom_label(ts)

    def _update_zoom_label(self, ts: TabState):
        if ts.zoom_pct_lbl:
            ts.zoom_pct_lbl.setText(f"{int(ts.pdf_view._zoom * 100)}%")
        self._reposition_zoom_panel(ts)

    # ── Page navigation ───────────────────────────────────────────────────
    def _go_to_page(self, page: int):
        ts = self._current_tab()
        if not ts:
            return
        total = ts.pdf_document.pageCount()
        page = max(0, min(total - 1, page))
        ts.pdf_view.pageNavigator().jump(page, QPointF(), 0)
        self._update_page_label(ts)

    def _update_page_label(self, ts: TabState):
        current = ts.pdf_view.pageNavigator().currentPage()
        total   = ts.pdf_document.pageCount()
        self.page_label.setText(f"Page: {current + 1} / {total}")

    # ── Input mode ────────────────────────────────────────────────────────
    def _set_input_mode(self, mode: str):
        # Apply to all open tabs
        for i in range(1, self._tabs.count()):
            ts = self._tab_state_at(i)
            if ts:
                ts.pdf_view._input_mode = mode
        if mode == "marker":
            self.act_mode_marker.setChecked(True)
            self.act_mode_text.setChecked(False)
            self._mode_label.setText("📍 Marker Mode")
            self._mode_label.setStyleSheet(
                "padding:2px 10px; border-radius:4px;"
                "background:#1A3A5C; color:#7EC8F0; font-weight:bold; font-size: 8pt;")
        else:
            self.act_mode_text.setChecked(True)
            self.act_mode_marker.setChecked(False)
            self._mode_label.setText("💬 Text Mode")
            self._mode_label.setStyleSheet(
                "padding:2px 10px; border-radius:4px;"
                "background:#3B1F4E; color:#CE9CF5; font-weight:bold; font-size: 8pt;")

    # ── keyPressEvent ─────────────────────────────────────────────────────
    def keyPressEvent(self, event):
        ts = self._current_tab()
        if not ts:
            super().keyPressEvent(event)
            return
        overlay = ts.pdf_view._overlay
        key = event.key()
        if key in (Qt.Key.Key_Delete, Qt.Key.Key_Backspace):
            if overlay._selected_marker:
                overlay.keyPressEvent(event)
                return
        if key in (Qt.Key.Key_Left, Qt.Key.Key_Right,
                   Qt.Key.Key_Up,   Qt.Key.Key_Down):
            if overlay._selected_marker:
                overlay.keyPressEvent(event)
                return
        super().keyPressEvent(event)

    # ── resizeEvent ───────────────────────────────────────────────────────
    def resizeEvent(self, event):
        super().resizeEvent(event)
        ts = self._current_tab()
        if ts:
            self._reposition_zoom_panel(ts)

    # ── Configure dialogs ─────────────────────────────────────────────────
    def open_config(self):
        dlg = SignalTypeConfigDialog(self._signal_types, parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            new_groups = dlg.signal_types()
            if not new_groups:
                QMessageBox.warning(self, "Configuration",
                                    "You must keep at least one group.")
                return
            self._signal_types[:] = new_groups
            db_save_signal_types(new_groups)
            def _count_leaves(nodes):
                n = 0
                for nd in nodes:
                    n += _count_leaves(nd.get("children") or []) if nd.get("children") else 1
                return n
            self.statusBar().showMessage(
                f"Signal types updated: {len(new_groups)} groups, "
                f"{_count_leaves(new_groups)} signal types")
    
    def open_signal_compositions_config(self):
        """Open signal typicals config for the current project."""
        ts = self._current_tab()
        if not ts:
            QMessageBox.information(
                self, "No PDF Open",
                "Please open a PDF file first to configure signal typicals.\n\n"
                "Signal typicals are managed per project.")
            return
        
        # Check if the PDF belongs to a project
        projects = db_find_pdf_in_projects(ts.pdf_path)
        
        if not projects:
            QMessageBox.information(
                self, "PDF Not in Project",
                f"The file '{os.path.basename(ts.pdf_path)}' is not registered in any project.\n\n"
                "To configure signal typicals:\n"
                "1. Create or open a project from the Project Panel\n"
                "2. Add this PDF to the project\n"
                "3. Then access signal typical configuration")
            return
        
        if len(projects) == 1:
            project_id = projects[0]["project_id"]
            project_name = projects[0]["project_name"]
        else:
            # Multiple projects - let user choose
            names = [p["project_name"] for p in projects]
            choice, ok = QInputDialog.getItem(
                self, "Select Project",
                "This PDF belongs to multiple projects.\n"
                "Configure typicals for:",
                names, 0, False)
            if not ok:
                return
            project_id = projects[names.index(choice)]["project_id"]
            project_name = choice
        
        # Get or create the project owner
        owner_id = db_get_or_create_project_owner(project_id)
        
        comp_dlg = SignalCompositionConfigDialog(owner_id, project_name, parent=self)
        comp_dlg.exec()
        # Refresh all open tab overlays so marker labels reflect any changes
        self._refresh_composition_markers()
        if comp_dlg.result() == QDialog.DialogCode.Accepted:
            self.statusBar().showMessage(
                f"Signal typicals configured for '{project_name}'")
        
    def _refresh_composition_markers(self) -> None:
        """Re-calculate the display label for every composition marker in all
        open tabs and repaint their overlays.  Called after the signal typical
        config dialog closes so that label changes are visible immediately."""
        for idx in range(self._tabs.count()):
            ts = self._tab_state_at(idx)
            if ts is None:
                continue
            changed = False
            for m in ts.io_list:
                if not m.get("is_composition") or not m.get("composition_id"):
                    continue
                comp = db_load_signal_composition(m["composition_id"])
                if not comp:
                    continue
                tag_parts  = m.get("tag_parts") or {}
                new_count  = int(tag_parts.get("count", m.get("count", 1)) or 1)
                signal_counts: dict[str, int] = {}
                for sig in comp.get("signals", []):
                    sig_type = sig.get("signal_type", "")
                    if sig_type:
                        sig_count = int(sig.get("count", 1) or 1)
                        signal_counts[sig_type] = (
                            signal_counts.get(sig_type, 0) + sig_count * new_count)
                parts = [f"{c}{t}" for t, c in sorted(signal_counts.items())]
                new_label = " ".join(parts) if parts else comp["title"]
                if m.get("type") != new_label:
                    m["type"] = new_label
                    changed = True
            if changed:
                ts.pdf_view._overlay.update()

    def open_export_config(self):
        current = db_load_export_columns()
        dlg = ExportColumnConfigDialog(current, parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            cols = dlg.export_columns()
            db_save_export_columns(cols)
            self.statusBar().showMessage(f"Export columns saved: {len(cols)} column(s)")

    # ── Signal-type tree  →  Excel ────────────────────────────────────────
    def export_signal_types_xlsx(self) -> None:
        """Export the current signal-type tree to an Excel template file."""
        if not self._signal_types:
            QMessageBox.warning(self, "Nothing to export",
                                "No signal types configured yet.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Signal Types to Excel",
            "signal_types.xlsx", "Excel Files (*.xlsx)")
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        try:
            export_signal_types_to_xlsx(path, self._signal_types)
            self.statusBar().showMessage(f"Signal types exported: {path}")
            QMessageBox.information(
                self, "Export successful",
                f"Signal types exported to:\n{path}\n\n"
                f"Edit the file, save it as .xlsx or as\n"
                f"'XML Spreadsheet 2003 (*.xml)' in Excel,\n"
                f"then import it back via\n"
                f"Edit → Import Signal Types from XML/Excel…")
        except Exception as exc:
            QMessageBox.critical(self, "Export failed", str(exc))

    # ── Signal-type tree  ←  XML / xlsx ──────────────────────────────────
    def import_signal_types_file(self) -> None:
        """Import a signal-type tree from an .xlsx or XML Spreadsheet file."""
        path, _ = QFileDialog.getOpenFileName(
            self, "Import Signal Types",
            "", "Excel / XML Files (*.xlsx *.xml);;All Files (*)")
        if not path:
            return
        try:
            ext = os.path.splitext(path)[1].lower()
            if ext == ".xml":
                new_tree = import_signal_types_from_xml(path)
            elif ext == ".xlsx":
                new_tree = import_signal_types_from_xlsx(path)
            else:
                QMessageBox.warning(self, "Unsupported format",
                                    "Please select an .xlsx or .xml file.")
                return
        except Exception as exc:
            QMessageBox.critical(self, "Import failed",
                                 f"Could not parse the file:\n\n{exc}")
            return

        if not new_tree:
            QMessageBox.warning(self, "Empty",
                                "The file contained no signal-type data.")
            return

        def _count_leaves(nodes):
            n = 0
            for nd in nodes:
                n += _count_leaves(nd.get("children") or []) if nd.get("children") else 1
            return n

        ans = QMessageBox.question(
            self, "Replace signal types?",
            f"Import {len(new_tree)} group(s) / "
            f"{_count_leaves(new_tree)} signal type(s) from:\n"
            f"{os.path.basename(path)}\n\n"
            f"This will replace the current signal-type configuration.\n"
            f"Existing markers on open PDFs are not affected.\n\n"
            f"Continue?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if ans != QMessageBox.StandardButton.Yes:
            return

        self._signal_types[:] = new_tree
        db_save_signal_types(new_tree)
        self.statusBar().showMessage(
            f"Signal types imported: {len(new_tree)} group(s), "
            f"{_count_leaves(new_tree)} signal type(s) — from {os.path.basename(path)}")

    # ── Manual session link ───────────────────────────────────────────────
    def open_manual_link(self):
        ts = self._current_tab()
        if not ts:
            QMessageBox.information(self, "No file open", "Please open a PDF first.")
            return
        sessions = [s for s in db_all_sessions() if s["pdf_path"] != ts.pdf_path]
        if not sessions:
            QMessageBox.information(self, "No sessions",
                                    "No previous sessions found in the database.")
            return
        dlg = ManualLinkDialog(ts.pdf_path, sessions, self)
        if dlg.exec() != QDialog.DialogCode.Accepted or not dlg.chosen_session:
            return
        old_path = dlg.chosen_session["pdf_path"]
        ans = QMessageBox.question(
            self, "Replace current markers?",
            f"Replace markers on <b>{os.path.basename(ts.pdf_path)}</b> with "
            f"{dlg.chosen_session['marker_count']} markers from "
            f"<b>{dlg.chosen_session['pdf_filename']}</b>?\n\nContinue?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if ans != QMessageBox.StandardButton.Yes:
            return
        self._push_undo(ts)
        ts.io_list.clear()
        ts.pdf_view.clear_markers()
        pdf_id, partial_hash = _pdf_fingerprint(ts.pdf_path)
        n = self._apply_session_match(ts, old_path, ts.pdf_path, pdf_id, partial_hash)
        self.statusBar().showMessage(
            f"{n} marker{'s' if n!=1 else ''} linked from "
            f"'{dlg.chosen_session['pdf_filename']}'")
        
    def show_marker_info(self, marker: dict):
        """Show information dialog for a marker."""
        dlg = MarkerInfoDialog(marker, self)
        dlg.exec()

    # ── Excel export ──────────────────────────────────────────────────────
    def export_xlsx(self):
        ts = self._current_tab()
        if not ts:
            return
        if not ts.io_list:
            self.statusBar().showMessage("No markers to export.")
            return
        meta = db_load_export_metadata(ts.pdf_path)
        meta_dlg = ExportMetadataDialog(meta, parent=self)
        if meta_dlg.exec() != QDialog.DialogCode.Accepted:
            return
        meta = meta_dlg.metadata
        db_save_export_metadata(ts.pdf_path, meta)
        col_config = db_load_export_columns()
        pdf_dir  = os.path.dirname(ts.pdf_path)
        pdf_stem = os.path.splitext(os.path.basename(ts.pdf_path))[0]
        out_path = os.path.join(pdf_dir, "comments", f"{pdf_stem}_comment.xlsx")
        path, _ = QFileDialog.getSaveFileName(
            self, "Export to Excel", out_path, "Excel Files (*.xlsx)")
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        os.makedirs(os.path.dirname(path), exist_ok=True)
        try:
            export_to_excel(path, ts.pdf_path, ts.io_list, meta, col_config)
            self.statusBar().showMessage(f"Excel exported: {path}")
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))

    # ── Project IO List export ────────────────────────────────────────────
    def export_project_io_list(self):
        """
        Export IO signal markers to Excel.
        The user first picks a project, then chooses:
          • Entire project  — all PDFs, one sheet per drawing + summary
          • A single PDF    — just that drawing
        Existing markers are found via pdf path even if saved before the
        project_file_id column was added (auto-repaired on export).
        """
        projects = db_load_projects()
        if not projects:
            QMessageBox.information(
                self, "No projects",
                "Create a project and add PDFs to it before exporting an IO list.")
            return

        # ── Step 1: pick project ─────────────────────────────────────────
        if len(projects) == 1:
            project = projects[0]
        else:
            ts = self._current_tab()
            pre_select = 0
            if ts:
                for i, p in enumerate(projects):
                    files = db_load_project_files(p["id"])
                    if any(os.path.normpath(f["file_path"]) ==
                           os.path.normpath(ts.pdf_path) for f in files):
                        pre_select = i
                        break
            names = [p["name"] + (f"  ({p['number']})" if p["number"] else "")
                     for p in projects]
            choice, ok = QInputDialog.getItem(
                self, "Export Project IO List",
                "Select project:", names, pre_select, False)
            if not ok:
                return
            project = projects[names.index(choice)]

        project_id = project["id"]
        project_files = db_load_project_files(project_id)

        if not project_files:
            QMessageBox.information(
                self, "No files",
                f"Project \"{project['name']}\" has no PDF files registered.")
            return

        # ── Step 2: entire project or single PDF? ────────────────────────
        scope_choices = (
            ["Entire project — all PDFs"] +
            [f["file_name"] for f in project_files]
        )
        scope, ok = QInputDialog.getItem(
            self, "Export scope",
            f"Export IO list for project \"{project['name']}\":",
            scope_choices, 0, False)
        if not ok:
            return

        if scope == scope_choices[0]:
            # Entire project
            selected_file_path = None
        else:
            # Single PDF — find the matching file_path
            chosen_file = next(
                f for f in project_files if f["file_name"] == scope)
            selected_file_path = chosen_file["file_path"]

        # ── Step 3: load markers via path-based join ─────────────────────
        markers = db_load_project_markers(project_id, selected_file_path)

        if not markers:
            # Build a diagnostic so the user knows what the DB actually has
            diag_lines = []
            for pf in project_files:
                if selected_file_path and pf["file_path"] != selected_file_path:
                    continue
                with _db_connect() as con:
                    total = con.execute(
                        "SELECT COUNT(*) FROM markers WHERE pdf = ?",
                        (pf["file_path"],)).fetchone()[0]
                    io_count = con.execute(
                        "SELECT COUNT(*) FROM markers WHERE pdf = ? AND kind != 'text'",
                        (pf["file_path"],)).fetchone()[0]
                diag_lines.append(
                    f"• {pf['file_name']}: {total} total markers, "
                    f"{io_count} IO markers in DB")

            diag = "\n".join(diag_lines) if diag_lines else "(no files checked)"
            QMessageBox.information(
                self, "No IO markers found",
                f"No IO signal markers were found for the selected scope.\n\n"
                f"Database check:\n{diag}\n\n"
                f"If the counts above show 0, place a marker on the PDF and\n"
                f"the system will save it automatically.")
            return

        # ── Step 4: pick save path ────────────────────────────────────────
        if selected_file_path:
            safe_stem = os.path.splitext(
                os.path.basename(selected_file_path))[0]
            default_name = f"{safe_stem}_IO_List.xlsx"
        else:
            safe_name = "".join(
                c for c in project["name"] if c.isalnum() or c in " _-")
            default_name = f"{safe_name}_IO_List.xlsx"

        out_path = os.path.join(os.path.dirname(DB_PATH), default_name)
        path, _ = QFileDialog.getSaveFileName(
            self, "Export IO List", out_path, "Excel Files (*.xlsx)")
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        os.makedirs(os.path.dirname(path), exist_ok=True)
        try:
            export_project_io_list(path, project_id, selected_file_path)
            n = len(markers)
            drawings = len(set(m["file_name"] for m in markers))
            self.statusBar().showMessage(
                f"IO List exported: {path}  ({n} marker(s), {drawings} drawing(s))")
            QMessageBox.information(
                self, "Export complete",
                f"IO List exported for <b>{project['name']}</b>.<br>"
                f"{n} IO marker(s) across {drawings} drawing(s).<br><br>"
                f"{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))

    def _export_project_io_by_id(self, project_id: int):
        """Export IO list for a specific project_id (called from project panel context menu)."""
        projects = db_load_projects()
        project = next((p for p in projects if p["id"] == project_id), None)
        if not project:
            QMessageBox.warning(self, "Project Not Found",
                                "Could not find the selected project.")
            return
        safe_name = "".join(
            c for c in project["name"] if c.isalnum() or c in " _-")
        default_name = f"{safe_name}_IO_List.xlsx"
        out_path = os.path.join(os.path.dirname(DB_PATH), default_name)
        path, _ = QFileDialog.getSaveFileName(
            self, "Export IO List", out_path, "Excel Files (*.xlsx)")
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        os.makedirs(os.path.dirname(path), exist_ok=True)
        try:
            export_project_io_list(path, project_id, None)
            markers = db_load_project_markers(project_id, None)
            n = len(markers)
            drawings = len(set(m["file_name"] for m in markers))
            self.statusBar().showMessage(
                f"IO List exported: {path}  ({n} marker(s), {drawings} drawing(s))")
            QMessageBox.information(
                self, "Export complete",
                f"IO List exported for <b>{project['name']}</b>.<br>"
                f"{n} IO marker(s) across {drawings} drawing(s).<br><br>"
                f"{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))
    def save_fdf(self):
        ts = self._current_tab()
        if not ts or not ts.io_list:
            self.statusBar().showMessage("No markers to save.")
            return
        pdf_dir  = os.path.dirname(ts.pdf_path)
        pdf_stem = os.path.splitext(os.path.basename(ts.pdf_path))[0]
        default_path = os.path.join(pdf_dir, "comments", f"{pdf_stem}_comment.fdf")
        path, _ = QFileDialog.getSaveFileName(
            self, "Save FDF", default_path, "FDF Files (*.fdf)")
        if not path:
            return
        if not path.lower().endswith(".fdf"):
            path += ".fdf"
        os.makedirs(os.path.dirname(path), exist_ok=True)
        _write_fdf(path, ts.pdf_path, ts.io_list)
        self.statusBar().showMessage(f"FDF saved: {path}")

    # ── PDF export ────────────────────────────────────────────────────────
    def save_pdf(self):
        ts = self._current_tab()
        if not ts or not ts.io_list:
            self.statusBar().showMessage("No markers to save.")
            return
        pdf_dir  = os.path.dirname(ts.pdf_path)
        pdf_stem = os.path.splitext(os.path.basename(ts.pdf_path))[0]
        default_path = os.path.join(pdf_dir, "comments", f"{pdf_stem}_comment.pdf")
        path, _ = QFileDialog.getSaveFileName(
            self, "Save PDF with comments", default_path, "PDF Files (*.pdf)")
        if not path:
            return
        if not path.lower().endswith(".pdf"):
            path += ".pdf"
        os.makedirs(os.path.dirname(path), exist_ok=True)
        try:
            reader = PdfReader(ts.pdf_path)
            writer = PdfWriter()
            writer.append(reader)
            for m in ts.io_list:
                page_idx = int(m["page"])
                if page_idx >= len(writer.pages):
                    continue
                pdf_x = float(m["pdf_x"])
                pdf_y = float(m["pdf_y"])
                kind  = m.get("kind", "marker")

                if kind == "text":
                    # ── FreeText / Typewriter annotation ─────────────────
                    # Renders as direct text on the page with no border,
                    # no background, and Subj(Typewriter) recognised by
                    # Adobe Acrobat, Foxit, PDF-XChange, etc.
                    text = m.get("type", "")
                    font_size_pt = 10
                    # Estimate rect width/height from text content
                    lines     = text.split("\n") if text else [""]
                    char_w    = font_size_pt * 0.55   # approx pts per char
                    line_h    = font_size_pt * 1.4    # line height in pts
                    rect_w    = max(40.0, max(len(l) for l in lines) * char_w)
                    rect_h    = max(line_h, len(lines) * line_h)
                    rect = (
                        pdf_x,
                        pdf_y - rect_h,        # PDF Y grows upward
                        pdf_x + rect_w,
                        pdf_y,
                    )
                    annot = PdfFreeTextAnnotation(
                        text=text,
                        rect=rect,
                        font="Helvetica",
                        font_size=f"{font_size_pt}pt",
                        font_color="000000",
                        border_color=None,       # no border
                        background_color=None,   # transparent background
                    )
                    # Flags: 4 = Print  (annotation will print)
                    annot[NameObject("/F")] = NumberObject(4)
                    # Intent = FreeTextTypewriter  →  Subj(Typewriter) in viewers
                    annot[NameObject("/IT")] = NameObject("/FreeTextTypewriter")
                    # Subject tag recognised by Acrobat/Foxit as Typewriter
                    annot[NameObject("/Subj")] = NameObject("/Typewriter")
                    # Quadding: 0 = left-aligned
                    annot[NameObject("/Q")] = NumberObject(0)
                    writer.add_annotation(page_number=page_idx, annotation=annot)

                else:
                    # ── Sticky-note Text annotation for IO signal markers ─
                    count   = int(m.get("count", 1))
                    label   = f"{count}{m['type']}" if count > 1 else m["type"]
                    comment = m.get("comment", "")
                    desc    = m.get("description", "")
                    parts   = [label]
                    if comment:
                        parts.append(comment)
                    if desc:
                        parts.append(desc)
                    contents = "\n".join(parts)
                    HALF = 12
                    annot = PdfTextAnnotation(
                        rect=(pdf_x - HALF, pdf_y - HALF,
                              pdf_x + HALF, pdf_y + HALF),
                        text=contents,
                        open=False)
                    writer.add_annotation(page_number=page_idx, annotation=annot)

            with open(path, "wb") as f:
                writer.write(f)
            self.statusBar().showMessage(f"PDF saved: {path}")
        except Exception as e:
            QMessageBox.critical(self, "Save PDF failed", str(e))   

def _test_signal_compositions():
    """Test the new signal composition functions."""
    print("Testing Signal Composition System...\n")
    
    # 1. Create Default compositions
    print("1. Creating Default compositions...")
    
    comp1_id = db_save_signal_composition(
        title="On-Off Valve",
        description="2/2 Solenoid Valve",
        signals=[
            {"signal_name": "XS", "signal_type": "HDO", "signal_description": "Valve position"},
            {"signal_name": "ZSH", "signal_type": "HDI", "signal_description": "Open limit"},
            {"signal_name": "ZSL", "signal_type": "HDI", "signal_description": "Closed limit"}
        ]
    )
    print(f"   ✓ Created 'On-Off Valve' (ID: {comp1_id})")
    
    comp2_id = db_save_signal_composition(
        title="Check Valve",
        description="Inline Check",
        signals=[
            {"signal_name": "P_in", "signal_type": "AI", "signal_description": "Inlet Pressure"},
            {"signal_name": "P_out", "signal_type": "AI", "signal_description": "Outlet Pressure"}
        ]
    )
    print(f"   ✓ Created 'Check Valve' (ID: {comp2_id})")
    
    # 2. Assign to Default owner
    print("\n2. Assigning to Default owner...")
    default_owner_id = db_ensure_default_owner()
    print(f"   ✓ Default owner ID: {default_owner_id}")
    
    db_assign_composition_to_owner(comp1_id, default_owner_id, 1)
    db_assign_composition_to_owner(comp2_id, default_owner_id, 2)
    print(f"   ✓ Assigned both compositions to Default owner")
    
    # 3. Create a test project
    print("\n3. Creating test project...")
    test_project_id = db_create_project("Test Project", "TP-001", "For testing")
    print(f"   ✓ Created project ID: {test_project_id}")
    
    # 4. Get project owner and assign composition
    print("\n4. Assigning composition to project...")
    project_owner_id = db_get_or_create_project_owner(test_project_id)
    print(f"   ✓ Project owner ID: {project_owner_id}")
    
    comp3_id = db_save_signal_composition(
        title="Flow Meter",
        description="Project-specific composition",
        signals=[
            {"signal_name": "FLOW", "signal_type": "AI", "signal_description": "Flow Rate"}
        ]
    )
    db_assign_composition_to_owner(comp3_id, project_owner_id, 1)
    print(f"   ✓ Created and assigned 'Flow Meter' to project")
    
    # 5. Load all compositions for project
    print("\n5. Loading all compositions for project...")
    all_comps = db_load_all_compositions_for_project(test_project_id)
    
    for owner_name, compositions in all_comps.items():
        print(f"\n   {owner_name}:")
        for comp in compositions:
            signals_str = " ".join([s["signal_type"] for s in comp["signals"]])
            print(f"     - {comp['title']}: {signals_str}")
    
    # 6. Cleanup
    print("\n6. Cleaning up...")
    db_delete_signal_composition(comp1_id)
    db_delete_signal_composition(comp2_id)
    db_delete_signal_composition(comp3_id)
    db_delete_project(test_project_id)
    print("   ✓ Cleanup complete")
    
    print("\n✓ All tests passed!")
    
if __name__ == "__main__":
    # Initialize database and default compositions
    _db_connect()
    
    app = QApplication(sys.argv)
    # Set the application-wide font with an explicit point size before any
    # widgets are created.
    from PySide6.QtGui import QFont as _AppFont
    app.setFont(_AppFont("Montserrat", 9))
    viewer = PDFViewer()
    viewer.show()
    sys.exit(app.exec())
